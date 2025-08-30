#!/usr/bin/env python3
"""
Improved Fleet Management System
Î’ÎµÎ»Ï„Î¹Ï‰Î¼Î­Î½Î¿ Î£ÏÏƒÏ„Î·Î¼Î± Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ·Ï‚ Î£Ï„ÏŒÎ»Î¿Ï…

Key improvements:
- Better code organization with separate modules
- Enhanced UI with modern components
- Improved error handling and validation
- Better user experience with tooltips and status updates
- Enhanced search and filtering capabilities
- Comprehensive logging and audit trail
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
from datetime import datetime
import os
import sys
import logging
import csv

# Import custom modules
from database import DatabaseManager
from ui_components import (
    ModernButton, ModernFrame, ModernEntry, ModernHeader, SearchableCombobox, 
    StatusBar, ProgressDialog, ConfirmDialog, ValidationMixin, TooltipManager
)
from utils import (
    normalize_plate, normalize_name, ensure_dir, validate_date,
    format_currency, format_distance, format_fuel, hash_password,
    export_to_csv, backup_file, DataValidator, log_user_action
)
from config import (
    THEMES, BUTTON_STYLES, FONT_BIG, FONT_NORMAL, FONT_SMALL,
    FONT_TITLE, FONT_SUBTITLE, IMAGES_DIR, TANK_MIN_LEVEL,
    TANK_CAPACITY, DEFAULT_PASSWORD, LOGGING_CONFIG
)

# Configure logging
logging.basicConfig(**LOGGING_CONFIG)

# Optional imports with graceful fallback
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    logging.warning("PIL not available - photo features will be limited")

try:
    from openpyxl import Workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not available - Excel export disabled")

class FleetManagerImproved(ValidationMixin):
    """Improved Fleet Management System with better UX and code organization"""
    
    def __init__(self, root):
        self.root = root
        # Load saved theme preference
        from config import load_user_setting
        self.current_theme = load_user_setting("theme", "light")
        self.db = None
        self.tooltip_manager = TooltipManager()
        
        # Initialize application
        self._setup_window()
        self._setup_database()
        
        # Authenticate user
        if not self._authenticate_user():
            self.root.destroy()
            return
        
        # Build UI
        self._build_main_ui()
        self._setup_event_handlers()
        
        # Initialize data
        self._initialize_data()
        
        log_user_action("Application started")
    
    def _setup_window(self):
        """Setup main window properties with adaptive sizing"""
        self.root.title("ğŸš— Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· Î£Ï„ÏŒÎ»Î¿Ï… - Fleet Management System v2.0")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calculate optimal window size based on screen size
        if screen_width < 1366:  # Small laptop screens
            window_width = int(screen_width * 0.95)
            window_height = int(screen_height * 0.90)
        elif screen_width < 1920:  # Standard laptop/desktop screens
            window_width = int(screen_width * 0.85)
            window_height = int(screen_height * 0.85)
        else:  # Large desktop monitors
            window_width = int(screen_width * 0.75)
            window_height = int(screen_height * 0.75)
        
        # Set minimum and maximum sizes
        min_width = 1000
        min_height = 700
        # Allow window to use full screen size when maximized
        max_width = screen_width
        max_height = screen_height
        
        # Ensure window size is within bounds
        window_width = max(min_width, min(window_width, max_width))
        window_height = max(min_height, min(window_height, max_height))
        
        # Set window size and constraints
        self.root.geometry(f"{window_width}x{window_height}")
        self.root.minsize(min_width, min_height)
        # Remove maxsize restriction to allow full screen maximize
        # self.root.maxsize(max_width, max_height)
        
        # Enable window resizing in all directions
        self.root.resizable(True, True)
        
        # Update font sizes based on screen size
        self._update_font_sizes(screen_width, screen_height)
        
        # Configure background
        self.root.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Center window on screen
        self.root.update_idletasks()
        x = (screen_width // 2) - (window_width // 2)
        y = (screen_height // 2) - (window_height // 2)
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Store window dimensions for later use
        self.window_width = window_width
        self.window_height = window_height
        
        # Additional window configuration for better behavior
        self.root.state('normal')  # Ensure window starts in normal state
        
        # Configure window attributes for better maximize behavior
        self.root.attributes('-topmost', False)  # Don't keep always on top
        
        # Set window to start in the center of the primary monitor
        # This ensures proper behavior across multiple monitors
        try:
            # Force window to be visible and properly positioned
            self.root.withdraw()  # Hide window temporarily
            self.root.update_idletasks()  # Process all pending events
            self.root.deiconify()  # Show window again
        except Exception as e:
            logging.warning(f"Error in window positioning: {e}")
        
        # Bind window state events for better maximize handling
        self.root.bind('<Configure>', self._on_window_configure)
        
        # Set window icon if available
        self._set_window_icon()
        
        # Log window setup
        logging.info(f"Window setup: {window_width}x{window_height} on {screen_width}x{screen_height} screen")
        
        # Ensure window gets focus and is brought to front
        self.root.focus_force()
        self.root.lift()
        self.root.attributes('-topmost', True)  # Temporarily bring to front
        self.root.after(100, lambda: self.root.attributes('-topmost', False))  # Remove topmost after 100ms
    
    def _set_window_icon(self):
        """Set application icon"""
        try:
            if getattr(sys, 'frozen', False):
                application_path = sys._MEIPASS  # type: ignore
            else:
                application_path = os.path.dirname(os.path.abspath(__file__))

            icon_path = os.path.join(application_path, "fleet_icon.png")
            if os.path.exists(icon_path):
                icon = tk.PhotoImage(file=icon_path)
                self.root.iconphoto(True, icon)
        except Exception as e:
            logging.warning(f"Could not set window icon: {e}")
    
    def _on_window_configure(self, event):
        """Handle window configuration changes for better resize behavior"""
        # Only handle main window configure events, not child widgets
        if event.widget == self.root:
            # Get current window dimensions
            current_width = self.root.winfo_width()
            current_height = self.root.winfo_height()
            
            # Update stored dimensions if they've changed significantly
            if (abs(current_width - self.window_width) > 10 or 
                abs(current_height - self.window_height) > 10):
                self.window_width = current_width
                self.window_height = current_height
                
                # Optional: Adjust font sizes if window becomes very large or small
                self._adjust_fonts_for_window_size(current_width, current_height)
    
    def _adjust_fonts_for_window_size(self, width, height):
        """Adjust font sizes based on current window size"""
        try:
            # Calculate a scaling factor based on window size
            base_width = 1200
            base_height = 800
            
            width_scale = width / base_width
            height_scale = height / base_height
            scale_factor = min(width_scale, height_scale)
            
            # Don't scale too much
            scale_factor = max(0.8, min(scale_factor, 1.5))
            
            # Update global font sizes (optional - only if you want dynamic scaling)
            # You can uncomment these if you want fonts to scale with window size
            # global FONT_TITLE, FONT_SUBTITLE, FONT_NORMAL, FONT_SMALL
            # base_title_size = 16
            # base_subtitle_size = 12
            # base_normal_size = 10
            # base_small_size = 8
            
            # FONT_TITLE = ("Segoe UI", int(base_title_size * scale_factor), "bold")
            # FONT_SUBTITLE = ("Segoe UI", int(base_subtitle_size * scale_factor), "bold")
            # FONT_NORMAL = ("Segoe UI", int(base_normal_size * scale_factor))
            # FONT_SMALL = ("Segoe UI", int(base_small_size * scale_factor))
            
        except Exception as e:
            logging.warning(f"Error adjusting fonts for window size: {e}")
    
    def _setup_database(self):
        """Initialize database connection"""
        try:
            ensure_dir(IMAGES_DIR)
            self.db = DatabaseManager()
            logging.info("Database initialized successfully")
        except Exception as e:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î± Î’Î¬ÏƒÎ·Ï‚ Î”ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½", 
                               f"Î‘Î´Ï…Î½Î±Î¼Î¯Î± ÏƒÏÎ½Î´ÎµÏƒÎ·Ï‚ Î¼Îµ Ï„Î· Î²Î¬ÏƒÎ· Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½:\n{str(e)}")
            self.root.destroy()
            raise
    
    def _authenticate_user(self):
        """Authenticate user with improved security"""
        max_attempts = 3
        hash_pwd = hash_password(DEFAULT_PASSWORD)
        
        # Create a custom authentication dialog that stays on top
        return self._show_auth_dialog(max_attempts, hash_pwd)
    
    def _show_auth_dialog(self, max_attempts, hash_pwd):
        """Show authentication dialog that stays on top"""
        for attempt in range(max_attempts):
            remaining = max_attempts - attempt
            
            # Create authentication window
            auth_window = tk.Toplevel(self.root)
            auth_window.title("ÎšÏ‰Î´Î¹ÎºÏŒÏ‚ Î ÏÏŒÏƒÎ²Î±ÏƒÎ·Ï‚")
            auth_window.geometry("350x180")
            auth_window.resizable(False, False)
            auth_window.configure(bg='#f0f0f0')
            
            # Center the window on screen
            auth_window.update_idletasks()
            x = (auth_window.winfo_screenwidth() // 2) - (350 // 2)
            y = (auth_window.winfo_screenheight() // 2) - (180 // 2)
            auth_window.geometry(f"350x180+{x}+{y}")
            
            # Make it stay on top and modal
            auth_window.transient(self.root)
            auth_window.grab_set()
            auth_window.attributes('-topmost', True)
            auth_window.focus_force()
            
            # Store the result
            result = {'password': None, 'cancelled': False}
            
            # Create UI elements
            tk.Label(auth_window, text="ğŸ” Î•Î¹ÏƒÎ¬Î³ÎµÏ„Îµ ÎºÏ‰Î´Î¹ÎºÏŒ Ï€ÏÏŒÏƒÎ²Î±ÏƒÎ·Ï‚:", 
                    font=("Segoe UI", 12), bg='#f0f0f0').pack(pady=15)
            
            tk.Label(auth_window, text=f"(Î‘Ï€Î¿Î¼Î­Î½Î¿Ï…Î½ {remaining} Ï€ÏÎ¿ÏƒÏ€Î¬Î¸ÎµÎ¹ÎµÏ‚)", 
                    font=("Segoe UI", 10), bg='#f0f0f0', fg='#666').pack(pady=5)
            
            # Password entry
            password_var = tk.StringVar()
            password_entry = tk.Entry(auth_window, textvariable=password_var, 
                                    show="*", font=("Segoe UI", 12), width=20)
            password_entry.pack(pady=10)
            password_entry.focus_set()
            
            # Button frame
            btn_frame = tk.Frame(auth_window, bg='#f0f0f0')
            btn_frame.pack(pady=15)
            
            def on_ok():
                result['password'] = password_var.get()
                auth_window.destroy()
            
            def on_cancel():
                result['cancelled'] = True
                auth_window.destroy()
            
            def on_enter(event):
                on_ok()
            
            # Bind Enter key to OK
            password_entry.bind('<Return>', on_enter)
            auth_window.bind('<Return>', on_enter)
            
            # Buttons
            tk.Button(btn_frame, text="OK", command=on_ok, 
                     font=("Segoe UI", 10), width=8).pack(side=tk.LEFT, padx=5)
            tk.Button(btn_frame, text="Cancel", command=on_cancel, 
                     font=("Segoe UI", 10), width=8).pack(side=tk.LEFT, padx=5)
            
            # Wait for the dialog to close
            auth_window.wait_window()
            
            # Check result
            if result['cancelled']:
                return False
            
            pwd = result['password']
            if pwd and hash_password(pwd) == hash_pwd:
                log_user_action("User authenticated successfully")
                return True
            
            if attempt < max_attempts - 1:  # Not the last attempt
                # Show error message
                error_window = tk.Toplevel(self.root)
                error_window.title("Î›Î¬Î¸Î¿Ï‚ ÎšÏ‰Î´Î¹ÎºÏŒÏ‚")
                error_window.geometry("300x120")
                error_window.resizable(False, False)
                error_window.configure(bg='#f0f0f0')
                
                # Center error window
                error_window.update_idletasks()
                x = (error_window.winfo_screenwidth() // 2) - (300 // 2)
                y = (error_window.winfo_screenheight() // 2) - (120 // 2)
                error_window.geometry(f"300x120+{x}+{y}")
                
                error_window.transient(self.root)
                error_window.grab_set()
                error_window.attributes('-topmost', True)
                error_window.focus_force()
                
                tk.Label(error_window, text="âŒ Î›Î¬Î¸Î¿Ï‚ ÎºÏ‰Î´Î¹ÎºÏŒÏ‚!", 
                        font=("Segoe UI", 12), bg='#f0f0f0', fg='red').pack(pady=15)
                tk.Label(error_window, text=f"Î‘Ï€Î¿Î¼Î­Î½Î¿Ï…Î½ {remaining - 1} Ï€ÏÎ¿ÏƒÏ€Î¬Î¸ÎµÎ¹ÎµÏ‚", 
                        font=("Segoe UI", 10), bg='#f0f0f0').pack(pady=5)
                
                def close_error():
                    error_window.destroy()
                
                tk.Button(error_window, text="OK", command=close_error, 
                         font=("Segoe UI", 10)).pack(pady=10)
                
                error_window.wait_window()
        
        # Final error - max attempts exceeded
        final_error = tk.Toplevel(self.root)
        final_error.title("Î ÏÏŒÏƒÎ²Î±ÏƒÎ· Î‘Ï€Î¿ÏÏÎ¯Ï†Î¸Î·ÎºÎµ")
        final_error.geometry("350x120")
        final_error.resizable(False, False)
        final_error.configure(bg='#f0f0f0')
        
        # Center final error window
        final_error.update_idletasks()
        x = (final_error.winfo_screenwidth() // 2) - (350 // 2)
        y = (final_error.winfo_screenheight() // 2) - (120 // 2)
        final_error.geometry(f"350x120+{x}+{y}")
        
        final_error.transient(self.root)
        final_error.grab_set()
        final_error.attributes('-topmost', True)
        final_error.focus_force()
        
        tk.Label(final_error, text="âŒ Î ÏÏŒÏƒÎ²Î±ÏƒÎ· Î‘Ï€Î¿ÏÏÎ¯Ï†Î¸Î·ÎºÎµ!", 
                font=("Segoe UI", 12), bg='#f0f0f0', fg='red').pack(pady=15)
        tk.Label(final_error, text="Î¥Ï€ÎµÏÎ²Î®ÎºÎ±Ï„Îµ Ï„Î¿Î½ Î¼Î­Î³Î¹ÏƒÏ„Î¿ Î±ÏÎ¹Î¸Î¼ÏŒ Ï€ÏÎ¿ÏƒÏ€Î±Î¸ÎµÎ¹ÏÎ½", 
                font=("Segoe UI", 10), bg='#f0f0f0').pack(pady=5)
        
        def close_final():
            final_error.destroy()
        
        tk.Button(final_error, text="OK", command=close_final, 
                 font=("Segoe UI", 10)).pack(pady=10)
        
        final_error.wait_window()
        
        log_user_action("Authentication failed - max attempts exceeded")
        return False
    
    def _build_main_ui(self):
        """Build the main user interface with modern styling"""
        # Create main container with modern frame
        self.main_container = ModernFrame(self.root, theme=self.current_theme)
        self.main_container.pack(fill="both", expand=True)
        
        # Create menu bar
        self._create_menu_bar()
        
        # Create tab control
        self._create_tab_control()
        
        # Create modern status bar with tank indicator
        self.status_bar = StatusBar(self.main_container, theme=self.current_theme)
        self.status_bar.pack(side="bottom", fill="x")
        
        # Add tank indicator to status bar with modern styling
        self.tank_indicator = tk.Label(
            self.status_bar, 
            text="â›½ Î¦ÏŒÏÏ„Ï‰ÏƒÎ·...", 
            font=FONT_SMALL,
            bg=THEMES[self.current_theme]["header_bg"],
            fg=THEMES[self.current_theme]["text_muted"],
            padx=15,
            pady=8
        )
        self.tank_indicator.pack(side="right", padx=0)
        
        self.status_bar.set_status("Î•Ï†Î±ÏÎ¼Î¿Î³Î® Ï†Î¿ÏÏ„ÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
    
    def _create_menu_bar(self):
        """Create application menu bar"""
        menubar = tk.Menu(self.root, bg=THEMES[self.current_theme]["bg"], 
                         fg=THEMES[self.current_theme]["fg"])
        self.root.config(menu=menubar)
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0, bg=THEMES[self.current_theme]["bg"], 
                           fg=THEMES[self.current_theme]["fg"])
        menubar.add_cascade(label="ğŸ“ Î‘ÏÏ‡ÎµÎ¯Î¿", menu=file_menu)
        file_menu.add_command(label="ğŸ’¾ Backup Î’Î¬ÏƒÎ·Ï‚", command=self._backup_database)
        file_menu.add_separator()
        file_menu.add_command(label="ğŸšª ÎˆÎ¾Î¿Î´Î¿Ï‚", command=self._on_close)
        
        # View menu
        view_menu = tk.Menu(menubar, tearoff=0, bg=THEMES[self.current_theme]["bg"], 
                           fg=THEMES[self.current_theme]["fg"])
        menubar.add_cascade(label="ğŸ¨ Î•Î¼Ï†Î¬Î½Î¹ÏƒÎ·", menu=view_menu)
        
        # Theme submenu
        theme_menu = tk.Menu(view_menu, tearoff=0, bg=THEMES[self.current_theme]["bg"], 
                            fg=THEMES[self.current_theme]["fg"])
        view_menu.add_cascade(label="ğŸ¨ Î˜Î­Î¼Î±Ï„Î±", menu=theme_menu)
        
        # Add theme options
        theme_menu.add_command(label="â˜€ï¸ Î¦Ï‰Ï„ÎµÎ¹Î½ÏŒ", command=lambda: self._change_theme("light"))
        theme_menu.add_command(label="ğŸŒ™ Î£ÎºÎ¿Ï„ÎµÎ¹Î½ÏŒ", command=lambda: self._change_theme("dark"))
        theme_menu.add_command(label="ğŸ’™ ÎœÏ€Î»Îµ", command=lambda: self._change_theme("blue"))
        theme_menu.add_command(label="ğŸ’š Î ÏÎ¬ÏƒÎ¹Î½Î¿", command=lambda: self._change_theme("green"))
        theme_menu.add_command(label="ğŸ’œ ÎœÏ‰Î²", command=lambda: self._change_theme("purple"))
        
        view_menu.add_separator()
        view_menu.add_command(label="ğŸ”„ Î‘Î½Î±Î½Î­Ï‰ÏƒÎ· Î”ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½", command=self._refresh_all_data)
        
        # Tools menu
        tools_menu = tk.Menu(menubar, tearoff=0, bg=THEMES[self.current_theme]["bg"], 
                            fg=THEMES[self.current_theme]["fg"])
        menubar.add_cascade(label="ğŸ”§ Î•ÏÎ³Î±Î»ÎµÎ¯Î±", menu=tools_menu)
        
        # Theme tools submenu
        theme_tools_menu = tk.Menu(tools_menu, tearoff=0, bg=THEMES[self.current_theme]["bg"], 
                                  fg=THEMES[self.current_theme]["fg"])
        tools_menu.add_cascade(label="ğŸ¨ Î•Î½Î±Î»Î»Î±Î³Î® Î˜Î­Î¼Î±Ï„Î¿Ï‚", menu=theme_tools_menu)
        
        # Quick theme switching
        theme_tools_menu.add_command(label="ğŸ”„ Î•Ï€ÏŒÎ¼ÎµÎ½Î¿ Î˜Î­Î¼Î± (Ctrl+T)", command=self._toggle_theme)
        theme_tools_menu.add_separator()
        theme_tools_menu.add_command(label="â˜€ï¸ Î¦Ï‰Ï„ÎµÎ¹Î½ÏŒ", command=lambda: self._change_theme("light"))
        theme_tools_menu.add_command(label="ğŸŒ™ Î£ÎºÎ¿Ï„ÎµÎ¹Î½ÏŒ", command=lambda: self._change_theme("dark"))
        theme_tools_menu.add_command(label="ğŸ’™ ÎœÏ€Î»Îµ", command=lambda: self._change_theme("blue"))
        theme_tools_menu.add_command(label="ğŸ’š Î ÏÎ¬ÏƒÎ¹Î½Î¿", command=lambda: self._change_theme("green"))
        theme_tools_menu.add_command(label="ğŸ’œ ÎœÏ‰Î²", command=lambda: self._change_theme("purple"))
        
        tools_menu.add_separator()
        tools_menu.add_command(label="ğŸ”„ Î‘Î½Î±Î½Î­Ï‰ÏƒÎ· Î”ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½ (F5)", command=self._refresh_all_data)
        tools_menu.add_command(label="ğŸ’¾ Backup Î’Î¬ÏƒÎ·Ï‚ (Ctrl+S)", command=self._backup_database)
        tools_menu.add_separator()
        tools_menu.add_command(label="ğŸ“Š Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÎ¬ Î£Ï…ÏƒÏ„Î®Î¼Î±Ï„Î¿Ï‚", command=self._show_system_stats)
        tools_menu.add_command(label="ğŸ§¹ ÎšÎ±Î¸Î±ÏÎ¹ÏƒÎ¼ÏŒÏ‚ Î Î±Î»Î¹ÏÎ½ Î”ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½", command=self._cleanup_old_data)
        
        # Help menu
        help_menu = tk.Menu(menubar, tearoff=0, bg=THEMES[self.current_theme]["bg"], 
                           fg=THEMES[self.current_theme]["fg"])
        menubar.add_cascade(label="â“ Î’Î¿Î®Î¸ÎµÎ¹Î±", menu=help_menu)
        help_menu.add_command(label="ğŸ“– ÎŸÎ´Î·Î³Î¯ÎµÏ‚ Î§ÏÎ®ÏƒÎ·Ï‚", command=self._show_help)
        help_menu.add_command(label="â„¹ï¸ Î£Ï‡ÎµÏ„Î¹ÎºÎ¬", command=self._show_about)
    
    def _toggle_theme(self):
        """Toggle between themes in sequence"""
        themes_sequence = ["light", "dark", "blue", "green", "purple"]
        current_index = themes_sequence.index(self.current_theme)
        next_index = (current_index + 1) % len(themes_sequence)
        self._change_theme(themes_sequence[next_index])
    
    def _create_tab_control(self):
        """Create main tab control with modern styling"""
        # Configure ttk styles
        style = ttk.Style()
        style.theme_use('clam')
        
        theme_colors = THEMES[self.current_theme]
        
        # Configure modern tab styling
        style.configure('TNotebook', 
                       background=theme_colors["bg"],
                       borderwidth=0,
                       tabmargins=[0, 0, 0, 0])
        
        style.configure('TNotebook.Tab', 
                       padding=[24, 16], 
                       font=(FONT_NORMAL[0], FONT_NORMAL[1], "normal"),
                       background=theme_colors["frame_bg"],
                       foreground=theme_colors["text_secondary"],
                       borderwidth=0,
                       relief="flat")
        
        style.configure('TNotebook.Tab', 
                       focuscolor='none')
        
        # Modern tab hover and selection effects
        style.map('TNotebook.Tab',
                 background=[('selected', theme_colors["select_bg"]),
                            ('active', theme_colors["border_light"])],
                 foreground=[('selected', theme_colors["select_fg"]),
                            ('active', theme_colors["fg"])],
                 relief=[('selected', 'flat'),
                        ('active', 'flat')])
        
        self.tab_control = ttk.Notebook(self.main_container, style='TNotebook')
        
        # Create tabs with more distinct soft colors
        self.tab_movements = tk.Frame(self.tab_control, bg='#E8F5E8')    # Î‘Ï€Î±Î»ÏŒ Ï€ÏÎ¬ÏƒÎ¹Î½Î¿ Ï†ÏÎ»Î»Î¿Ï…
        self.tab_vehicles = tk.Frame(self.tab_control, bg='#FFF4E6')     # Î‘Ï€Î±Î»ÏŒ Ï€Î¿ÏÏ„Î¿ÎºÎ±Î»Î¯/ÎºÏÎµÎ¼
        self.tab_drivers = tk.Frame(self.tab_control, bg='#E6F3FF')      # Î‘Ï€Î±Î»ÏŒ Î³Î±Î»Î¬Î¶Î¹Î¿ Î¿Ï…ÏÎ±Î½Î¿Ï
        self.tab_driver_analytics = tk.Frame(self.tab_control, bg='#F0F8F0')  # Î‘Ï€Î±Î»ÏŒ Ï€ÏÎ¬ÏƒÎ¹Î½Î¿ Î¼Î­Î½Ï„Î±Ï‚
        self.tab_fuel = tk.Frame(self.tab_control, bg='#F0E6FF')         # Î‘Ï€Î±Î»ÏŒ Î»ÎµÎ²Î¬Î½Ï„Î±/Î¼Ï‰Î²
        self.tab_purposes = tk.Frame(self.tab_control, bg='#F5F5DC')     # Î‘Ï€Î±Î»ÏŒ Î¼Ï€ÎµÎ¶
        self.tab_reports = tk.Frame(self.tab_control, bg='#FFE6F0')      # Î‘Ï€Î±Î»ÏŒ ÏÏŒÎ¶
        
        # Add tabs with improved icons and labels
        self.tab_control.add(self.tab_movements, text="ğŸš— ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ ÎŸÏ‡Î·Î¼Î¬Ï„Ï‰Î½")
        self.tab_control.add(self.tab_vehicles, text="ğŸš™ Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· ÎŸÏ‡Î·Î¼Î¬Ï„Ï‰Î½")
        self.tab_control.add(self.tab_drivers, text="ğŸ‘¤ Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· ÎŸÎ´Î·Î³ÏÎ½")
        self.tab_control.add(self.tab_driver_analytics, text="ğŸ“ˆ Î‘Î½Î±Î»Ï…Ï„Î¹ÎºÎ¬ ÎŸÎ´Î·Î³ÏÎ½")
        self.tab_control.add(self.tab_fuel, text="â›½ Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½")
        self.tab_control.add(self.tab_purposes, text="ğŸ¯ Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· Î£ÎºÎ¿Ï€ÏÎ½")
        self.tab_control.add(self.tab_reports, text="ğŸ“Š Î‘Î½Î±Ï†Î¿ÏÎ­Ï‚ & Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÎ¬")
        
        self.tab_control.pack(expand=1, fill="both", padx=10, pady=5)
        
        # Build individual tabs
        self._build_movements_tab()
        self._build_vehicles_tab()
        self._build_drivers_tab()
        self._build_driver_analytics_tab()
        self._build_fuel_tab()
        self._build_purposes_tab()
        self._build_reports_tab()
    
    def _setup_event_handlers(self):
        """Setup global event handlers"""
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)
        self.root.bind('<F5>', lambda e: self._refresh_all_data())
        self.root.bind('<Control-s>', lambda e: self._backup_database())
        self.root.bind('<Control-q>', lambda e: self._on_close())
        self.root.bind('<Control-t>', lambda e: self._toggle_theme())  # Theme toggle shortcut
        
        # Window management shortcuts
        self.root.bind('<F11>', self._toggle_fullscreen)
        self.root.bind('<Alt-Return>', self._toggle_maximize)
        self.root.bind('<Control-0>', self._reset_window_size)
        
        # Track fullscreen state
        self.is_fullscreen = False
        
    def _toggle_fullscreen(self, event=None):
        """Toggle fullscreen mode"""
        try:
            self.is_fullscreen = not self.is_fullscreen
            if self.is_fullscreen:
                self.root.attributes('-fullscreen', True)
                self.root.bind('<Escape>', self._exit_fullscreen)
            else:
                self.root.attributes('-fullscreen', False)
                self.root.unbind('<Escape>')
        except Exception as e:
            logging.warning(f"Error toggling fullscreen: {e}")
    
    def _exit_fullscreen(self, event=None):
        """Exit fullscreen mode"""
        try:
            self.is_fullscreen = False
            self.root.attributes('-fullscreen', False)
            self.root.unbind('<Escape>')
        except Exception as e:
            logging.warning(f"Error exiting fullscreen: {e}")
    
    def _toggle_maximize(self, event=None):
        """Toggle maximized state"""
        try:
            if self.root.state() == 'zoomed':
                self.root.state('normal')
            else:
                self.root.state('zoomed')
        except Exception as e:
            logging.warning(f"Error toggling maximize: {e}")
    
    def _reset_window_size(self, event=None):
        """Reset window to default size"""
        try:
            screen_width = self.root.winfo_screenwidth()
            screen_height = self.root.winfo_screenheight()
            
            # Calculate default size (same logic as in _setup_window)
            if screen_width < 1366:
                window_width = int(screen_width * 0.95)
                window_height = int(screen_height * 0.90)
            elif screen_width < 1920:
                window_width = int(screen_width * 0.85)
                window_height = int(screen_height * 0.85)
            else:
                window_width = int(screen_width * 0.75)
                window_height = int(screen_height * 0.75)
            
            # Center window
            x = (screen_width // 2) - (window_width // 2)
            y = (screen_height // 2) - (window_height // 2)
            
            self.root.state('normal')
            self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
            
        except Exception as e:
            logging.warning(f"Error resetting window size: {e}")
    
    def _initialize_data(self):
        """Initialize application data"""
        try:
            self._refresh_all_data()
            self._update_tank_level()
            self.status_bar.set_status("Î”ÎµÎ´Î¿Î¼Î­Î½Î± Ï†Î¿ÏÏ„ÏÎ¸Î·ÎºÎ±Î½ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
        except Exception as e:
            logging.error(f"Error initializing data: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½: {str(e)}")
    
    def _build_movements_tab(self):
        """Build improved movements tab"""
        # Create main scrollable container
        canvas = tk.Canvas(self.tab_movements, bg=THEMES[self.current_theme]["bg"])
        scrollbar = ttk.Scrollbar(self.tab_movements, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # New movement form
        self._create_movement_form(scrollable_frame)
        
        # Active movements section
        self._create_active_movements_section(scrollable_frame)
        
        # Completed movements section
        self._create_completed_movements_section(scrollable_frame)
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def _create_movement_form(self, parent):
        """Create new movement form with improved validation"""
        form_frame = ModernFrame(parent, theme=self.current_theme, card=True, shadow=True)
        form_frame.pack(fill="x", padx=20, pady=20)
        
        # Title
        title_label = tk.Label(
            form_frame, 
            text="ğŸ“ ÎˆÎºÎ´Î¿ÏƒÎ· ÎÎ­Î±Ï‚ ÎšÎ¯Î½Î·ÏƒÎ·Ï‚ ÎŸÏ‡Î®Î¼Î±Ï„Î¿Ï‚", 
            font=FONT_TITLE, 
            fg=THEMES[self.current_theme]["accent"],
            bg=THEMES[self.current_theme]["frame_bg"]
        )
        title_label.pack(pady=(15, 20))
        
        # Form container with grid layout
        form_container = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        form_container.pack(fill="x", padx=20, pady=10)
        
        # Configure grid weights for responsive design
        for i in range(3):
            form_container.grid_columnconfigure(i, weight=1)
        
        # Date field
        tk.Label(form_container, text="ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0, sticky="w", padx=5, pady=8)
        
        self.mov_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.mov_date_entry = tk.Entry(
            form_container, 
            textvariable=self.mov_date_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.mov_date_entry.grid(row=0, column=1, sticky="ew", padx=5, pady=8)
        
        # Add tooltip
        self.tooltip_manager.add_tooltip(self.mov_date_entry, "ÎœÎ¿ÏÏ†Î®: YYYY-MM-DD")
        
        # Driver field with improved search
        tk.Label(form_container, text="ğŸ‘¤ ÎŸÎ´Î·Î³ÏŒÏ‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=1, column=0, sticky="w", padx=5, pady=8)
        
        self.mov_driver_combo = SearchableCombobox(
            form_container, 
            font=FONT_NORMAL, 
            placeholder="Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ· Î¿Î´Î·Î³Î¿Ï..."
        )
        self.mov_driver_combo.grid(row=1, column=1, sticky="ew", padx=5, pady=8)
        
        # Vehicle field with improved search
        tk.Label(form_container, text="ğŸš— ÎŒÏ‡Î·Î¼Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=2, column=0, sticky="w", padx=5, pady=8)
        
        self.mov_vehicle_combo = SearchableCombobox(
            form_container, 
            font=FONT_NORMAL,
            placeholder="Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ· Î¿Ï‡Î®Î¼Î±Ï„Î¿Ï‚..."
        )
        self.mov_vehicle_combo.grid(row=2, column=1, sticky="ew", padx=5, pady=8)
        
        # Start km field with validation
        tk.Label(form_container, text="ğŸ›£ï¸ Î§Î»Î¼ Î‘Î½Î±Ï‡ÏÏÎ·ÏƒÎ·Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=3, column=0, sticky="w", padx=5, pady=8)
        
        self.mov_start_km_var = tk.StringVar()
        self.mov_start_km_entry = tk.Entry(
            form_container, 
            textvariable=self.mov_start_km_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.mov_start_km_entry.grid(row=3, column=1, sticky="ew", padx=5, pady=8)
        
        # Route field
        tk.Label(form_container, text="ğŸ—ºï¸ Î”Î¹Î±Î´ÏÎ¿Î¼Î®:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=4, column=0, sticky="w", padx=5, pady=8)
        
        self.mov_route_entry = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.mov_route_entry.grid(row=4, column=1, sticky="ew", padx=5, pady=8)

        # Purpose field
        tk.Label(form_container, text="ğŸ¯ Î£ÎºÎ¿Ï€ÏŒÏ‚:", font=FONT_NORMAL,
                 bg=THEMES[self.current_theme]["frame_bg"],
                 fg=THEMES[self.current_theme]["fg"]).grid(row=5, column=0, sticky="w", padx=5, pady=8)
        # Get purposes from database
        self.purpose_options = self.db.get_purpose_names(active_only=True)
        self.mov_purpose_combobox = SearchableCombobox(
            form_container,
            values=self.purpose_options,
            font=FONT_NORMAL,
            width=30,
            placeholder="Î•Ï€Î¹Î»Î­Î¾Ï„Îµ ÏƒÎºÎ¿Ï€ÏŒ..."
        )
        self.mov_purpose_combobox.grid(row=5, column=1, sticky="ew", padx=5, pady=8)
        
        # Submit button
        button_frame = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        button_frame.pack(pady=20)
        
        # Movement submission button
        submit_btn = ModernButton(
            button_frame, 
            text="âœ… ÎˆÎºÎ´Î¿ÏƒÎ· ÎšÎ¯Î½Î·ÏƒÎ·Ï‚", 
            style="success",
            command=self._add_movement
        )
        submit_btn.pack(side="left", padx=5)
        
        # Browse movements button
        browse_btn = ModernButton(
            button_frame, 
            text="ğŸ“ Î ÎµÏÎ¹Î®Î³Î·ÏƒÎ· ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½", 
            style="info",
            command=self._browse_movement_documents
        )
        browse_btn.pack(side="left", padx=5)
        
        # Add tooltips
        self.tooltip_manager.add_tooltip(submit_btn, "Î”Î·Î¼Î¹Î¿Ï…ÏÎ³ÎµÎ¯ Î½Î­Î± ÎºÎ¯Î½Î·ÏƒÎ· ÎºÎ±Î¹ ÎµÎºÏ„Ï…Ï€ÏÏƒÎ¹Î¼Î¿ Î­Î³Î³ÏÎ±Ï†Î¿")
        self.tooltip_manager.add_tooltip(browse_btn, "Î ÎµÏÎ¹Î®Î³Î·ÏƒÎ· ÏƒÏ„Î¹Ï‚ Î±Ï€Î¿Î¸Î·ÎºÎµÏ…Î¼Î­Î½ÎµÏ‚ ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ Î±Î½Î¬ Î­Ï„Î¿Ï‚/Î¼Î®Î½Î±")
        
        # Bind vehicle selection to auto-fill
        self.mov_vehicle_combo.var.trace("w", self._auto_fill_last_km)
    
    def _create_active_movements_section(self, parent):
        """Create active movements section"""
        active_frame = ModernFrame(parent, theme=self.current_theme, card=True, shadow=True)
        active_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Title
        title_label = tk.Label(
            active_frame, 
            text="ğŸš¨ Î•Î½ÎµÏÎ³Î­Ï‚ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ (Î”ÎµÎ½ Î­Ï‡Î¿Ï…Î½ ÎµÏ€Î¹ÏƒÏ„ÏÎ­ÏˆÎµÎ¹)", 
            font=FONT_SUBTITLE, 
            fg=THEMES[self.current_theme]["warning"],
            bg=THEMES[self.current_theme]["frame_bg"]
        )
        title_label.pack(pady=(15, 10))
        
        # Search frame
        search_frame = tk.Frame(active_frame, bg=THEMES[self.current_theme]["frame_bg"])
        search_frame.pack(fill="x", padx=15, pady=5)
        
        tk.Label(search_frame, text="ğŸ” Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ·:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.active_search_var = tk.StringVar()
        self.active_search_entry = tk.Entry(
            search_frame, 
            textvariable=self.active_search_var,
            font=FONT_NORMAL, 
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.active_search_entry.pack(side="left", padx=10, fill="x", expand=True)
        self.active_search_var.trace("w", lambda *args: self._load_movements())
        
        # Tree for active movements
        tree_frame = tk.Frame(active_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tree_frame.pack(fill="both", expand=True, padx=15, pady=10)
        
        self.tree_active = ttk.Treeview(
            tree_frame,
            columns=("movement_num", "date", "driver", "vehicle", "start_km", "route", "purpose"),
            show="headings",
            height=6
        )
        
        # Configure columns
        columns_config = [
            ("movement_num", "ğŸ”¢ Î‘Ï. ÎšÎ¯Î½Î·ÏƒÎ·Ï‚", 100),
            ("date", "ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", 120),
            ("driver", "ğŸ‘¤ ÎŸÎ´Î·Î³ÏŒÏ‚", 150),
            ("vehicle", "ğŸš— ÎŒÏ‡Î·Î¼Î±", 120),
            ("start_km", "ğŸ›£ï¸ Î§Î»Î¼ Î‘Î½Î±Ï‡.", 100),
            ("route", "ğŸ—ºï¸ Î”Î¹Î±Î´ÏÎ¿Î¼Î®", 200),
            ("purpose", "ğŸ¯ Î£ÎºÎ¿Ï€ÏŒÏ‚", 150)
        ]
        
        for col, text, width in columns_config:
            self.tree_active.heading(col, text=text)
            self.tree_active.column(col, width=width, anchor="center")
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_active.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_active.xview)
        self.tree_active.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree_active.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Bind double-click for editing
        self.tree_active.bind("<Double-1>", self._edit_movement_return)
    
    def _create_completed_movements_section(self, parent):
        """Create completed movements section"""
        completed_frame = ModernFrame(parent, theme=self.current_theme)
        completed_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Title
        title_label = tk.Label(
            completed_frame, 
            text="âœ… ÎŸÎ»Î¿ÎºÎ»Î·ÏÏ‰Î¼Î­Î½ÎµÏ‚ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ (Î£Î®Î¼ÎµÏÎ±)", 
            font=FONT_SUBTITLE, 
            fg=THEMES[self.current_theme]["success"],
            bg=THEMES[self.current_theme]["frame_bg"]
        )
        title_label.pack(pady=(15, 10))
        
        # Search frame
        search_frame = tk.Frame(completed_frame, bg=THEMES[self.current_theme]["frame_bg"])
        search_frame.pack(fill="x", padx=15, pady=5)
        
        tk.Label(search_frame, text="ğŸ” Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ·:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.completed_search_var = tk.StringVar()
        self.completed_search_entry = tk.Entry(
            search_frame, 
            textvariable=self.completed_search_var,
            font=FONT_NORMAL, 
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.completed_search_entry.pack(side="left", padx=10, fill="x", expand=True)
        self.completed_search_var.trace("w", lambda *args: self._load_movements())
        
        # Tree for completed movements
        tree_frame = tk.Frame(completed_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tree_frame.pack(fill="both", expand=True, padx=15, pady=10)
        
        self.tree_completed = ttk.Treeview(
            tree_frame,
            columns=("movement_num", "date", "driver", "vehicle", "start_km", "end_km", "total_km", "route", "purpose"),
            show="headings",
            height=6
        )
        
        # Configure columns
        columns_config = [
            ("movement_num", "ğŸ”¢ Î‘Ï. ÎšÎ¯Î½Î·ÏƒÎ·Ï‚", 100),
            ("date", "ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", 120),
            ("driver", "ğŸ‘¤ ÎŸÎ´Î·Î³ÏŒÏ‚", 150),
            ("vehicle", "ğŸš— ÎŒÏ‡Î·Î¼Î±", 120),
            ("start_km", "ğŸ›£ï¸ Î§Î»Î¼ Î‘Î½Î±Ï‡.", 100),
            ("end_km", "ğŸ›£ï¸ Î§Î»Î¼ Î•Ï€Î¹ÏƒÏ„Ï.", 100),
            ("total_km", "ğŸ“Š Î£ÏÎ½Î¿Î»Î¿", 100),
            ("route", "ğŸ—ºï¸ Î”Î¹Î±Î´ÏÎ¿Î¼Î®", 200),
            ("purpose", "ğŸ¯ Î£ÎºÎ¿Ï€ÏŒÏ‚", 150)
        ]
        
        for col, text, width in columns_config:
            self.tree_completed.heading(col, text=text)
            self.tree_completed.column(col, width=width, anchor="center")
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_completed.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_completed.xview)
        self.tree_completed.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree_completed.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Bind double-click for editing
        self.tree_completed.bind("<Double-1>", self._edit_completed_movement)
    
    def _build_vehicles_tab(self):
        """Build improved vehicles tab with two-column layout"""
        # Main content frame
        main_content = tk.Frame(self.tab_vehicles, bg=THEMES[self.current_theme]["bg"])
        main_content.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Configure grid weights for responsive design  
        main_content.grid_columnconfigure(0, weight=1)  # Left column (form)
        main_content.grid_columnconfigure(1, weight=2)  # Right column (list)
        main_content.grid_rowconfigure(0, weight=1)
        
        # LEFT COLUMN - Vehicle form
        form_frame = ModernFrame(main_content, bg=THEMES[self.current_theme]["frame_bg"])
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=0)
        
        # Title with modern styling
        title_label = tk.Label(form_frame, text="ğŸš™ Î ÏÎ¿ÏƒÎ¸Î®ÎºÎ·/Î•Ï€ÎµÎ¾ÎµÏÎ³Î±ÏƒÎ¯Î± ÎŸÏ‡Î®Î¼Î±Ï„Î¿Ï‚", 
                              font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["frame_bg"])
        title_label.pack(pady=(15, 15))

        # Create form container
        form_container = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        form_container.pack(fill="x", padx=15, pady=5)
        
        # Configure grid weights for responsive design
        form_container.grid_columnconfigure(1, weight=1)
        
        # Plate field
        tk.Label(form_container, text="ğŸ”¢ Î Î¹Î½Î±ÎºÎ¯Î´Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0, sticky="w", padx=5, pady=5)
        
        self.ent_plate = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_plate.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        # Brand field
        tk.Label(form_container, text="ğŸ­ ÎœÎ¬ÏÎºÎ±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=1, column=0, sticky="w", padx=5, pady=5)
        
        self.ent_brand = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_brand.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        
        # Type field
        tk.Label(form_container, text="ğŸš— Î¤ÏÏ€Î¿Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=2, column=0, sticky="w", padx=5, pady=5)
        
        self.ent_vtype = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_vtype.grid(row=2, column=1, sticky="ew", padx=5, pady=5)

        # Purpose field
        tk.Label(form_container, text="ğŸ¯ Î£ÎºÎ¿Ï€ÏŒÏ‚:", font=FONT_NORMAL,
                 bg=THEMES[self.current_theme]["frame_bg"],
                 fg=THEMES[self.current_theme]["fg"]).grid(row=3, column=0, sticky="w", padx=5, pady=5)
        # Get purposes from database
        self.vehicle_purpose_options = self.db.get_purpose_names(active_only=True)
        self.ent_vpurpose = SearchableCombobox(
            form_container,
            values=self.vehicle_purpose_options,
            font=FONT_NORMAL,
            width=25,
            placeholder="Î•Ï€Î¹Î»Î­Î¾Ï„Îµ ÏƒÎºÎ¿Ï€ÏŒ..."
        )
        self.ent_vpurpose.grid(row=3, column=1, sticky="ew", padx=5, pady=5)
        
        # Photo field
        tk.Label(form_container, text="ğŸ–¼ï¸ Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=4, column=0, sticky="w", padx=5, pady=5)

        # Photo handling with modern styling
        photo_container = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        photo_container.grid(row=4, column=1, sticky="ew", padx=5, pady=5)
        
        self.photo_path_var = tk.StringVar()
        self.photo_label = tk.Label(photo_container, text="Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ ÎµÏ€Î¹Î»ÎµÎ³ÎµÎ¯ Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±", 
                                   fg=THEMES[self.current_theme]["text_muted"],
                                   bg=THEMES[self.current_theme]["frame_bg"], font=FONT_SMALL)
        self.photo_label.pack(side="left")
        
        photo_btn_frame = tk.Frame(photo_container, bg=THEMES[self.current_theme]["frame_bg"])
        photo_btn_frame.pack(side="right")
        
        ModernButton(photo_btn_frame, text="ğŸ“", style="secondary", 
                    command=self._select_photo).pack(side="left", padx=1)
        ModernButton(photo_btn_frame, text="ğŸ‘ï¸", style="info", 
                    command=self._view_photo).pack(side="left", padx=1)

        # Buttons with compact modern styling
        btn_frame = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        btn_frame.pack(pady=15)
        
        ModernButton(btn_frame, text="â• Î ÏÎ¿ÏƒÎ¸Î®ÎºÎ·", style="success", 
                    command=self._add_vehicle).pack(side="left", padx=3)
        ModernButton(btn_frame, text="âœï¸ Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ·", style="primary", 
                    command=self._update_vehicle).pack(side="left", padx=3)
        ModernButton(btn_frame, text="ğŸ—‘ï¸ Î”Î¹Î±Î³ÏÎ±Ï†Î®", style="danger", 
                    command=self._delete_vehicle).pack(side="left", padx=3)
        ModernButton(btn_frame, text="ğŸ§¹ Î•ÎºÎºÎ±Î¸Î¬ÏÎ¹ÏƒÎ·", style="warning", 
                    command=self._clear_vehicle_form).pack(side="left", padx=3)

        # RIGHT COLUMN - Vehicles list section
        vehicles_frame = ModernFrame(main_content, bg=THEMES[self.current_theme]["frame_bg"])
        vehicles_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=0)
        
        # Title for vehicles list
        vehicles_title = tk.Label(vehicles_frame, text="ğŸ“‹ Î›Î¯ÏƒÏ„Î± ÎŸÏ‡Î·Î¼Î¬Ï„Ï‰Î½", 
                                 font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                                 bg=THEMES[self.current_theme]["frame_bg"])
        vehicles_title.pack(pady=(15, 10))

        # Search frame with modern styling
        search_frame_v = tk.Frame(vehicles_frame, bg=THEMES[self.current_theme]["frame_bg"])
        search_frame_v.pack(fill="x", pady=5, padx=15)
        tk.Label(search_frame_v, text="ğŸ” Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ·:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        self.vehicles_search_var = tk.StringVar()
        self.vehicles_search_entry = tk.Entry(search_frame_v, textvariable=self.vehicles_search_var, 
                                             font=FONT_NORMAL, relief="flat", borderwidth=1, highlightthickness=1,
                                             highlightbackground=THEMES[self.current_theme]["border"],
                                             highlightcolor=THEMES[self.current_theme]["accent"])
        self.vehicles_search_entry.pack(side="left", padx=10, fill="x", expand=True)
        self.vehicles_search_var.trace("w", lambda *args: self._load_vehicles())

        # Tree with modern styling
        tree_container = tk.Frame(vehicles_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tree_container.pack(fill="both", expand=True, padx=15, pady=5)
        
        self.tree_vehicles = ttk.Treeview(tree_container, 
                                        columns=("plate", "brand", "vtype", "purpose"), 
                                        show="headings", height=12)
        
        # Configure columns
        columns_config = [
            ("plate", "ğŸ”¢ Î Î¹Î½Î±ÎºÎ¯Î´Î±", 120),
            ("brand", "ğŸ­ ÎœÎ¬ÏÎºÎ±", 150),
            ("vtype", "ğŸš— Î¤ÏÏ€Î¿Ï‚", 120),
            ("purpose", "ğŸ¯ Î£ÎºÎ¿Ï€ÏŒÏ‚", 200)
        ]
        
        for col, text, width in columns_config:
            self.tree_vehicles.heading(col, text=text)
            self.tree_vehicles.column(col, width=width, anchor="center")
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_vehicles.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree_vehicles.xview)
        self.tree_vehicles.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree_vehicles.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Bind events
        self.tree_vehicles.bind("<Double-1>", self._select_vehicle_from_tree)
        self.tree_vehicles.bind("<Button-3>", self._show_vehicle_context_menu)  # Right click
        
        # Create context menu for vehicles with modern styling
        self.vehicle_context_menu = tk.Menu(self.root, tearoff=0, 
                                           bg=THEMES[self.current_theme]["frame_bg"], 
                                           fg=THEMES[self.current_theme]["fg"],
                                           activebackground=THEMES[self.current_theme]["select_bg"],
                                           activeforeground=THEMES[self.current_theme]["select_fg"])
        self.vehicle_context_menu.add_command(label="ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½", command=self._show_vehicle_history)
        self.vehicle_context_menu.add_command(label="â›½ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½", command=self._show_vehicle_fuel_history)
        self.vehicle_context_menu.add_separator()
        self.vehicle_context_menu.add_command(label="ğŸ“Š Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÎ¬ ÎŸÏ‡Î®Î¼Î±Ï„Î¿Ï‚", command=self._show_vehicle_statistics)
        self.vehicle_context_menu.add_separator()
        self.vehicle_context_menu.add_command(label="âœï¸ Î•Ï€ÎµÎ¾ÎµÏÎ³Î±ÏƒÎ¯Î±", command=self._select_vehicle_from_tree)
        self.vehicle_context_menu.add_command(label="ğŸ–¼ï¸ Î ÏÎ¿Î²Î¿Î»Î® Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±Ï‚", command=self._view_vehicle_photo_from_tree)

        # Load initial data
        self._load_vehicles()
    
    def _build_drivers_tab(self):
        """Build improved drivers tab with two-column layout"""
        # Main container with scrollable frame
        main_canvas = tk.Canvas(self.tab_drivers, bg=THEMES[self.current_theme]["bg"], 
                               highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_drivers, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)

        # Title section
        title_frame = tk.Frame(scrollable_frame, bg=THEMES[self.current_theme]["bg"])
        title_frame.pack(fill="x", padx=10, pady=(10, 5))  # Reduced padding
        
        title_label = tk.Label(title_frame, text="ğŸ‘¤ Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· ÎŸÎ´Î·Î³ÏÎ½", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack()

        # Main content section with two columns
        content_section = tk.Frame(scrollable_frame, bg=THEMES[self.current_theme]["bg"])
        content_section.pack(fill="both", expand=True, padx=10, pady=5)  # Reduced padding
        
        # Configure grid weights - optimize space usage
        content_section.grid_columnconfigure(0, weight=2, minsize=300)  # Form column (reduced weight)
        content_section.grid_columnconfigure(1, weight=3, minsize=600)  # List column (increased weight)
        content_section.grid_rowconfigure(0, weight=1)  # Make row expandable

        # Left column - Driver form
        form_frame = ModernFrame(content_section, bg=THEMES[self.current_theme]["frame_bg"])
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5), pady=0)  # Reduced padding
        
        # Form title
        form_title = tk.Label(form_frame, text="â• ÎšÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ· ÎŸÎ´Î·Î³Î¿Ï", 
                             font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                             bg=THEMES[self.current_theme]["frame_bg"])
        form_title.pack(pady=(10, 15))  # Reduced padding

        # Form fields container
        form_container = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        form_container.pack(fill="x", padx=10, pady=5)  # Reduced padding
        
        # Name field
        tk.Label(form_container, text="ğŸ‘¤ ÎŒÎ½Î¿Î¼Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        self.ent_name = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_name.pack(fill="x", pady=(0, 15))
        
        # Surname field
        tk.Label(form_container, text="ğŸ‘¤ Î•Ï€ÏÎ½Ï…Î¼Î¿:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        self.ent_surname = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_surname.pack(fill="x", pady=(0, 15))
        
        # Notes field
        tk.Label(form_container, text="ğŸ“ Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        self.ent_dnotes = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_dnotes.pack(fill="x", pady=(0, 15))  # Reduced padding

        # Buttons
        btn_frame = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        btn_frame.pack(pady=10)  # Reduced padding
        
        ModernButton(btn_frame, text="â• Î ÏÎ¿ÏƒÎ¸Î®ÎºÎ·", style="success", 
                    command=self._add_driver).pack(pady=2)
        ModernButton(btn_frame, text="âœï¸ Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ·", style="primary", 
                    command=self._update_driver).pack(pady=2)
        ModernButton(btn_frame, text="ğŸ—‘ï¸ Î”Î¹Î±Î³ÏÎ±Ï†Î®", style="danger", 
                    command=self._delete_driver).pack(pady=2)
        ModernButton(btn_frame, text="ğŸ§¹ Î•ÎºÎºÎ±Î¸Î¬ÏÎ¹ÏƒÎ·", style="warning", 
                    command=self._clear_driver_form).pack(pady=2)

        # Right column - Drivers list
        drivers_frame = ModernFrame(content_section, bg=THEMES[self.current_theme]["frame_bg"])
        drivers_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0), pady=0)  # Reduced padding
        
        # List title
        drivers_title = tk.Label(drivers_frame, text="ğŸ“‹ Î›Î¯ÏƒÏ„Î± ÎŸÎ´Î·Î³ÏÎ½", 
                                font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                                bg=THEMES[self.current_theme]["frame_bg"])
        drivers_title.pack(pady=(10, 8))  # Reduced padding

        # Search frame
        search_frame_d = tk.Frame(drivers_frame, bg=THEMES[self.current_theme]["frame_bg"])
        search_frame_d.pack(fill="x", pady=(0, 8), padx=10)  # Reduced padding
        
        tk.Label(search_frame_d, text="ğŸ” Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ·:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.drivers_search_var = tk.StringVar()
        self.drivers_search_entry = tk.Entry(search_frame_d, textvariable=self.drivers_search_var, 
                                           font=FONT_NORMAL, relief="flat", borderwidth=1, highlightthickness=1,
                                           highlightbackground=THEMES[self.current_theme]["border"],
                                           highlightcolor=THEMES[self.current_theme]["accent"])
        self.drivers_search_entry.pack(side="left", padx=(10, 0), fill="x", expand=True)
        self.drivers_search_var.trace("w", lambda *args: self._load_drivers())

        # Tree container
        tree_container = tk.Frame(drivers_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tree_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))  # Reduced padding
        
        self.tree_drivers = ttk.Treeview(tree_container, 
                                       columns=("name", "surname", "notes"), 
                                       show="headings", height=15)  # Increased height
        
        # Configure columns
        columns_config = [
            ("name", "ğŸ‘¤ ÎŒÎ½Î¿Î¼Î±", 120),
            ("surname", "ğŸ‘¤ Î•Ï€ÏÎ½Ï…Î¼Î¿", 120),
            ("notes", "ğŸ“ Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚", 200)
        ]
        
        for col, text, width in columns_config:
            self.tree_drivers.heading(col, text=text)
            self.tree_drivers.column(col, width=width, anchor="center")
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_drivers.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree_drivers.xview)
        self.tree_drivers.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree_drivers.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        
        # Bind events
        self.tree_drivers.bind("<Double-1>", self._select_driver_from_tree)

        # Pack canvas and scrollbar
        main_canvas.pack(side="left", fill="both", expand=True, padx=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10))

        # Load initial data
        self._load_drivers()
    
    def _build_driver_analytics_tab(self):
        """Build driver analytics tab"""
        # Main container
        main_frame = tk.Frame(self.tab_driver_analytics, bg=THEMES[self.current_theme]["bg"])
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Title
        title_label = tk.Label(main_frame, text="ğŸ“ˆ Î‘Î½Î±Î»Ï…Ï„Î¹ÎºÎ¬ Î£Ï„Î¿Î¹Ï‡ÎµÎ¯Î± ÎŸÎ´Î·Î³ÏÎ½", 
                              font=FONT_BIG, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack(pady=(0, 15))
        
        # Controls frame
        controls_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        controls_frame.pack(fill="x", pady=(0, 15))
        
        # Driver selection
        tk.Label(controls_frame, text="ğŸ§‘ Î•Ï€Î¹Î»Î¿Î³Î® ÎŸÎ´Î·Î³Î¿Ï:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left", padx=(0, 10))
        
        self.analytics_driver_combo = SearchableCombobox(
            controls_frame, 
            font=FONT_NORMAL, width=25
        )
        self.analytics_driver_combo.pack(side="left", padx=(0, 20))
        
        # Date range
        tk.Label(controls_frame, text="ğŸ“… Î‘Ï€ÏŒ:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left", padx=(0, 5))
        
        self.analytics_start_date = tk.Entry(controls_frame, font=FONT_NORMAL, width=12)
        self.analytics_start_date.pack(side="left", padx=(0, 10))
        
        tk.Label(controls_frame, text="ğŸ“… ÎˆÏ‰Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left", padx=(0, 5))
        
        self.analytics_end_date = tk.Entry(controls_frame, font=FONT_NORMAL, width=12)
        self.analytics_end_date.pack(side="left", padx=(0, 15))
        
        # Buttons
        ModernButton(controls_frame, text="ğŸ“Š Î‘Î½Î¬Î»Ï…ÏƒÎ·", style="primary",
                    command=self._show_driver_analytics).pack(side="left", padx=(0, 10))
        
        ModernButton(controls_frame, text="ğŸ‘¥ ÎŒÎ»Î¿Î¹ ÎŸÎ´Î·Î³Î¿Î¯", style="info",
                    command=self._show_all_drivers_summary).pack(side="left")
        
        # Results area with notebook for different views
        results_notebook = ttk.Notebook(main_frame)
        results_notebook.pack(fill="both", expand=True)
        
        # Summary tab
        self.summary_frame = tk.Frame(results_notebook, bg=THEMES[self.current_theme]["frame_bg"])
        results_notebook.add(self.summary_frame, text="ğŸ“‹ Î£ÏÎ½Î¿ÏˆÎ·")
        
        # Details tab
        self.details_frame = tk.Frame(results_notebook, bg=THEMES[self.current_theme]["frame_bg"])
        results_notebook.add(self.details_frame, text="ğŸ“ˆ Î›ÎµÏ€Ï„Î¿Î¼Î­ÏÎµÎ¹ÎµÏ‚")
        
        # Comparison tab
        self.comparison_frame = tk.Frame(results_notebook, bg=THEMES[self.current_theme]["frame_bg"])
        results_notebook.add(self.comparison_frame, text="âš–ï¸ Î£ÏÎ³ÎºÏÎ¹ÏƒÎ·")
        
        # Load drivers for selection
        self._load_analytics_drivers()
    
    def _build_fuel_tab(self):
        """Build improved fuel tab"""
        # Main container with scrollable frame
        main_canvas = tk.Canvas(self.tab_fuel, bg=THEMES[self.current_theme]["bg"], 
                               highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_fuel, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)

        # Top section with tank and pump side by side
        top_section = tk.Frame(scrollable_frame, bg=THEMES[self.current_theme]["bg"])
        top_section.pack(fill="x", padx=20, pady=20)
        
        # Configure grid weights for equal columns
        top_section.grid_columnconfigure(0, weight=1)
        top_section.grid_columnconfigure(1, weight=1)

        # Tank display (left column)
        tank_frame = ModernFrame(top_section, bg=THEMES[self.current_theme]["frame_bg"])
        tank_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        
        tank_title = tk.Label(tank_frame, text="â›½ ÎšÎ±Ï„Î¬ÏƒÏ„Î±ÏƒÎ· Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚", 
                             font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                             bg=THEMES[self.current_theme]["frame_bg"])
        tank_title.pack(pady=(15, 10))
        
        # Tank info container
        tank_info_frame = tk.Frame(tank_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tank_info_frame.pack(fill="x", padx=20, pady=10)
        
        self.tank_level_label = tk.Label(tank_info_frame, text="Î¦ÏŒÏÏ„Ï‰ÏƒÎ·...", 
                                        font=FONT_BIG, fg=THEMES[self.current_theme]["success"],
                                        bg=THEMES[self.current_theme]["frame_bg"])
        self.tank_level_label.pack(pady=10)
        
        # Tank capacity info
        capacity_label = tk.Label(tank_info_frame, 
                                 text=f"ğŸ›¢ï¸ Î£Ï…Î½Î¿Î»Î¹ÎºÎ® Î§Ï‰ÏÎ·Ï„Î¹ÎºÏŒÏ„Î·Ï„Î±: {TANK_CAPACITY:,.0f} Î›Î¯Ï„ÏÎ±", 
                                 font=FONT_NORMAL, fg=THEMES[self.current_theme]["text_muted"],
                                 bg=THEMES[self.current_theme]["frame_bg"])
        capacity_label.pack(pady=(0, 5))
        
        # Progress bar for tank level
        progress_frame = tk.Frame(tank_info_frame, bg=THEMES[self.current_theme]["frame_bg"])
        progress_frame.pack(fill="x", padx=20, pady=10)
        
        tk.Label(progress_frame, text="Î•Ï€Î¯Ï€ÎµÎ´Î¿ Î Î»Î®ÏÏ‰ÏƒÎ·Ï‚:", 
                font=FONT_SMALL, fg=THEMES[self.current_theme]["text_muted"],
                bg=THEMES[self.current_theme]["frame_bg"]).pack(anchor="w")
        
        self.tank_progress = ttk.Progressbar(progress_frame, length=300, mode='determinate')
        self.tank_progress.pack(fill="x", pady=(5, 10))
        
        # Tank management buttons
        tank_btn_frame = tk.Frame(tank_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tank_btn_frame.pack(pady=15)
        
        ModernButton(tank_btn_frame, text="â›½ Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚", style="primary", 
                    command=self._refill_tank).pack(side="top", pady=2)
        ModernButton(tank_btn_frame, text="ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ", style="info", 
                    command=self._show_tank_history).pack(side="top", pady=2)

        # Pump display (right column)
        pump_frame = ModernFrame(top_section, bg=THEMES[self.current_theme]["frame_bg"])
        pump_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
        
        pump_title = tk.Label(pump_frame, text="â›½ ÎšÎ±Ï„Î¬ÏƒÏ„Î±ÏƒÎ· Î‘Î½Ï„Î»Î¯Î±Ï‚", 
                             font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                             bg=THEMES[self.current_theme]["frame_bg"])
        pump_title.pack(pady=(15, 10))
        
        # Pump info container
        pump_info_frame = tk.Frame(pump_frame, bg=THEMES[self.current_theme]["frame_bg"])
        pump_info_frame.pack(fill="x", padx=20, pady=10)
        
        self.pump_reading_label = tk.Label(pump_info_frame, text="Î¦ÏŒÏÏ„Ï‰ÏƒÎ·...", 
                                          font=FONT_BIG, fg=THEMES[self.current_theme]["info"],
                                          bg=THEMES[self.current_theme]["frame_bg"])
        self.pump_reading_label.pack(pady=10)
        
        # Pump info
        pump_info_label = tk.Label(pump_info_frame, 
                                  text="ğŸ“Š Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ Î»Î¯Ï„ÏÎ± Ï€Î¿Ï… Î´Î¹Î±Î½ÎµÎ¼Î®Î¸Î·ÎºÎ±Î½ Î±Ï€ÏŒ Ï„Î·Î½ Î±Î½Ï„Î»Î¯Î±", 
                                  font=FONT_NORMAL, fg=THEMES[self.current_theme]["text_muted"],
                                  bg=THEMES[self.current_theme]["frame_bg"])
        pump_info_label.pack(pady=(0, 5))
        
        # Spacer for visual balance
        spacer_frame = tk.Frame(pump_info_frame, bg=THEMES[self.current_theme]["frame_bg"], height=45)
        spacer_frame.pack(fill="x")
        
        # Pump management buttons
        pump_btn_frame = tk.Frame(pump_frame, bg=THEMES[self.current_theme]["frame_bg"])
        pump_btn_frame.pack(pady=15)
        
        ModernButton(pump_btn_frame, text="ğŸ“Š Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ·", style="success", 
                    command=self._update_pump_reading).pack(side="top", pady=2)
        ModernButton(pump_btn_frame, text="ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ", style="info", 
                    command=self._show_pump_history).pack(side="top", pady=2)

        # Form for fuel addition with modern styling - horizontal layout
        content_section = tk.Frame(scrollable_frame, bg=THEMES[self.current_theme]["bg"])
        content_section.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Configure grid weights for two columns
        content_section.grid_columnconfigure(0, weight=1)  # Form column
        content_section.grid_columnconfigure(1, weight=2)  # History column (wider)

        # Left side - Form
        form_frame = ModernFrame(content_section, bg=THEMES[self.current_theme]["frame_bg"])
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        
        # Title with modern styling
        title_label = tk.Label(form_frame, text="â›½ ÎšÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ· ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["frame_bg"])
        title_label.pack(pady=(15, 20))

        # Create compact form container
        form_container = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        form_container.pack(fill="x", padx=15, pady=10)
        
        # Date field - compact layout
        date_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        date_frame.pack(fill="x", pady=5)
        
        tk.Label(date_frame, text="ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±:", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.fuel_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.fuel_date_entry = tk.Entry(
            date_frame, 
            textvariable=self.fuel_date_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"],
            width=12
        )
        self.fuel_date_entry.pack(side="right", fill="x", expand=True)
        
        # Driver field - compact layout
        driver_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        driver_frame.pack(fill="x", pady=5)
        
        tk.Label(driver_frame, text="ğŸ‘¤ ÎŸÎ´Î·Î³ÏŒÏ‚:", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.fuel_driver_combo = SearchableCombobox(
            driver_frame, 
            font=FONT_NORMAL, 
            placeholder="Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ· Î¿Î´Î·Î³Î¿Ï...",
            width=12
        )
        self.fuel_driver_combo.pack(side="right", fill="x", expand=True)
        
        # Vehicle field - compact layout
        vehicle_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        vehicle_frame.pack(fill="x", pady=5)
        
        tk.Label(vehicle_frame, text="ğŸš— ÎŒÏ‡Î·Î¼Î±:", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.fuel_vehicle_combo = SearchableCombobox(
            vehicle_frame, 
            font=FONT_NORMAL,
            placeholder="Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ· Î¿Ï‡Î®Î¼Î±Ï„Î¿Ï‚...",
            width=12
        )
        self.fuel_vehicle_combo.pack(side="right", fill="x", expand=True)
        
        # Liters field - compact layout
        liters_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        liters_frame.pack(fill="x", pady=5)
        
        tk.Label(liters_frame, text="â›½ Î›Î¯Ï„ÏÎ±:", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.fuel_liters_var = tk.StringVar()
        self.fuel_liters_entry = tk.Entry(
            liters_frame, 
            textvariable=self.fuel_liters_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"],
            width=12
        )
        self.fuel_liters_entry.pack(side="right", fill="x", expand=True)
        
        # Mileage field - compact layout
        mileage_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        mileage_frame.pack(fill="x", pady=5)
        
        tk.Label(mileage_frame, text="ğŸ›£ï¸ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±:", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.fuel_mileage_var = tk.StringVar()
        self.fuel_mileage_entry = tk.Entry(
            mileage_frame, 
            textvariable=self.fuel_mileage_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"],
            width=12
        )
        self.fuel_mileage_entry.pack(side="right", fill="x", expand=True)
        
        # Cost field - compact layout
        cost_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        cost_frame.pack(fill="x", pady=5)
        
        tk.Label(cost_frame, text="ğŸ’° ÎšÏŒÏƒÏ„Î¿Ï‚ (â‚¬):", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.fuel_cost_var = tk.StringVar()
        self.fuel_cost_entry = tk.Entry(
            cost_frame, 
            textvariable=self.fuel_cost_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"],
            width=12
        )
        self.fuel_cost_entry.pack(side="right", fill="x", expand=True)

        # Pump reading field - compact layout
        pump_frame = tk.Frame(form_container, bg=THEMES[self.current_theme]["frame_bg"])
        pump_frame.pack(fill="x", pady=5)
        
        tk.Label(pump_frame, text="ğŸ“Š ÎœÎ­Ï„ÏÎ·ÏƒÎ·:", font=FONT_NORMAL, width=12, anchor="w",
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        pump_input_frame = tk.Frame(pump_frame, bg=THEMES[self.current_theme]["frame_bg"])
        pump_input_frame.pack(side="right", fill="x", expand=True)
        
        self.fuel_pump_reading_var = tk.StringVar()
        self.fuel_pump_reading_entry = tk.Entry(
            pump_input_frame, 
            textvariable=self.fuel_pump_reading_var, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"],
            width=10
        )
        self.fuel_pump_reading_entry.pack(side="left", fill="x", expand=True)
        
        # Button to get current pump reading
        ModernButton(pump_input_frame, text="ğŸ“Š", style="info", 
                    command=self._get_current_pump_reading).pack(side="right", padx=(5, 0))

        # Buttons with modern styling
        btn_frame = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        btn_frame.pack(pady=20)
        
        ModernButton(btn_frame, text="â• Î ÏÎ¿ÏƒÎ¸Î®ÎºÎ·", style="success", 
                    command=self._add_fuel).pack(pady=5, fill="x")
        ModernButton(btn_frame, text="ğŸ§¹ Î•ÎºÎºÎ±Î¸Î¬ÏÎ¹ÏƒÎ·", style="warning", 
                    command=self._clear_fuel_form).pack(pady=5, fill="x")

        # Right side - Fuel records with modern styling
        fuel_frame = ModernFrame(content_section, bg=THEMES[self.current_theme]["frame_bg"])
        fuel_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
        
        # Title for fuel records
        fuel_title = tk.Label(fuel_frame, text="ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½", 
                             font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                             bg=THEMES[self.current_theme]["frame_bg"])
        fuel_title.pack(pady=(15, 10))

        # Tree with modern styling
        tree_container = tk.Frame(fuel_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tree_container.pack(fill="both", expand=True, padx=15, pady=10)
        
        self.tree_fuel = ttk.Treeview(tree_container, 
                                    columns=("date", "driver", "vehicle", "liters", "mileage", "cost"), 
                                    show="headings", height=8)
        
        # Configure columns
        columns_config = [
            ("date", "ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", 120),
            ("driver", "ğŸ‘¤ ÎŸÎ´Î·Î³ÏŒÏ‚", 150),
            ("vehicle", "ğŸš— ÎŒÏ‡Î·Î¼Î±", 120),
            ("liters", "â›½ Î›Î¯Ï„ÏÎ±", 100),
            ("mileage", "ğŸ›£ï¸ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", 120),
            ("cost", "ğŸ’° ÎšÏŒÏƒÏ„Î¿Ï‚", 100)
        ]
        
        for col, text, width in columns_config:
            self.tree_fuel.heading(col, text=text)
            self.tree_fuel.column(col, width=width, anchor="center")
        
        # Add scrollbars
        v_scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_fuel.yview)
        h_scrollbar = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree_fuel.xview)
        self.tree_fuel.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        self.tree_fuel.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        # Pack canvas and scrollbar
        main_canvas.pack(side="left", fill="both", expand=True, padx=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10))

        # Load initial data
        self._load_fuel()
        self._load_fuel_combos()
    
    def _build_purposes_tab(self):
        """Build purposes management tab with two-column layout"""
        # Main container with scrollable frame
        main_canvas = tk.Canvas(self.tab_purposes, bg=THEMES[self.current_theme]["bg"], 
                               highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.tab_purposes, orient="vertical", command=main_canvas.yview)
        scrollable_frame = ttk.Frame(main_canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=scrollbar.set)

        # Title section
        title_frame = tk.Frame(scrollable_frame, bg=THEMES[self.current_theme]["bg"])
        title_frame.pack(fill="x", padx=20, pady=(20, 10))
        
        title_label = tk.Label(title_frame, text="ğŸ¯ Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· Î£ÎºÎ¿Ï€ÏÎ½", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack()

        # Main content section with two columns
        content_section = tk.Frame(scrollable_frame, bg=THEMES[self.current_theme]["bg"])
        content_section.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Configure grid weights for equal columns
        content_section.grid_columnconfigure(0, weight=1)
        content_section.grid_columnconfigure(1, weight=2)  # Give more space to the list

        # Left column - Purpose form
        form_frame = ModernFrame(content_section, bg=THEMES[self.current_theme]["frame_bg"])
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        
        # Form title
        form_title = tk.Label(form_frame, text="â• ÎšÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ· Î£ÎºÎ¿Ï€Î¿Ï", 
                             font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                             bg=THEMES[self.current_theme]["frame_bg"])
        form_title.pack(pady=(15, 20))

        # Form fields container
        form_container = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        form_container.pack(fill="x", padx=15, pady=10)
        
        # Name field
        tk.Label(form_container, text="ğŸ¯ ÎŒÎ½Î¿Î¼Î± Î£ÎºÎ¿Ï€Î¿Ï:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        self.ent_purpose_name = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_purpose_name.pack(fill="x", pady=(0, 15))
        
        # Description field
        tk.Label(form_container, text="ğŸ“ Î ÎµÏÎ¹Î³ÏÎ±Ï†Î®:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        self.ent_purpose_description = tk.Entry(
            form_container, 
            font=FONT_NORMAL,
            relief="flat", 
            borderwidth=1, 
            highlightthickness=1,
            highlightbackground=THEMES[self.current_theme]["border"],
            highlightcolor=THEMES[self.current_theme]["accent"]
        )
        self.ent_purpose_description.pack(fill="x", pady=(0, 15))
        
        # Category field
        tk.Label(form_container, text="ğŸ“‚ ÎšÎ±Ï„Î·Î³Î¿ÏÎ¯Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        self.purpose_categories = ["general", "operation", "transport", "maintenance", "cleaning", "repair", "inspection"]
        self.purpose_category_var = tk.StringVar(value="general")
        self.purpose_category_combo = ttk.Combobox(
            form_container,
            textvariable=self.purpose_category_var,
            values=self.purpose_categories,
            state="readonly",
            font=FONT_NORMAL
        )
        self.purpose_category_combo.pack(fill="x", pady=(0, 20))

        # Buttons
        btn_frame = tk.Frame(form_frame, bg=THEMES[self.current_theme]["frame_bg"])
        btn_frame.pack(pady=15)
        
        ModernButton(btn_frame, text="â• Î ÏÎ¿ÏƒÎ¸Î®ÎºÎ·", style="success", 
                    command=self._add_purpose).pack(pady=2)
        ModernButton(btn_frame, text="âœï¸ Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ·", style="primary", 
                    command=self._update_purpose).pack(pady=2)
        ModernButton(btn_frame, text="ğŸ—‘ï¸ Î‘Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¯Î·ÏƒÎ·", style="danger", 
                    command=self._delete_purpose).pack(pady=2)
        ModernButton(btn_frame, text="â™»ï¸ Î•Ï€Î±Î½Î±Ï†Î¿ÏÎ¬", style="warning", 
                    command=self._restore_purpose).pack(pady=2)
        ModernButton(btn_frame, text="ğŸ§¹ Î•ÎºÎºÎ±Î¸Î¬ÏÎ¹ÏƒÎ·", style="warning", 
                    command=self._clear_purpose_form).pack(pady=2)

        # Right column - Purposes list
        purposes_frame = ModernFrame(content_section, bg=THEMES[self.current_theme]["frame_bg"])
        purposes_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
        
        # List title
        purposes_title = tk.Label(purposes_frame, text="ğŸ“‹ Î›Î¯ÏƒÏ„Î± Î£ÎºÎ¿Ï€ÏÎ½", 
                                font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                                bg=THEMES[self.current_theme]["frame_bg"])
        purposes_title.pack(pady=(15, 10))

        # Search and filter frame
        search_frame = tk.Frame(purposes_frame, bg=THEMES[self.current_theme]["frame_bg"])
        search_frame.pack(fill="x", pady=(0, 10), padx=15)
        
        # Search field
        tk.Label(search_frame, text="ğŸ” Î‘Î½Î±Î¶Î®Ï„Î·ÏƒÎ·:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.purposes_search_var = tk.StringVar()
        self.purposes_search_entry = tk.Entry(search_frame, textvariable=self.purposes_search_var, 
                                           font=FONT_NORMAL, relief="flat", borderwidth=1, highlightthickness=1,
                                           highlightbackground=THEMES[self.current_theme]["border"],
                                           highlightcolor=THEMES[self.current_theme]["accent"])
        self.purposes_search_entry.pack(side="left", padx=(10, 15), fill="x", expand=True)
        self.purposes_search_var.trace("w", lambda *args: self._load_purposes())
        
        # Filter by category
        tk.Label(search_frame, text="ğŸ“‚ ÎšÎ±Ï„Î·Î³Î¿ÏÎ¯Î±:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(side="left")
        
        self.purpose_filter_var = tk.StringVar(value="all")
        filter_options = ["all"] + self.purpose_categories
        self.purpose_filter_combo = ttk.Combobox(
            search_frame,
            textvariable=self.purpose_filter_var,
            values=filter_options,
            state="readonly",
            font=FONT_NORMAL,
            width=12
        )
        self.purpose_filter_combo.pack(side="left", padx=(5, 0))
        self.purpose_filter_var.trace("w", lambda *args: self._load_purposes())

        # Show inactive checkbox
        self.show_inactive_var = tk.BooleanVar()
        show_inactive_cb = tk.Checkbutton(
            search_frame,
            text="Î•Î¼Ï†Î¬Î½Î¹ÏƒÎ· Î±Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¹Î·Î¼Î­Î½Ï‰Î½",
            variable=self.show_inactive_var,
            font=FONT_SMALL,
            bg=THEMES[self.current_theme]["frame_bg"],
            fg=THEMES[self.current_theme]["fg"],
            command=self._load_purposes
        )
        show_inactive_cb.pack(side="right", padx=(15, 0))

        # Tree container
        tree_container = tk.Frame(purposes_frame, bg=THEMES[self.current_theme]["frame_bg"])
        tree_container.pack(fill="both", expand=True, padx=15, pady=(0, 15))
        
        self.tree_purposes = ttk.Treeview(tree_container, 
                                       columns=("name", "description", "category", "active"), 
                                       show="headings", height=12)
        
        # Configure columns
        columns_config = [
            ("name", "ğŸ¯ ÎŒÎ½Î¿Î¼Î±", 150),
            ("description", "ğŸ“ Î ÎµÏÎ¹Î³ÏÎ±Ï†Î®", 200),
            ("category", "ğŸ“‚ ÎšÎ±Ï„Î·Î³Î¿ÏÎ¯Î±", 120),
            ("active", "âœ… Î•Î½ÎµÏÎ³ÏŒÏ‚", 80)
        ]
        
        for col, text, width in columns_config:
            self.tree_purposes.heading(col, text=text)
            self.tree_purposes.column(col, width=width, anchor="center")
        
        # Add scrollbars
        v_scrollbar_p = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_purposes.yview)
        h_scrollbar_p = ttk.Scrollbar(tree_container, orient="horizontal", command=self.tree_purposes.xview)
        self.tree_purposes.configure(yscrollcommand=v_scrollbar_p.set, xscrollcommand=h_scrollbar_p.set)
        
        self.tree_purposes.pack(side="left", fill="both", expand=True)
        v_scrollbar_p.pack(side="right", fill="y")
        h_scrollbar_p.pack(side="bottom", fill="x")
        
        # Bind events
        self.tree_purposes.bind("<Double-1>", self._select_purpose_from_tree)

        # Pack canvas and scrollbar
        main_canvas.pack(side="left", fill="both", expand=True, padx=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10))

        # Load initial data
        self._load_purposes()
    
    def _build_reports_tab(self):
        """Build improved reports tab"""
        # Main container
        main_frame = tk.Frame(self.tab_reports, bg=THEMES[self.current_theme]["bg"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_frame, text="ğŸ“Š Î‘Î½Î±Ï†Î¿ÏÎ­Ï‚ & Î•Î¾Î±Î³Ï‰Î³Î­Ï‚", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack(pady=(0, 20))
        
        # Monthly Reports Section
        monthly_frame = ModernFrame(main_frame, bg=THEMES[self.current_theme]["frame_bg"])
        monthly_frame.pack(fill="x", pady=(0, 20))
        
        monthly_title = tk.Label(monthly_frame, text="ğŸ“… ÎœÎ·Î½Î¹Î±Î¯ÎµÏ‚ Î‘Î½Î±Ï†Î¿ÏÎ­Ï‚", 
                                font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                                bg=THEMES[self.current_theme]["frame_bg"])
        monthly_title.pack(pady=(15, 15))
        
        # Month/Year selection
        date_frame = tk.Frame(monthly_frame, bg=THEMES[self.current_theme]["frame_bg"])
        date_frame.pack(pady=10)
        
        tk.Label(date_frame, text="ÎœÎ®Î½Î±Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0, padx=5)
        self.month_var = tk.StringVar(value=str(datetime.now().month))
        month_combo = ttk.Combobox(date_frame, textvariable=self.month_var, values=list(range(1, 13)), 
                                  width=10, state="readonly")
        month_combo.grid(row=0, column=1, padx=5)
        
        tk.Label(date_frame, text="ÎˆÏ„Î¿Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=2, padx=5)
        self.year_var = tk.StringVar(value=str(datetime.now().year))
        year_combo = ttk.Combobox(date_frame, textvariable=self.year_var, 
                                 values=list(range(2020, datetime.now().year + 5)), 
                                 width=10, state="readonly")
        year_combo.grid(row=0, column=3, padx=5)
        
        # Monthly export buttons
        btn_frame_monthly = tk.Frame(monthly_frame, bg=THEMES[self.current_theme]["frame_bg"])
        btn_frame_monthly.pack(pady=15)
        
        ModernButton(btn_frame_monthly, text="ğŸ“„ Î•Î¾Î±Î³Ï‰Î³Î® CSV", style="primary", 
                    command=self._export_monthly_report_csv).pack(side="left", padx=5)
        
        if EXCEL_AVAILABLE:
            ModernButton(btn_frame_monthly, text="ğŸ“Š Î•Î¾Î±Î³Ï‰Î³Î® Excel", style="success", 
                        command=self._export_monthly_report_excel).pack(side="left", padx=5)
        
        # General Reports Section
        general_frame = ModernFrame(main_frame, bg=THEMES[self.current_theme]["frame_bg"])
        general_frame.pack(fill="x", pady=(0, 20))
        
        general_title = tk.Label(general_frame, text="ğŸ“‹ Î“ÎµÎ½Î¹ÎºÎ­Ï‚ Î•Î¾Î±Î³Ï‰Î³Î­Ï‚", 
                                font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                                bg=THEMES[self.current_theme]["frame_bg"])
        general_title.pack(pady=(15, 15))
        
        btn_frame_general = tk.Frame(general_frame, bg=THEMES[self.current_theme]["frame_bg"])
        btn_frame_general.pack(pady=10)
        
        ModernButton(btn_frame_general, text="ğŸš— Î•Î¾Î±Î³Ï‰Î³Î® ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½", style="primary", 
                    command=self._export_movements_csv).pack(side="left", padx=5)
        
        ModernButton(btn_frame_general, text="â›½ Î•Î¾Î±Î³Ï‰Î³Î® ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½", style="primary", 
                    command=self._export_fuel_csv).pack(side="left", padx=5)
        
        if EXCEL_AVAILABLE:
            ModernButton(btn_frame_general, text="ğŸ“Š Î“ÎµÎ½Î¹ÎºÎ® Î‘Î½Î±Ï†Î¿ÏÎ¬ Excel", style="success", 
                        command=self._export_excel).pack(side="left", padx=5)
        
        # Statistics Section
        stats_frame = ModernFrame(main_frame, bg=THEMES[self.current_theme]["frame_bg"])
        stats_frame.pack(fill="both", expand=True)
        
        stats_title = tk.Label(stats_frame, text="ğŸ“ˆ Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÎ¬", 
                              font=FONT_SUBTITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["frame_bg"])
        stats_title.pack(pady=(15, 15))
        
        # Text widget for statistics
        stats_container = tk.Frame(stats_frame, bg=THEMES[self.current_theme]["frame_bg"])
        stats_container.pack(fill="both", expand=True, padx=15, pady=10)
        
        self.stats_text = tk.Text(stats_container, height=12, font=FONT_SMALL, wrap=tk.WORD,
                                 bg=THEMES[self.current_theme]["entry_bg"], 
                                 fg=THEMES[self.current_theme]["fg"],
                                 relief="flat", borderwidth=1)
        self.stats_text.pack(side="left", fill="both", expand=True)
        
        # Add scrollbar for stats
        stats_scrollbar = ttk.Scrollbar(stats_container, orient="vertical", command=self.stats_text.yview)
        self.stats_text.configure(yscrollcommand=stats_scrollbar.set)
        stats_scrollbar.pack(side="right", fill="y")
        
        # Refresh button
        refresh_btn_frame = tk.Frame(stats_frame, bg=THEMES[self.current_theme]["frame_bg"])
        refresh_btn_frame.pack(pady=10)
        ModernButton(refresh_btn_frame, text="ğŸ”„ Î‘Î½Î±Î½Î­Ï‰ÏƒÎ· Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÏÎ½", style="info", 
                    command=self._update_statistics).pack()
        
        self._update_statistics()
    
    # Vehicle management methods
    def _add_vehicle(self):
        """Add a new vehicle with improved validation"""
        try:
            plate = normalize_plate(self.ent_plate.get())
            brand = self.ent_brand.get().strip()
            vtype = self.ent_vtype.get().strip()
            purpose = self.ent_vpurpose.get().strip() if hasattr(self, 'ent_vpurpose') else ""
            photo_path = self.photo_path_var.get()
            
            # Validate required fields
            required_fields = {
                "Î Î¹Î½Î±ÎºÎ¯Î´Î±": plate
            }
            
            is_valid, error_msg = self.validate_required_fields(required_fields)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î± Î•Ï€Î¹ÎºÏÏÏ‰ÏƒÎ·Ï‚", error_msg)
                return
            
            # Validate plate format
            is_valid, validated_plate = DataValidator.is_valid_plate(plate)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", validated_plate)
                return
            
            self.db.cursor.execute("""
                INSERT INTO vehicles (plate, brand, vtype, purpose, photo_path)
                VALUES (?, ?, ?, ?, ?)
            """, (validated_plate, brand, vtype, purpose, photo_path))
            
            self.db.conn.commit()
            self._clear_vehicle_form()
            self._load_vehicles()
            self._refresh_movement_combos()
            self._load_fuel_combos()  # Update fuel combos
            
            self.status_bar.set_status(f"ÎŒÏ‡Î·Î¼Î± {validated_plate} Ï€ÏÎ¿ÏƒÏ„Î­Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
            log_user_action("Vehicle added", f"Plate: {validated_plate}")
            messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "Î¤Î¿ ÏŒÏ‡Î·Î¼Î± Ï€ÏÎ¿ÏƒÏ„Î­Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
            
        except sqlite3.IntegrityError:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î— Ï€Î¹Î½Î±ÎºÎ¯Î´Î± Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Î®Î´Î·!")
        except Exception as e:
            logging.error(f"Error adding vehicle: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Ï€ÏÎ¿ÏƒÎ¸Î®ÎºÎ·: {str(e)}")

    def _update_vehicle(self):
        """Update selected vehicle"""
        selection = self.tree_vehicles.selection()
        if not selection:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ ÏŒÏ‡Î·Î¼Î±!")
            return
        
        try:
            item = selection[0]
            old_plate = self.tree_vehicles.item(item, "values")[0]
            
            plate = normalize_plate(self.ent_plate.get())
            brand = self.ent_brand.get().strip()
            vtype = self.ent_vtype.get().strip()
            purpose = self.ent_vpurpose.get().strip() if hasattr(self, 'ent_vpurpose') else ""
            photo_path = self.photo_path_var.get()
            
            # Validate plate
            is_valid, validated_plate = DataValidator.is_valid_plate(plate)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", validated_plate)
                return
            
            self.db.cursor.execute("""
                UPDATE vehicles 
                SET plate = ?, brand = ?, vtype = ?, purpose = ?, photo_path = ?
                WHERE plate = ?
            """, (validated_plate, brand, vtype, purpose, photo_path, old_plate))
            
            self.db.conn.commit()
            self._clear_vehicle_form()
            self._load_vehicles()
            self._refresh_movement_combos()
            
            self.status_bar.set_status(f"ÎŒÏ‡Î·Î¼Î± {validated_plate} ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
            log_user_action("Vehicle updated", f"Plate: {validated_plate}")
            messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "Î¤Î¿ ÏŒÏ‡Î·Î¼Î± ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
            
        except sqlite3.IntegrityError:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î— Ï€Î¹Î½Î±ÎºÎ¯Î´Î± Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Î®Î´Î·!")
        except Exception as e:
            logging.error(f"Error updating vehicle: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ·: {str(e)}")

    def _delete_vehicle(self):
        """Delete selected vehicle"""
        selection = self.tree_vehicles.selection()
        if not selection:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ ÏŒÏ‡Î·Î¼Î±!")
            return
        
        item = selection[0]
        plate = self.tree_vehicles.item(item, "values")[0]
        
        if ConfirmDialog(
            self.root,
            "ğŸ—‘ï¸ Î•Ï€Î¹Î²ÎµÎ²Î±Î¯Ï‰ÏƒÎ· Î”Î¹Î±Î³ÏÎ±Ï†Î®Ï‚",
            f"Î•Î¯ÏƒÏ„Îµ ÏƒÎ¯Î³Î¿Ï…ÏÎ¿Ï‚ ÏŒÏ„Î¹ Î¸Î­Î»ÎµÏ„Îµ Î½Î± Î´Î¹Î±Î³ÏÎ¬ÏˆÎµÏ„Îµ Ï„Î¿ ÏŒÏ‡Î·Î¼Î± {plate};\n\nÎ‘Ï…Ï„Î® Î· ÎµÎ½Î­ÏÎ³ÎµÎ¹Î± Î¸Î± Î´Î¹Î±Î³ÏÎ¬ÏˆÎµÎ¹ ÎºÎ±Î¹ ÏŒÎ»ÎµÏ‚ Ï„Î¹Ï‚ ÏƒÏ‡ÎµÏ„Î¹ÎºÎ­Ï‚ ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ ÎºÎ±Î¹ ÎºÎ±ÏÏƒÎ¹Î¼Î±!"
        ).show():
            try:
                self.db.cursor.execute("DELETE FROM vehicles WHERE plate = ?", (plate,))
                self.db.conn.commit()
                self._clear_vehicle_form()
                self._load_vehicles()
                self._refresh_movement_combos()
                
                self.status_bar.set_status(f"ÎŒÏ‡Î·Î¼Î± {plate} Î´Î¹Î±Î³ÏÎ¬Ï†Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                log_user_action("Vehicle deleted", f"Plate: {plate}")
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "Î¤Î¿ ÏŒÏ‡Î·Î¼Î± Î´Î¹Î±Î³ÏÎ¬Ï†Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
                
            except Exception as e:
                logging.error(f"Error deleting vehicle: {e}")
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î´Î¹Î±Î³ÏÎ±Ï†Î®: {str(e)}")

    def _clear_vehicle_form(self):
        """Clear vehicle form"""
        self.ent_plate.delete(0, tk.END)
        self.ent_brand.delete(0, tk.END)
        self.ent_vtype.delete(0, tk.END)
        if hasattr(self, 'ent_vpurpose'):
            self.ent_vpurpose.set("")
        self.photo_path_var.set("")
        self.photo_label.config(text="Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ ÎµÏ€Î¹Î»ÎµÎ³ÎµÎ¯ Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±")

    def _select_vehicle_from_tree(self, event=None):
        """Select vehicle from tree"""
        selection = self.tree_vehicles.selection()
        if selection:
            item = selection[0]
            values = self.tree_vehicles.item(item, "values")
            
            self.ent_plate.delete(0, tk.END)
            self.ent_plate.insert(0, values[0])
            self.ent_brand.delete(0, tk.END)
            self.ent_brand.insert(0, values[1])
            self.ent_vtype.delete(0, tk.END)
            self.ent_vtype.insert(0, values[2])
            if hasattr(self, 'ent_vpurpose'):
                self.ent_vpurpose.set(values[3])
            
            # Get photo path
            try:
                self.db.cursor.execute("SELECT photo_path FROM vehicles WHERE plate = ?", (values[0],))
                result = self.db.cursor.fetchone()
                if result and result[0]:
                    self.photo_path_var.set(result[0])
                    self.photo_label.config(text=os.path.basename(result[0]))
                else:
                    self.photo_path_var.set("")
                    self.photo_label.config(text="Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ ÎµÏ€Î¹Î»ÎµÎ³ÎµÎ¯ Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±")
            except Exception as e:
                logging.error(f"Error getting photo path: {e}")

    def _load_vehicles(self):
        """Load vehicles into tree with search filtering"""
        for item in self.tree_vehicles.get_children():
            self.tree_vehicles.delete(item)
        
        search_term = self.vehicles_search_var.get().upper() if hasattr(self, 'vehicles_search_var') else ""
        
        try:
            query = "SELECT plate, brand, vtype, purpose FROM vehicles"
            params = ()
            
            if search_term:
                query += " WHERE UPPER(plate) LIKE ?"
                params = (f"%{search_term}%",)
            
            query += " ORDER BY plate"
            
            self.db.cursor.execute(query, params)
            for row in self.db.cursor.fetchall():
                self.tree_vehicles.insert("", "end", values=row)
                
        except Exception as e:
            logging.error(f"Error loading vehicles: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· Î¿Ï‡Î·Î¼Î¬Ï„Ï‰Î½: {str(e)}")

    def _select_photo(self):
        """Select photo for vehicle"""
        file_path = filedialog.askopenfilename(
            title="Î•Ï€Î¹Î»Î¿Î³Î® Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±Ï‚",
            filetypes=[("Image files", "*.jpg *.jpeg *.png *.bmp *.gif")]
        )
        
        if file_path:
            # Copy file to images directory
            ensure_dir(IMAGES_DIR)
            filename = os.path.basename(file_path)
            dest_path = os.path.join(IMAGES_DIR, filename)
            
            try:
                import shutil
                shutil.copy2(file_path, dest_path)
                self.photo_path_var.set(dest_path)
                self.photo_label.config(text=filename)
                self.status_bar.set_status("Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î± ÎµÏ€Î¹Î»Î­Ï‡Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
            except Exception as e:
                logging.error(f"Error copying photo: {e}")
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î±Î½Ï„Î¹Î³ÏÎ±Ï†Î®: {str(e)}")

    def _view_photo(self):
        """View selected photo"""
        photo_path = self.photo_path_var.get()
        if not photo_path or not os.path.exists(photo_path):
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±!")
            return
        
        try:
            if PIL_AVAILABLE:
                # Create photo viewer window
                photo_window = tk.Toplevel(self.root)
                photo_window.title("Î ÏÎ¿Î²Î¿Î»Î® Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±Ï‚")
                photo_window.configure(bg=THEMES[self.current_theme]["bg"])
                
                # Load and resize image
                from PIL import Image, ImageTk
                pil_image = Image.open(photo_path)
                pil_image.thumbnail((600, 400), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(pil_image)
                
                label = tk.Label(photo_window, image=photo, bg=THEMES[self.current_theme]["bg"])
                label.image = photo  # Keep a reference
                label.pack(padx=10, pady=10)
            else:
                messagebox.showinfo("ğŸ“· Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±", f"Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±: {photo_path}")
        except Exception as e:
            logging.error(f"Error viewing photo: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Ï€ÏÎ¿Î²Î¿Î»Î®: {str(e)}")
    
    # Driver management methods
    def _add_driver(self):
        """Add a new driver with improved validation"""
        try:
            name = normalize_name(self.ent_name.get())
            surname = normalize_name(self.ent_surname.get())
            notes = self.ent_dnotes.get().strip()
            
            # Validate required fields
            required_fields = {
                "ÎŒÎ½Î¿Î¼Î±": name,
                "Î•Ï€ÏÎ½Ï…Î¼Î¿": surname
            }
            
            is_valid, error_msg = self.validate_required_fields(required_fields)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î± Î•Ï€Î¹ÎºÏÏÏ‰ÏƒÎ·Ï‚", error_msg)
                return
            
            # Validate names
            is_valid, validated_name = DataValidator.is_valid_name(name)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"ÎŒÎ½Î¿Î¼Î±: {validated_name}")
                return
                
            is_valid, validated_surname = DataValidator.is_valid_name(surname)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î•Ï€ÏÎ½Ï…Î¼Î¿: {validated_surname}")
                return
            
            self.db.cursor.execute("""
                INSERT INTO drivers (name, surname, notes)
                VALUES (?, ?, ?)
            """, (validated_name, validated_surname, notes))
            
            self.db.conn.commit()
            self._clear_driver_form()
            self._load_drivers()
            self._refresh_movement_combos()
            self._load_fuel_combos()  # Update fuel combos
            
            self.status_bar.set_status(f"ÎŸÎ´Î·Î³ÏŒÏ‚ {validated_name} {validated_surname} Ï€ÏÎ¿ÏƒÏ„Î­Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
            log_user_action("Driver added", f"Name: {validated_name} {validated_surname}")
            messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "ÎŸ Î¿Î´Î·Î³ÏŒÏ‚ Ï€ÏÎ¿ÏƒÏ„Î­Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
            
        except Exception as e:
            logging.error(f"Error adding driver: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Ï€ÏÎ¿ÏƒÎ¸Î®ÎºÎ·: {str(e)}")

    def _update_driver(self):
        """Update selected driver"""
        selection = self.tree_drivers.selection()
        if not selection:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î¿Î´Î·Î³ÏŒ!")
            return
        
        try:
            item = selection[0]
            old_values = self.tree_drivers.item(item, "values")
            
            name = normalize_name(self.ent_name.get())
            surname = normalize_name(self.ent_surname.get())
            notes = self.ent_dnotes.get().strip()
            
            # Validate names
            is_valid, validated_name = DataValidator.is_valid_name(name)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"ÎŒÎ½Î¿Î¼Î±: {validated_name}")
                return
                
            is_valid, validated_surname = DataValidator.is_valid_name(surname)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î•Ï€ÏÎ½Ï…Î¼Î¿: {validated_surname}")
                return
            
            self.db.cursor.execute("""
                UPDATE drivers 
                SET name = ?, surname = ?, notes = ?
                WHERE name = ? AND surname = ?
            """, (validated_name, validated_surname, notes, old_values[0], old_values[1]))
            
            self.db.conn.commit()
            self._clear_driver_form()
            self._load_drivers()
            self._refresh_movement_combos()
            
            self.status_bar.set_status(f"ÎŸÎ´Î·Î³ÏŒÏ‚ {validated_name} {validated_surname} ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
            log_user_action("Driver updated", f"Name: {validated_name} {validated_surname}")
            messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "ÎŸ Î¿Î´Î·Î³ÏŒÏ‚ ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
            
        except Exception as e:
            logging.error(f"Error updating driver: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ·: {str(e)}")

    def _delete_driver(self):
        """Delete selected driver"""
        selection = self.tree_drivers.selection()
        if not selection:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î¿Î´Î·Î³ÏŒ!")
            return
        
        item = selection[0]
        values = self.tree_drivers.item(item, "values")
        driver_name = f"{values[0]} {values[1]}"
        
        if ConfirmDialog(
            self.root,
            "ğŸ—‘ï¸ Î•Ï€Î¹Î²ÎµÎ²Î±Î¯Ï‰ÏƒÎ· Î”Î¹Î±Î³ÏÎ±Ï†Î®Ï‚",
            f"Î•Î¯ÏƒÏ„Îµ ÏƒÎ¯Î³Î¿Ï…ÏÎ¿Ï‚ ÏŒÏ„Î¹ Î¸Î­Î»ÎµÏ„Îµ Î½Î± Î´Î¹Î±Î³ÏÎ¬ÏˆÎµÏ„Îµ Ï„Î¿Î½ Î¿Î´Î·Î³ÏŒ {driver_name};\n\nÎ‘Ï…Ï„Î® Î· ÎµÎ½Î­ÏÎ³ÎµÎ¹Î± Î¸Î± Î´Î¹Î±Î³ÏÎ¬ÏˆÎµÎ¹ ÎºÎ±Î¹ ÏŒÎ»ÎµÏ‚ Ï„Î¹Ï‚ ÏƒÏ‡ÎµÏ„Î¹ÎºÎ­Ï‚ ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ ÎºÎ±Î¹ ÎºÎ±ÏÏƒÎ¹Î¼Î±!"
        ).show():
            try:
                self.db.cursor.execute("DELETE FROM drivers WHERE name = ? AND surname = ?", (values[0], values[1]))
                self.db.conn.commit()
                self._clear_driver_form()
                self._load_drivers()
                self._refresh_movement_combos()
                
                self.status_bar.set_status(f"ÎŸÎ´Î·Î³ÏŒÏ‚ {driver_name} Î´Î¹Î±Î³ÏÎ¬Ï†Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                log_user_action("Driver deleted", f"Name: {driver_name}")
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "ÎŸ Î¿Î´Î·Î³ÏŒÏ‚ Î´Î¹Î±Î³ÏÎ¬Ï†Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
                
            except Exception as e:
                logging.error(f"Error deleting driver: {e}")
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î´Î¹Î±Î³ÏÎ±Ï†Î®: {str(e)}")

    def _clear_driver_form(self):
        """Clear driver form"""
        self.ent_name.delete(0, tk.END)
        self.ent_surname.delete(0, tk.END)
        self.ent_dnotes.delete(0, tk.END)

    def _select_driver_from_tree(self, event=None):
        """Select driver from tree"""
        selection = self.tree_drivers.selection()
        if selection:
            item = selection[0]
            values = self.tree_drivers.item(item, "values")
            
            self.ent_name.delete(0, tk.END)
            self.ent_name.insert(0, values[0])
            self.ent_surname.delete(0, tk.END)
            self.ent_surname.insert(0, values[1])
            self.ent_dnotes.delete(0, tk.END)
            self.ent_dnotes.insert(0, values[2])

    def _load_drivers(self):
        """Load drivers into tree with search filtering"""
        for item in self.tree_drivers.get_children():
            self.tree_drivers.delete(item)
        
        search_term = self.drivers_search_var.get().upper() if hasattr(self, 'drivers_search_var') else ""
        
        try:
            query = "SELECT name, surname, notes FROM drivers"
            params = ()
            
            if search_term:
                query += " WHERE UPPER(name || ' ' || surname) LIKE ?"
                params = (f"%{search_term}%",)
            
            query += " ORDER BY name, surname"
            
            self.db.cursor.execute(query, params)
            for row in self.db.cursor.fetchall():
                self.tree_drivers.insert("", "end", values=row)
                
        except Exception as e:
            logging.error(f"Error loading drivers: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· Î¿Î´Î·Î³ÏÎ½: {str(e)}")
    
    # Fuel management methods
    def _add_fuel(self):
        """Add fuel record with improved validation"""
        try:
            date = self.fuel_date_var.get()
            driver = self.fuel_driver_combo.get()
            vehicle = self.fuel_vehicle_combo.get()
            liters_str = self.fuel_liters_var.get()
            mileage_str = self.fuel_mileage_var.get()
            cost_str = self.fuel_cost_var.get()
            pump_reading_str = self.fuel_pump_reading_var.get()
            
            # Validate required fields
            required_fields = {
                "Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±": date,
                "ÎŸÎ´Î·Î³ÏŒÏ‚": driver,
                "ÎŒÏ‡Î·Î¼Î±": vehicle,
                "Î›Î¯Ï„ÏÎ±": liters_str
            }
            
            is_valid, error_msg = self.validate_required_fields(required_fields)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î± Î•Ï€Î¹ÎºÏÏÏ‰ÏƒÎ·Ï‚", error_msg)
                return
            
            # Validate date
            if not validate_date(date):
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "ÎœÎ· Î­Î³ÎºÏ…ÏÎ· Î·Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±! Î§ÏÎ·ÏƒÎ¹Î¼Î¿Ï€Î¿Î¹Î®ÏƒÏ„Îµ Ï„Î· Î¼Î¿ÏÏ†Î® YYYY-MM-DD")
                return
            
            # Validate fuel amount
            is_valid, liters = DataValidator.is_valid_fuel(liters_str)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", liters)
                return
            
            # Validate mileage if provided
            mileage = None
            if mileage_str:
                is_valid, mileage = DataValidator.is_valid_km(mileage_str)
                if not is_valid:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", mileage)
                    return
            
            # Validate cost if provided
            cost = 0.0
            if cost_str:
                try:
                    cost = float(cost_str)
                    if cost < 0:
                        messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î¤Î¿ ÎºÏŒÏƒÏ„Î¿Ï‚ Î´ÎµÎ½ Î¼Ï€Î¿ÏÎµÎ¯ Î½Î± ÎµÎ¯Î½Î±Î¹ Î±ÏÎ½Î·Ï„Î¹ÎºÏŒ!")
                        return
                except ValueError:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î¤Î¿ ÎºÏŒÏƒÏ„Î¿Ï‚ Ï€ÏÎ­Ï€ÎµÎ¹ Î½Î± ÎµÎ¯Î½Î±Î¹ Î­Î³ÎºÏ…ÏÎ¿Ï‚ Î±ÏÎ¹Î¸Î¼ÏŒÏ‚!")
                    return
            
            # Get IDs
            driver_id = self._get_driver_id(driver)
            vehicle_id = self._get_vehicle_id(vehicle)
            
            if not driver_id:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "ÎœÎ· Î­Î³ÎºÏ…ÏÎ¿Ï‚ Î¿Î´Î·Î³ÏŒÏ‚!")
                return
            
            if not vehicle_id:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "ÎœÎ· Î­Î³ÎºÏ…ÏÎ¿ ÏŒÏ‡Î·Î¼Î±!")
                return
            
            # Check tank level
            self.db.cursor.execute("SELECT SUM(CASE WHEN type = 'fill' THEN liters ELSE -liters END) FROM tank")
            current_level = self.db.cursor.fetchone()[0] or 0
            
            if current_level < liters:
                if not ConfirmDialog(
                    self.root,
                    "âš ï¸ Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ· Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚",
                    f"Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡Î¿Ï…Î½ Î±ÏÎºÎµÏ„Î¬ ÎºÎ±ÏÏƒÎ¹Î¼Î± ÏƒÏ„Î· Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®!\n\nÎ”Î¹Î±Î¸Î­ÏƒÎ¹Î¼Î±: {current_level:.1f}L\nÎ‘Ï€Î±Î¹Ï„Î¿ÏÎ½Ï„Î±Î¹: {liters:.1f}L\n\nÎ˜Î­Î»ÎµÏ„Îµ Î½Î± ÏƒÏ…Î½ÎµÏ‡Î¯ÏƒÎµÏ„Îµ;"
                ).show():
                    return
            
            # Add fuel record
            self.db.cursor.execute("""
                INSERT INTO fuel (vehicle_id, driver_id, date, liters, mileage, cost)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (vehicle_id, driver_id, date, liters, mileage, cost))
            
            # Update tank
            self.db.cursor.execute("""
                INSERT INTO tank (date, liters, type, notes)
                VALUES (?, ?, 'consume', ?)
            """, (date, liters, f"Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚ {vehicle}"))
            
            # Update pump reading if provided
            if pump_reading_str:
                try:
                    new_pump_reading = float(pump_reading_str)
                    current_pump_reading = self.db.get_pump_reading()
                    
                    if new_pump_reading >= current_pump_reading:
                        calculated_liters = new_pump_reading - current_pump_reading
                        
                        # Check if pump liters match fuel liters (with small tolerance)
                        if abs(calculated_liters - liters) <= 1.0:  # 1L tolerance
                            self.db.update_pump_reading(
                                new_pump_reading, 
                                calculated_liters,
                                vehicle, 
                                driver,
                                f"Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚ {vehicle} - {liters:.1f}L"
                            )
                        else:
                            logging.warning(f"Pump reading mismatch: pump={calculated_liters:.1f}L, fuel={liters:.1f}L")
                    else:
                        logging.warning(f"New pump reading ({new_pump_reading}) is less than current ({current_pump_reading})")
                        
                except ValueError:
                    logging.warning(f"Invalid pump reading: {pump_reading_str}")
            
            self.db.conn.commit()
            self._clear_fuel_form()
            self._load_fuel()
            self._update_tank_level()
            
            self.status_bar.set_status(f"ÎšÎ±ÏÏƒÎ¹Î¼Î¿ ÎºÎ±Ï„Î±Ï‡Ï‰ÏÎ®Î¸Î·ÎºÎµ: {liters:.1f}L Î³Î¹Î± {vehicle}")
            log_user_action("Fuel added", f"Vehicle: {vehicle}, Liters: {liters:.1f}")
            messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "Î¤Î¿ ÎºÎ±ÏÏƒÎ¹Î¼Î¿ ÎºÎ±Ï„Î±Ï‡Ï‰ÏÎ®Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
            
        except Exception as e:
            logging.error(f"Error adding fuel: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎºÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ·: {str(e)}")

    def _clear_fuel_form(self):
        """Clear fuel form"""
        self.fuel_driver_combo.set("")
        self.fuel_vehicle_combo.set("")
        self.fuel_liters_var.set("")
        self.fuel_mileage_var.set("")
        self.fuel_cost_var.set("")
        self.fuel_pump_reading_var.set("")

    def _refill_tank(self):
        """Refill tank dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("â›½ Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚ Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚")
        dialog.geometry("400x300")
        dialog.resizable(False, False)
        dialog.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main frame
        main_frame = tk.Frame(dialog, bg=THEMES[self.current_theme]["bg"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_frame, text="â›½ Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚ Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack(pady=(0, 20))
        
        # Current level info
        try:
            self.db.cursor.execute(
                "SELECT SUM(CASE WHEN type = 'fill' THEN liters ELSE -liters END) FROM tank"
            )
            current_level = self.db.cursor.fetchone()[0] or 0
            remaining_capacity = TANK_CAPACITY - current_level
            
            info_text = f"ğŸ”µ Î¤ÏÎ­Ï‡Î¿Î½ Î•Ï€Î¯Ï€ÎµÎ´Î¿: {current_level:.1f}L\nğŸ›¢ï¸ Î£Ï…Î½Î¿Î»Î¹ÎºÎ® Î§Ï‰ÏÎ·Ï„Î¹ÎºÏŒÏ„Î·Ï„Î±: {TANK_CAPACITY:,.0f}L\nâœ¨ Î”Î¹Î±Î¸Î­ÏƒÎ¹Î¼Î¿Ï‚ Î§ÏÏÎ¿Ï‚: {remaining_capacity:.1f}L"
            info_label = tk.Label(main_frame, text=info_text, font=FONT_NORMAL,
                                 fg=THEMES[self.current_theme]["fg"],
                                 bg=THEMES[self.current_theme]["bg"], justify="left")
            info_label.pack(pady=(0, 20))
        except:
            current_level = 0
            remaining_capacity = TANK_CAPACITY
        
        # Input frame
        input_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        input_frame.pack(pady=10)
        
        tk.Label(input_frame, text="â›½ Î›Î¯Ï„ÏÎ± Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼Î¿Ï:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        liters_var = tk.StringVar()
        liters_entry = tk.Entry(input_frame, textvariable=liters_var, font=FONT_NORMAL,
                               width=20, relief="flat", borderwidth=1, highlightthickness=1)
        liters_entry.pack(pady=(0, 10))
        liters_entry.focus()
        
        tk.Label(input_frame, text="ğŸ“ Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(10, 5))
        
        notes_var = tk.StringVar(value="Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚ ÎºÎµÎ½Ï„ÏÎ¹ÎºÎ®Ï‚ Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚")
        notes_entry = tk.Entry(input_frame, textvariable=notes_var, font=FONT_NORMAL,
                              width=20, relief="flat", borderwidth=1, highlightthickness=1)
        notes_entry.pack(pady=(0, 20))
        
        def confirm_refill():
            try:
                liters_str = liters_var.get()
                notes = notes_var.get() or "Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚ Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚"
                
                if not liters_str:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÎ¹ÏƒÎ¬Î³ÎµÏ„Îµ Ï„Î± Î»Î¯Ï„ÏÎ±!")
                    return
                
                liters = float(liters_str)
                if liters <= 0:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î¤Î± Î»Î¯Ï„ÏÎ± Ï€ÏÎ­Ï€ÎµÎ¹ Î½Î± ÎµÎ¯Î½Î±Î¹ Î¸ÎµÏ„Î¹ÎºÏŒÏ‚ Î±ÏÎ¹Î¸Î¼ÏŒÏ‚!")
                    return
                
                if liters > remaining_capacity:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", 
                                       f"Î”ÎµÎ½ Ï‡Ï‰ÏÎ¬Î½Îµ {liters:.1f}L!\nÎœÎ­Î³Î¹ÏƒÏ„Î± Î´Î¹Î±Î¸Î­ÏƒÎ¹Î¼Î±: {remaining_capacity:.1f}L")
                    return
                
                # Add refill to tank
                self.db.cursor.execute("""
                    INSERT INTO tank (date, liters, type, notes)
                    VALUES (?, ?, 'fill', ?)
                """, (datetime.now().strftime("%Y-%m-%d"), liters, notes))
                
                self.db.conn.commit()
                self._update_tank_level()
                
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"Î— Î´ÎµÎ¾Î±Î¼ÎµÎ½Î® Î±Î½ÎµÏ†Î¿Î´Î¹Î¬ÏƒÏ„Î·ÎºÎµ Î¼Îµ {liters:.1f}L!")
                dialog.destroy()
                
            except ValueError:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÎ¹ÏƒÎ¬Î³ÎµÏ„Îµ Î­Î³ÎºÏ…ÏÎ¿ Î±ÏÎ¹Î¸Î¼ÏŒ Î»Î¯Ï„ÏÏ‰Î½!")
            except Exception as e:
                logging.error(f"Error refilling tank: {e}")
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î¿Î½ Î±Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒ: {str(e)}")
        
        # Buttons
        btn_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        btn_frame.pack(pady=20)
        
        ModernButton(btn_frame, text="âœ… Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚", style="success", 
                    command=confirm_refill).pack(side="left", padx=5)
        ModernButton(btn_frame, text="âŒ Î‘ÎºÏÏÏ‰ÏƒÎ·", style="danger", 
                    command=dialog.destroy).pack(side="left", padx=5)
        
        # Enter key binding
        dialog.bind('<Return>', lambda e: confirm_refill())

    def _show_tank_history(self):
        """Show tank history"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚")
        dialog.geometry("700x500")
        dialog.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main frame
        main_frame = tk.Frame(dialog, bg=THEMES[self.current_theme]["bg"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_frame, text="ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½ Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack(pady=(0, 20))
        
        # Tree for history
        tree_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        tree_frame.pack(fill="both", expand=True)
        
        tree = ttk.Treeview(tree_frame, columns=("date", "type", "liters", "notes"), 
                           show="headings", height=15)
        
        # Configure columns
        tree.heading("date", text="ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±")
        tree.heading("type", text="ğŸ”„ Î¤ÏÏ€Î¿Ï‚")
        tree.heading("liters", text="â›½ Î›Î¯Ï„ÏÎ±")
        tree.heading("notes", text="ğŸ“ Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚")
        
        tree.column("date", width=120, anchor="center")
        tree.column("type", width=100, anchor="center")
        tree.column("liters", width=100, anchor="center")
        tree.column("notes", width=300, anchor="w")
        
        # Load data
        try:
            self.db.cursor.execute("""
                SELECT date, type, liters, notes 
                FROM tank 
                ORDER BY date DESC, id DESC
            """)
            
            for row in self.db.cursor.fetchall():
                type_display = "ğŸ“ˆ Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒÏ‚" if row[1] == 'fill' else "ğŸ“‰ ÎšÎ±Ï„Î±Î½Î¬Î»Ï‰ÏƒÎ·"
                tree.insert("", "end", values=(
                    row[0],  # date
                    type_display,  # type
                    f"{row[2]:.1f}L",  # liters
                    row[3] or ""  # notes
                ))
        except Exception as e:
            logging.error(f"Error loading tank history: {e}")
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=v_scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        
        # Close button
        ModernButton(main_frame, text="âŒ ÎšÎ»ÎµÎ¯ÏƒÎ¹Î¼Î¿", style="secondary", 
                    command=dialog.destroy).pack(pady=20)

    def _get_current_pump_reading(self):
        """Get current pump reading and populate field"""
        try:
            current_reading = self.db.get_pump_reading()
            self.fuel_pump_reading_var.set(str(current_reading))
        except Exception as e:
            logging.error(f"Error getting pump reading: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î±Î½Î¬ÎºÏ„Î·ÏƒÎ· Î¼Î­Ï„ÏÎ·ÏƒÎ·Ï‚: {str(e)}")

    def _update_pump_reading(self):
        """Update pump reading dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ğŸ“Š Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ· ÎœÎµÏ„ÏÎ·Ï„Î® Î‘Î½Ï„Î»Î¯Î±Ï‚")
        dialog.geometry("450x350")
        dialog.resizable(False, False)
        dialog.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main frame
        main_frame = tk.Frame(dialog, bg=THEMES[self.current_theme]["bg"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_frame, text="ğŸ“Š Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ· ÎœÎµÏ„ÏÎ·Ï„Î® Î‘Î½Ï„Î»Î¯Î±Ï‚", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack(pady=(0, 20))
        
        # Current reading info
        try:
            current_reading = self.db.get_pump_reading()
            info_text = f"ğŸ”µ Î¤ÏÎ­Ï‡Î¿Ï…ÏƒÎ± ÎœÎ­Ï„ÏÎ·ÏƒÎ·: {current_reading:,.0f} Î»Î¯Ï„ÏÎ±"
            info_label = tk.Label(main_frame, text=info_text, font=FONT_NORMAL,
                                 fg=THEMES[self.current_theme]["fg"],
                                 bg=THEMES[self.current_theme]["bg"])
            info_label.pack(pady=(0, 20))
        except:
            current_reading = 0
        
        # Input frame
        input_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        input_frame.pack(pady=10)
        
        tk.Label(input_frame, text="ğŸ“Š ÎÎ­Î± ÎœÎ­Ï„ÏÎ·ÏƒÎ·:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(0, 5))
        
        new_reading_var = tk.StringVar()
        new_reading_entry = tk.Entry(input_frame, textvariable=new_reading_var, font=FONT_NORMAL,
                                    width=20, relief="flat", borderwidth=1, highlightthickness=1)
        new_reading_entry.pack(pady=(0, 10))
        new_reading_entry.focus()
        
        tk.Label(input_frame, text="ğŸ“ Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚:", font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"], 
                fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", pady=(10, 5))
        
        notes_var = tk.StringVar(value="Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ· Î¼ÎµÏ„ÏÎ·Ï„Î® Î±Î½Ï„Î»Î¯Î±Ï‚")
        notes_entry = tk.Entry(input_frame, textvariable=notes_var, font=FONT_NORMAL,
                              width=20, relief="flat", borderwidth=1, highlightthickness=1)
        notes_entry.pack(pady=(0, 20))
        
        def confirm_update():
            try:
                new_reading_str = new_reading_var.get()
                notes = notes_var.get() or "Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ· Î¼ÎµÏ„ÏÎ·Ï„Î®"
                
                if not new_reading_str:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÎ¹ÏƒÎ¬Î³ÎµÏ„Îµ Ï„Î· Î½Î­Î± Î¼Î­Ï„ÏÎ·ÏƒÎ·!")
                    return
                
                new_reading = float(new_reading_str)
                
                if new_reading < current_reading:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", 
                                       f"Î— Î½Î­Î± Î¼Î­Ï„ÏÎ·ÏƒÎ· ({new_reading:,.0f}) Î´ÎµÎ½ Î¼Ï€Î¿ÏÎµÎ¯ Î½Î± ÎµÎ¯Î½Î±Î¹ Î¼Î¹ÎºÏÏŒÏ„ÎµÏÎ· Î±Ï€ÏŒ Ï„Î·Î½ Ï„ÏÎ­Ï‡Î¿Ï…ÏƒÎ± ({current_reading:,.0f})!")
                    return
                
                liters_dispensed = new_reading - current_reading
                
                if liters_dispensed > 0:
                    # Update pump reading
                    if self.db.update_pump_reading(new_reading, liters_dispensed, notes=notes):
                        self._update_pump_display()
                        messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", 
                                          f"ÎŸ Î¼ÎµÏ„ÏÎ·Ï„Î®Ï‚ ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎµ!\nÎ”Î¹Î±Î½Î¿Î¼Î®: {liters_dispensed:,.1f}L")
                        dialog.destroy()
                    else:
                        messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ·!")
                else:
                    messagebox.showwarning("âš ï¸ Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Î±Î»Î»Î±Î³Î® ÏƒÏ„Î· Î¼Î­Ï„ÏÎ·ÏƒÎ·!")
                
            except ValueError:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÎ¹ÏƒÎ¬Î³ÎµÏ„Îµ Î­Î³ÎºÏ…ÏÎ¿ Î±ÏÎ¹Î¸Î¼ÏŒ!")
            except Exception as e:
                logging.error(f"Error updating pump reading: {e}")
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ·: {str(e)}")
        
        # Buttons
        btn_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        btn_frame.pack(pady=20)
        
        ModernButton(btn_frame, text="âœ… Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ·", style="success", 
                    command=confirm_update).pack(side="left", padx=5)
        ModernButton(btn_frame, text="âŒ Î‘ÎºÏÏÏ‰ÏƒÎ·", style="danger", 
                    command=dialog.destroy).pack(side="left", padx=5)
        
        # Enter key binding
        dialog.bind('<Return>', lambda e: confirm_update())

    def _show_pump_history(self):
        """Show pump history"""
        dialog = tk.Toplevel(self.root)
        dialog.title("ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ Î‘Î½Ï„Î»Î¯Î±Ï‚")
        dialog.geometry("800x500")
        dialog.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Center the dialog
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Main frame
        main_frame = tk.Frame(dialog, bg=THEMES[self.current_theme]["bg"])
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Title
        title_label = tk.Label(main_frame, text="ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½ Î‘Î½Ï„Î»Î¯Î±Ï‚", 
                              font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                              bg=THEMES[self.current_theme]["bg"])
        title_label.pack(pady=(0, 20))
        
        # Tree for history
        tree_frame = tk.Frame(main_frame, bg=THEMES[self.current_theme]["bg"])
        tree_frame.pack(fill="both", expand=True)
        
        tree = ttk.Treeview(tree_frame, columns=("date", "reading", "dispensed", "vehicle", "driver", "notes"), 
                           show="headings", height=15)
        
        # Configure columns
        tree.heading("date", text="ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±")
        tree.heading("reading", text="ğŸ“Š ÎœÎ­Ï„ÏÎ·ÏƒÎ·")
        tree.heading("dispensed", text="â›½ Î”Î¹Î±Î½Î¿Î¼Î®")
        tree.heading("vehicle", text="ğŸš— ÎŒÏ‡Î·Î¼Î±")
        tree.heading("driver", text="ğŸ‘¤ ÎŸÎ´Î·Î³ÏŒÏ‚")
        tree.heading("notes", text="ğŸ“ Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚")
        
        tree.column("date", width=120, anchor="center")
        tree.column("reading", width=100, anchor="center")
        tree.column("dispensed", width=100, anchor="center")
        tree.column("vehicle", width=100, anchor="center")
        tree.column("driver", width=120, anchor="center")
        tree.column("notes", width=200, anchor="w")
        
        # Load data
        try:
            history = self.db.get_pump_history()
            for row in history:
                tree.insert("", "end", values=(
                    row[0],  # date
                    f"{row[1]:,.0f}",  # reading
                    f"{row[2]:.1f}L",  # liters_dispensed
                    row[3] or "",  # vehicle_plate
                    row[4] or "",  # driver_name
                    row[5] or ""  # notes
                ))
        except Exception as e:
            logging.error(f"Error loading pump history: {e}")
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=v_scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        
        # Close button
        ModernButton(main_frame, text="âŒ ÎšÎ»ÎµÎ¯ÏƒÎ¹Î¼Î¿", style="secondary", 
                    command=dialog.destroy).pack(pady=20)

    def _update_pump_display(self):
        """Update pump reading display"""
        try:
            current_reading = self.db.get_pump_reading()
            self.pump_reading_label.config(
                text=f"ğŸ“Š Î¤ÏÎ­Ï‡Î¿Ï…ÏƒÎ± ÎœÎ­Ï„ÏÎ·ÏƒÎ·: {current_reading:,.0f} Î»Î¯Ï„ÏÎ±"
            )
        except Exception as e:
            logging.error(f"Error updating pump display: {e}")
            self.pump_reading_label.config(text="âš ï¸ Î£Ï†Î¬Î»Î¼Î± Ï†ÏŒÏÏ„Ï‰ÏƒÎ·Ï‚ Î¼Î­Ï„ÏÎ·ÏƒÎ·Ï‚")

    def _load_fuel(self):
        """Load fuel records with cost information"""
        for item in self.tree_fuel.get_children():
            self.tree_fuel.delete(item)
        
        try:
            self.db.cursor.execute("""
                SELECT f.date, d.name || ' ' || d.surname, v.plate, f.liters, f.mileage, f.cost
                FROM fuel f
                JOIN drivers d ON f.driver_id = d.id
                JOIN vehicles v ON f.vehicle_id = v.id
                ORDER BY f.date DESC, f.id DESC
            """)
            
            for row in self.db.cursor.fetchall():
                # Format the display values
                display_row = (
                    row[0],  # date
                    row[1],  # driver
                    row[2],  # vehicle
                    f"{row[3]:.1f}",  # liters
                    f"{row[4]:.0f}" if row[4] else "",  # mileage
                    format_currency(row[5]) if row[5] else "0.00 â‚¬"  # cost
                )
                self.tree_fuel.insert("", "end", values=display_row)
            
            # Update tank level display
            self._update_tank_level()
            # Update pump display
            self._update_pump_display()
                
        except Exception as e:
            logging.error(f"Error loading fuel: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· ÎºÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½: {str(e)}")

    def _load_fuel_combos(self):
        """Load drivers and vehicles into fuel comboboxes"""
        try:
            # Load drivers
            self.db.cursor.execute("SELECT name, surname FROM drivers ORDER BY name, surname")
            drivers = [f"{row[0]} {row[1]}" for row in self.db.cursor.fetchall()]
            
            if hasattr(self, 'fuel_driver_combo'):
                self.fuel_driver_combo.set_values(drivers)
            
            # Load vehicles  
            self.db.cursor.execute("SELECT plate FROM vehicles ORDER BY plate")
            vehicles = [row[0] for row in self.db.cursor.fetchall()]
            
            if hasattr(self, 'fuel_vehicle_combo'):
                self.fuel_vehicle_combo.set_values(vehicles)
                
        except Exception as e:
            logging.error(f"Error loading fuel combos: {e}")
    
    # ===== PURPOSE MANAGEMENT METHODS =====
    
    def _load_purposes(self):
        """Load purposes into the tree"""
        try:
            # Clear existing items
            for item in self.tree_purposes.get_children():
                self.tree_purposes.delete(item)
            
            # Get search and filter criteria
            search_term = self.purposes_search_var.get().lower()
            category_filter = self.purpose_filter_var.get()
            show_inactive = self.show_inactive_var.get()
            
            # Get purposes from database
            category = None if category_filter == "all" else category_filter
            purposes = self.db.get_purposes(category=category, active_only=not show_inactive)
            
            for purpose in purposes:
                purpose_id, name, description, category, active = purpose
                
                # Apply search filter
                if search_term and search_term not in name.lower() and search_term not in (description or "").lower():
                    continue
                
                # Format data for display
                active_text = "âœ… ÎÎ±Î¹" if active else "âŒ ÎŒÏ‡Î¹"
                
                self.tree_purposes.insert("", "end", iid=purpose_id, values=(
                    name, description or "", category, active_text
                ))
                
        except Exception as e:
            logging.error(f"Error loading purposes: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· ÏƒÎºÎ¿Ï€ÏÎ½: {e}")
    
    def _add_purpose(self):
        """Add a new purpose"""
        try:
            name = self.ent_purpose_name.get().strip()
            description = self.ent_purpose_description.get().strip()
            category = self.purpose_category_var.get()
            
            if not name:
                messagebox.showwarning("Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÎ¹ÏƒÎ¬Î³ÎµÏ„Îµ ÏŒÎ½Î¿Î¼Î± ÏƒÎºÎ¿Ï€Î¿Ï")
                return
            
            # Add to database
            purpose_id = self.db.add_purpose(name, description, category)
            
            if purpose_id:
                messagebox.showinfo("Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"ÎŸ ÏƒÎºÎ¿Ï€ÏŒÏ‚ '{name}' Ï€ÏÎ¿ÏƒÏ„Î­Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                self._clear_purpose_form()
                self._load_purposes()
                self._update_purpose_combos()  # Update comboboxes in other tabs
            else:
                messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"ÎŸ ÏƒÎºÎ¿Ï€ÏŒÏ‚ '{name}' Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Î®Î´Î·")
                
        except Exception as e:
            logging.error(f"Error adding purpose: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Ï€ÏÎ¿ÏƒÎ¸Î®ÎºÎ· ÏƒÎºÎ¿Ï€Î¿Ï: {e}")
    
    def _update_purpose(self):
        """Update selected purpose"""
        try:
            selected_items = self.tree_purposes.selection()
            if not selected_items:
                messagebox.showwarning("Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î­Î½Î±Î½ ÏƒÎºÎ¿Ï€ÏŒ Î³Î¹Î± ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ·")
                return
            
            purpose_id = selected_items[0]
            name = self.ent_purpose_name.get().strip()
            description = self.ent_purpose_description.get().strip()
            category = self.purpose_category_var.get()
            
            if not name:
                messagebox.showwarning("Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÎ¹ÏƒÎ¬Î³ÎµÏ„Îµ ÏŒÎ½Î¿Î¼Î± ÏƒÎºÎ¿Ï€Î¿Ï")
                return
            
            # Update in database
            if self.db.update_purpose(purpose_id, name=name, description=description, category=category):
                messagebox.showinfo("Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"ÎŸ ÏƒÎºÎ¿Ï€ÏŒÏ‚ '{name}' ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                self._clear_purpose_form()
                self._load_purposes()
                self._update_purpose_combos()  # Update comboboxes in other tabs
            else:
                messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", "Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ· Ï„Î¿Ï… ÏƒÎºÎ¿Ï€Î¿Ï")
                
        except Exception as e:
            logging.error(f"Error updating purpose: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ· ÏƒÎºÎ¿Ï€Î¿Ï: {e}")
    
    def _delete_purpose(self):
        """Deactivate selected purpose"""
        try:
            selected_items = self.tree_purposes.selection()
            if not selected_items:
                messagebox.showwarning("Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î­Î½Î±Î½ ÏƒÎºÎ¿Ï€ÏŒ Î³Î¹Î± Î±Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¯Î·ÏƒÎ·")
                return
            
            purpose_id = selected_items[0]
            purpose_data = self.tree_purposes.item(purpose_id)["values"]
            purpose_name = purpose_data[0]
            
            if messagebox.askyesno("Î•Ï€Î¹Î²ÎµÎ²Î±Î¯Ï‰ÏƒÎ·", f"Î˜Î­Î»ÎµÏ„Îµ Î½Î± Î±Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¹Î®ÏƒÎµÏ„Îµ Ï„Î¿Î½ ÏƒÎºÎ¿Ï€ÏŒ '{purpose_name}';"):
                if self.db.delete_purpose(purpose_id):
                    messagebox.showinfo("Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"ÎŸ ÏƒÎºÎ¿Ï€ÏŒÏ‚ '{purpose_name}' Î±Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¹Î®Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                    self._clear_purpose_form()
                    self._load_purposes()
                    self._update_purpose_combos()  # Update comboboxes in other tabs
                else:
                    messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", "Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î±Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¯Î·ÏƒÎ· Ï„Î¿Ï… ÏƒÎºÎ¿Ï€Î¿Ï")
                    
        except Exception as e:
            logging.error(f"Error deleting purpose: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î±Ï€ÎµÎ½ÎµÏÎ³Î¿Ï€Î¿Î¯Î·ÏƒÎ· ÏƒÎºÎ¿Ï€Î¿Ï: {e}")
    
    def _restore_purpose(self):
        """Restore selected purpose"""
        try:
            selected_items = self.tree_purposes.selection()
            if not selected_items:
                messagebox.showwarning("Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î­Î½Î±Î½ ÏƒÎºÎ¿Ï€ÏŒ Î³Î¹Î± ÎµÏ€Î±Î½Î±Ï†Î¿ÏÎ¬")
                return
            
            purpose_id = selected_items[0]
            purpose_data = self.tree_purposes.item(purpose_id)["values"]
            purpose_name = purpose_data[0]
            
            if self.db.restore_purpose(purpose_id):
                messagebox.showinfo("Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"ÎŸ ÏƒÎºÎ¿Ï€ÏŒÏ‚ '{purpose_name}' ÎµÏ€Î±Î½Î±Ï†Î­ÏÎ¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                self._clear_purpose_form()
                self._load_purposes()
                self._update_purpose_combos()  # Update comboboxes in other tabs
            else:
                messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", "Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÏ€Î±Î½Î±Ï†Î¿ÏÎ¬ Ï„Î¿Ï… ÏƒÎºÎ¿Ï€Î¿Ï")
                
        except Exception as e:
            logging.error(f"Error restoring purpose: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÏ€Î±Î½Î±Ï†Î¿ÏÎ¬ ÏƒÎºÎ¿Ï€Î¿Ï: {e}")
    
    def _clear_purpose_form(self):
        """Clear purpose form fields"""
        self.ent_purpose_name.delete(0, tk.END)
        self.ent_purpose_description.delete(0, tk.END)
        self.purpose_category_var.set("general")
        
        # Clear selection
        for item in self.tree_purposes.selection():
            self.tree_purposes.selection_remove(item)
    
    def _select_purpose_from_tree(self, event=None):
        """Select purpose from tree and populate form"""
        try:
            selected_items = self.tree_purposes.selection()
            if not selected_items:
                return
            
            purpose_id = selected_items[0]
            purpose_data = self.tree_purposes.item(purpose_id)["values"]
            
            # Clear form first
            self._clear_purpose_form()
            
            # Populate form
            self.ent_purpose_name.insert(0, purpose_data[0])  # name
            self.ent_purpose_description.insert(0, purpose_data[1])  # description
            self.purpose_category_var.set(purpose_data[2])  # category
            
        except Exception as e:
            logging.error(f"Error selecting purpose: {e}")
    
    def _update_purpose_combos(self):
        """Update purpose comboboxes in other tabs"""
        try:
            # Get active purposes
            purpose_names = self.db.get_purpose_names(active_only=True)
            
            # Update movement tab combobox
            if hasattr(self, 'mov_purpose_combobox'):
                self.mov_purpose_combobox.set_values(purpose_names)
            
            # Update vehicle tab combobox
            if hasattr(self, 'ent_vpurpose'):
                self.ent_vpurpose.set_values(purpose_names)
                
        except Exception as e:
            logging.error(f"Error updating purpose combos: {e}")
    
    # Export and reporting methods
    def _export_movements_csv(self):
        """Export movements to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialname=f"movements_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                self.db.cursor.execute("""
              SELECT m.date, d.name || ' ' || d.surname as driver, v.plate, 
                  m.start_km, m.end_km, m.route, m.purpose,
                           (COALESCE(m.end_km, 0) - COALESCE(m.start_km, 0)) as total_km,
                           COALESCE(m.movement_number, 0) as movement_num
                    FROM movements m
                    JOIN drivers d ON m.driver_id = d.id
                    JOIN vehicles v ON m.vehicle_id = v.id
                    ORDER BY m.date DESC
                """)
                
                headers = ["Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", "ÎŸÎ´Î·Î³ÏŒÏ‚", "ÎŒÏ‡Î·Î¼Î±", "Î§Î»Î¼ Î‘Î½Î±Ï‡.", "Î§Î»Î¼ Î•Ï€Î¹ÏƒÏ„Ï.", 
                          "Î”Î¹Î±Î´ÏÎ¿Î¼Î®", "Î£ÎºÎ¿Ï€ÏŒÏ‚", "Î£ÏÎ½Î¿Î»Î¿ Î§Î»Î¼", "Î‘Ï. ÎšÎ¯Î½Î·ÏƒÎ·Ï‚"]
                data = self.db.cursor.fetchall()
                
                if export_to_csv(data, headers, filename):
                    self.status_bar.set_status("ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ ÎµÎ¾Î®Ï‡Î¸Î·ÏƒÎ±Î½ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                    
        except Exception as e:
            logging.error(f"Error exporting movements: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ¾Î±Î³Ï‰Î³Î®: {str(e)}")

    def _export_fuel_csv(self):
        """Export fuel to CSV"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialname=f"fuel_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            )
            
            if filename:
                self.db.cursor.execute("""
                    SELECT f.date, d.name || ' ' || d.surname, v.plate, f.liters, f.mileage, f.cost
                    FROM fuel f
                    JOIN drivers d ON f.driver_id = d.id
                    JOIN vehicles v ON f.vehicle_id = v.id
                    ORDER BY f.date DESC
                """)
                
                headers = ["Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", "ÎŸÎ´Î·Î³ÏŒÏ‚", "ÎŒÏ‡Î·Î¼Î±", "Î›Î¯Ï„ÏÎ±", "Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", "ÎšÏŒÏƒÏ„Î¿Ï‚"]
                data = self.db.cursor.fetchall()
                
                if export_to_csv(data, headers, filename):
                    self.status_bar.set_status("ÎšÎ±ÏÏƒÎ¹Î¼Î± ÎµÎ¾Î®Ï‡Î¸Î·ÏƒÎ±Î½ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                    
        except Exception as e:
            logging.error(f"Error exporting fuel: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ¾Î±Î³Ï‰Î³Î®: {str(e)}")

    def _export_monthly_report_csv(self):
        """Export comprehensive monthly report to CSV"""
        try:
            month = int(self.month_var.get())
            year = int(self.year_var.get())
            
            data = self._get_monthly_data(month, year)
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                initialname=f"monthly_report_{year}_{month:02d}.csv"
            )
            
            if filename:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    
                    # Header
                    writer.writerow([f"ÎœÎ—ÎÎ™Î‘Î™Î‘ Î‘ÎÎ‘Î¦ÎŸÎ¡Î‘ - {month:02d}/{year}"])
                    writer.writerow([f"Î”Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
                    writer.writerow([])
                    
                    # Individual vehicle data
                    writer.writerow(["Î£Î¤ÎŸÎ™Î§Î•Î™Î‘ Î‘ÎÎ‘ ÎŸÎ§Î—ÎœÎ‘"])
                    writer.writerow([
                        "Î Î¹Î½Î±ÎºÎ¯Î´Î±", "ÎœÎ¬ÏÎºÎ±", "Î¤ÏÏ€Î¿Ï‚", "ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", "Î£ÏÎ½Î¿Î»Î¿ Î§Î»Î¼", 
                        "Î•Î»Î¬Ï‡Î¹ÏƒÏ„Î± Î§Î»Î¼", "ÎœÎ­Î³Î¹ÏƒÏ„Î± Î§Î»Î¼", "Î£ÏÎ½Î¿Î»Î¿ ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½ (L)", 
                        "Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼Î¿Î¯", "ÎœÎ­ÏƒÎ¿Ï‚ ÎŒÏÎ¿Ï‚/Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒ", "Î‘Ï€ÏŒÎ´Î¿ÏƒÎ· (Ï‡Î»Î¼/L)"
                    ])
                    
                    total_km = 0
                    total_fuel = 0
                    total_movements = 0
                    
                    for vehicle in data:
                        writer.writerow([
                            vehicle['plate'],
                            vehicle['brand'],
                            vehicle['type'],
                            vehicle['total_movements'],
                            f"{vehicle['total_km']:.1f}",
                            f"{vehicle['min_km']:.0f}",
                            f"{vehicle['max_km']:.0f}",
                            f"{vehicle['total_fuel']:.1f}",
                            vehicle['fuel_refills'],
                            f"{vehicle['avg_fuel_per_refill']:.1f}",
                            f"{vehicle['efficiency_km_per_liter']:.2f}"
                        ])
                        
                        total_km += vehicle['total_km']
                        total_fuel += vehicle['total_fuel']
                        total_movements += vehicle['total_movements']
                    
                    # Summary
                    writer.writerow([])
                    writer.writerow(["Î£Î¥ÎÎŸÎ›Î™ÎšÎ‘ Î£Î¤ÎŸÎ™Î§Î•Î™Î‘"])
                    writer.writerow(["Î£Ï…Î½Î¿Î»Î¹ÎºÎ­Ï‚ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", total_movements])
                    writer.writerow(["Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", f"{total_km:.1f}"])
                    writer.writerow(["Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ ÎšÎ±ÏÏƒÎ¹Î¼Î±", f"{total_fuel:.1f}"])
                    if total_fuel > 0:
                        writer.writerow(["ÎœÎ­ÏƒÎ· Î‘Ï€ÏŒÎ´Î¿ÏƒÎ·", f"{total_km/total_fuel:.2f} Ï‡Î»Î¼/L"])
                
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"Î— Î¼Î·Î½Î¹Î±Î¯Î± Î±Î½Î±Ï†Î¿ÏÎ¬ ÎµÎ¾Î®Ï‡Î¸Î· ÏƒÎµ: {filename}")
                self.status_bar.set_status("ÎœÎ·Î½Î¹Î±Î¯Î± Î±Î½Î±Ï†Î¿ÏÎ¬ ÎµÎ¾Î®Ï‡Î¸Î· ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                
        except Exception as e:
            logging.error(f"Error exporting monthly report: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ¾Î±Î³Ï‰Î³Î®: {str(e)}")

    def _export_monthly_report_excel(self):
        """Export comprehensive monthly report to Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î¤Î¿ Excel Î´ÎµÎ½ ÎµÎ¯Î½Î±Î¹ Î´Î¹Î±Î¸Î­ÏƒÎ¹Î¼Î¿!")
            return
        
        try:
            from openpyxl import Workbook
            
            month = int(self.month_var.get())
            year = int(self.year_var.get())
            
            data = self._get_monthly_data(month, year)
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialname=f"monthly_report_{year}_{month:02d}.xlsx"
            )
            
            if filename:
                wb = Workbook()
                ws = wb.active
                ws.title = f"Î‘Î½Î±Ï†Î¿ÏÎ¬ {month:02d}-{year}"
                
                # Header
                ws.append([f"ÎœÎ—ÎÎ™Î‘Î™Î‘ Î‘ÎÎ‘Î¦ÎŸÎ¡Î‘ - {month:02d}/{year}"])
                ws.append([f"Î”Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ: {datetime.now().strftime('%d/%m/%Y %H:%M')}"])
                ws.append([])
                
                # Individual vehicle data
                ws.append(["Î£Î¤ÎŸÎ™Î§Î•Î™Î‘ Î‘ÎÎ‘ ÎŸÎ§Î—ÎœÎ‘"])
                ws.append([
                    "Î Î¹Î½Î±ÎºÎ¯Î´Î±", "ÎœÎ¬ÏÎºÎ±", "Î¤ÏÏ€Î¿Ï‚", "ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", "Î£ÏÎ½Î¿Î»Î¿ Î§Î»Î¼", 
                    "Î•Î»Î¬Ï‡Î¹ÏƒÏ„Î± Î§Î»Î¼", "ÎœÎ­Î³Î¹ÏƒÏ„Î± Î§Î»Î¼", "Î£ÏÎ½Î¿Î»Î¿ ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½ (L)", 
                    "Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼Î¿Î¯", "ÎœÎ­ÏƒÎ¿Ï‚ ÎŒÏÎ¿Ï‚/Î‘Î½ÎµÏ†Î¿Î´Î¹Î±ÏƒÎ¼ÏŒ", "Î‘Ï€ÏŒÎ´Î¿ÏƒÎ· (Ï‡Î»Î¼/L)"
                ])
                
                total_km = 0
                total_fuel = 0
                total_movements = 0
                
                for vehicle in data:
                    ws.append([
                        vehicle['plate'],
                        vehicle['brand'],
                        vehicle['type'],
                        vehicle['total_movements'],
                        vehicle['total_km'],
                        vehicle['min_km'],
                        vehicle['max_km'],
                        vehicle['total_fuel'],
                        vehicle['fuel_refills'],
                        vehicle['avg_fuel_per_refill'],
                        vehicle['efficiency_km_per_liter']
                    ])
                    
                    total_km += vehicle['total_km']
                    total_fuel += vehicle['total_fuel']
                    total_movements += vehicle['total_movements']
                
                # Summary
                ws.append([])
                ws.append(["Î£Î¥ÎÎŸÎ›Î™ÎšÎ‘ Î£Î¤ÎŸÎ™Î§Î•Î™Î‘"])
                ws.append(["Î£Ï…Î½Î¿Î»Î¹ÎºÎ­Ï‚ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", total_movements])
                ws.append(["Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", total_km])
                ws.append(["Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ ÎšÎ±ÏÏƒÎ¹Î¼Î±", total_fuel])
                if total_fuel > 0:
                    ws.append(["ÎœÎ­ÏƒÎ· Î‘Ï€ÏŒÎ´Î¿ÏƒÎ·", f"{total_km/total_fuel:.2f} Ï‡Î»Î¼/L"])
                
                wb.save(filename)
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"Î— Î¼Î·Î½Î¹Î±Î¯Î± Î±Î½Î±Ï†Î¿ÏÎ¬ ÎµÎ¾Î®Ï‡Î¸Î· ÏƒÎµ: {filename}")
                self.status_bar.set_status("ÎœÎ·Î½Î¹Î±Î¯Î± Î±Î½Î±Ï†Î¿ÏÎ¬ Excel ÎµÎ¾Î®Ï‡Î¸Î· ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                
        except Exception as e:
            logging.error(f"Error exporting monthly Excel report: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ¾Î±Î³Ï‰Î³Î®: {str(e)}")

    def _export_excel(self):
        """Export comprehensive data to Excel"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î¤Î¿ Excel Î´ÎµÎ½ ÎµÎ¯Î½Î±Î¹ Î´Î¹Î±Î¸Î­ÏƒÎ¹Î¼Î¿!")
            return
        
        try:
            from openpyxl import Workbook
            
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialname=f"fleet_data_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if filename:
                wb = Workbook()
                
                # Movements sheet
                ws1 = wb.active
                ws1.title = "ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚"
                ws1.append(["Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", "ÎŸÎ´Î·Î³ÏŒÏ‚", "ÎŒÏ‡Î·Î¼Î±", "Î§Î»Î¼ Î‘Î½Î±Ï‡.", "Î§Î»Î¼ Î•Ï€Î¹ÏƒÏ„Ï.", "Î”Î¹Î±Î´ÏÎ¿Î¼Î®", "Î£ÎºÎ¿Ï€ÏŒÏ‚", "Î£ÏÎ½Î¿Î»Î¿ Î§Î»Î¼", "Î‘Ï. ÎšÎ¯Î½Î·ÏƒÎ·Ï‚"])
                
                self.db.cursor.execute("""
                    SELECT m.date, d.name || ' ' || d.surname, v.plate, 
                           m.start_km, m.end_km, m.route, m.purpose,
                           (COALESCE(m.end_km, 0) - COALESCE(m.start_km, 0)),
                           COALESCE(m.movement_number, 0)
                    FROM movements m
                    JOIN drivers d ON m.driver_id = d.id
                    JOIN vehicles v ON m.vehicle_id = v.id
                    ORDER BY m.date DESC
                """)
                
                for row in self.db.cursor.fetchall():
                    ws1.append(row)
                
                # Fuel sheet
                ws2 = wb.create_sheet("ÎšÎ±ÏÏƒÎ¹Î¼Î±")
                ws2.append(["Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±", "ÎŸÎ´Î·Î³ÏŒÏ‚", "ÎŒÏ‡Î·Î¼Î±", "Î›Î¯Ï„ÏÎ±", "Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", "ÎšÏŒÏƒÏ„Î¿Ï‚"])
                
                self.db.cursor.execute("""
                    SELECT f.date, d.name || ' ' || d.surname, v.plate, f.liters, f.mileage, f.cost
                    FROM fuel f
                    JOIN drivers d ON f.driver_id = d.id
                    JOIN vehicles v ON f.vehicle_id = v.id
                    ORDER BY f.date DESC
                """)
                
                for row in self.db.cursor.fetchall():
                    ws2.append(row)
                
                # Vehicles sheet
                ws3 = wb.create_sheet("ÎŸÏ‡Î®Î¼Î±Ï„Î±")
                ws3.append(["Î Î¹Î½Î±ÎºÎ¯Î´Î±", "ÎœÎ¬ÏÎºÎ±", "Î¤ÏÏ€Î¿Ï‚", "Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚"])
                
                self.db.cursor.execute("SELECT plate, brand, vtype, purpose FROM vehicles ORDER BY plate")
                for row in self.db.cursor.fetchall():
                    ws3.append(row)
                
                # Drivers sheet
                ws4 = wb.create_sheet("ÎŸÎ´Î·Î³Î¿Î¯")
                ws4.append(["ÎŒÎ½Î¿Î¼Î±", "Î•Ï€ÏÎ½Ï…Î¼Î¿", "Î£Î·Î¼ÎµÎ¹ÏÏƒÎµÎ¹Ï‚"])
                
                self.db.cursor.execute("SELECT name, surname, notes FROM drivers ORDER BY name, surname")
                for row in self.db.cursor.fetchall():
                    ws4.append(row)
                
                wb.save(filename)
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"Î¤Î± Î´ÎµÎ´Î¿Î¼Î­Î½Î± ÎµÎ¾Î®Ï‡Î¸Î·ÏƒÎ±Î½ ÏƒÎµ: {filename}")
                self.status_bar.set_status("Î“ÎµÎ½Î¹ÎºÎ® Î±Î½Î±Ï†Î¿ÏÎ¬ Excel ÎµÎ¾Î®Ï‡Î¸Î· ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
                
        except Exception as e:
            logging.error(f"Error exporting Excel: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ¾Î±Î³Ï‰Î³Î®: {str(e)}")

    def _get_monthly_data(self, month, year):
        """Get comprehensive monthly data for vehicles"""
        month_str = f"{year}-{month:02d}"
        
        # Get movements data for the month
        movements_query = """
            SELECT v.plate, v.brand, v.vtype,
                   COUNT(m.id) as total_movements,
                   SUM(COALESCE(m.end_km, 0) - COALESCE(m.start_km, 0)) as total_km,
                   MIN(m.start_km) as min_km,
                   MAX(COALESCE(m.end_km, m.start_km)) as max_km
            FROM vehicles v
            LEFT JOIN movements m ON v.id = m.vehicle_id 
                AND m.date LIKE ?
                AND m.end_km IS NOT NULL
            GROUP BY v.id, v.plate, v.brand, v.vtype
            ORDER BY v.plate
        """
        
        # Get fuel data for the month
        fuel_query = """
            SELECT v.plate,
                   SUM(f.liters) as total_fuel,
                   COUNT(f.id) as fuel_refills,
                   AVG(f.liters) as avg_fuel_per_refill
            FROM vehicles v
            LEFT JOIN fuel f ON v.id = f.vehicle_id 
                AND f.date LIKE ?
            GROUP BY v.id, v.plate
            ORDER BY v.plate
        """
        
        self.db.cursor.execute(movements_query, (f"{month_str}%",))
        movements_data = {row[0]: row for row in self.db.cursor.fetchall()}
        
        self.db.cursor.execute(fuel_query, (f"{month_str}%",))
        fuel_data = {row[0]: row for row in self.db.cursor.fetchall()}
        
        # Combine data
        combined_data = []
        for plate in movements_data:
            movement_row = movements_data[plate]
            fuel_row = fuel_data.get(plate, (plate, 0, 0, 0))
            
            # Calculate efficiency (km per liter)
            efficiency = 0
            if fuel_row[1] > 0 and movement_row[4] > 0:
                efficiency = movement_row[4] / fuel_row[1]
            
            combined_row = {
                'plate': plate,
                'brand': movement_row[1],
                'type': movement_row[2],
                'total_movements': movement_row[3],
                'total_km': movement_row[4] or 0,
                'min_km': movement_row[5] or 0,
                'max_km': movement_row[6] or 0,
                'total_fuel': fuel_row[1] or 0,
                'fuel_refills': fuel_row[2] or 0,
                'avg_fuel_per_refill': fuel_row[3] or 0,
                'efficiency_km_per_liter': efficiency
            }
            combined_data.append(combined_row)
        
        return combined_data

    def _update_statistics(self):
        """Update the statistics display"""
        try:
            self.stats_text.delete(1.0, tk.END)
            
            # Current month data
            current_month = datetime.now().month
            current_year = datetime.now().year
            
            stats_content = f"Î£Î¤Î‘Î¤Î™Î£Î¤Î™ÎšÎ‘ Î£Î¤ÎŸÎ™Î§Î•Î™Î‘ - {current_month:02d}/{current_year}\n"
            stats_content += "=" * 50 + "\n\n"
            
            # Total vehicles
            self.db.cursor.execute("SELECT COUNT(*) FROM vehicles")
            total_vehicles = self.db.cursor.fetchone()[0]
            stats_content += f"Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ ÎŸÏ‡Î®Î¼Î±Ï„Î±: {total_vehicles}\n"
            
            # Total drivers
            self.db.cursor.execute("SELECT COUNT(*) FROM drivers")
            total_drivers = self.db.cursor.fetchone()[0]
            stats_content += f"Î£Ï…Î½Î¿Î»Î¹ÎºÎ¿Î¯ ÎŸÎ´Î·Î³Î¿Î¯: {total_drivers}\n\n"
            
            # Current month movements
            month_str = f"{current_year}-{current_month:02d}"
            self.db.cursor.execute("""
                SELECT COUNT(*) FROM movements 
                WHERE date LIKE ? AND end_km IS NOT NULL
            """, (f"{month_str}%",))
            month_movements = self.db.cursor.fetchone()[0]
            stats_content += f"ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ Ï„ÏÎ­Ï‡Î¿Î½Ï„Î¿Ï‚ Î¼Î®Î½Î±: {month_movements}\n"
            
            # Current month kilometers
            self.db.cursor.execute("""
                SELECT SUM(COALESCE(end_km, 0) - COALESCE(start_km, 0)) 
                FROM movements 
                WHERE date LIKE ? AND end_km IS NOT NULL
            """, (f"{month_str}%",))
            month_km = self.db.cursor.fetchone()[0] or 0
            stats_content += f"Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ± Ï„ÏÎ­Ï‡Î¿Î½Ï„Î¿Ï‚ Î¼Î®Î½Î±: {month_km:.1f} Ï‡Î»Î¼\n"
            
            # Current month fuel
            self.db.cursor.execute("""
                SELECT SUM(liters) FROM fuel 
                WHERE date LIKE ?
            """, (f"{month_str}%",))
            month_fuel = self.db.cursor.fetchone()[0] or 0
            stats_content += f"ÎšÎ±ÏÏƒÎ¹Î¼Î± Ï„ÏÎ­Ï‡Î¿Î½Ï„Î¿Ï‚ Î¼Î®Î½Î±: {month_fuel:.1f} L\n\n"
            
            # Active movements (not returned)
            self.db.cursor.execute("SELECT COUNT(*) FROM movements WHERE end_km IS NULL")
            active_movements = self.db.cursor.fetchone()[0]
            stats_content += f"Î•Î½ÎµÏÎ³Î­Ï‚ ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ (Î´ÎµÎ½ Î­Ï‡Î¿Ï…Î½ ÎµÏ€Î¹ÏƒÏ„ÏÎ­ÏˆÎµÎ¹): {active_movements}\n"
            
            # Tank level
            self.db.cursor.execute("SELECT SUM(CASE WHEN type = 'fill' THEN liters ELSE -liters END) FROM tank")
            tank_level = self.db.cursor.fetchone()[0] or 0
            stats_content += f"Î•Ï€Î¯Ï€ÎµÎ´Î¿ Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚: {tank_level:.1f} L\n\n"
            
            # Top 5 vehicles by distance this month
            stats_content += "Î¤ÎŸÎ  5 ÎŸÎ§Î—ÎœÎ‘Î¤Î‘ (Î§Î™Î›Î™ÎŸÎœÎ•Î¤Î¡Î‘ ÎœÎ—ÎÎ‘):\n"
            stats_content += "-" * 30 + "\n"
            
            self.db.cursor.execute("""
                SELECT v.plate, SUM(COALESCE(m.end_km, 0) - COALESCE(m.start_km, 0)) as total_km
                FROM vehicles v
                LEFT JOIN movements m ON v.id = m.vehicle_id 
                    AND m.date LIKE ? AND m.end_km IS NOT NULL
                GROUP BY v.id, v.plate
                ORDER BY total_km DESC
                LIMIT 5
            """, (f"{month_str}%",))
            
            for i, (plate, km) in enumerate(self.db.cursor.fetchall(), 1):
                stats_content += f"{i}. {plate}: {km:.1f} Ï‡Î»Î¼\n"
            
            self.stats_text.insert(tk.END, stats_content)
            
        except Exception as e:
            logging.error(f"Error updating statistics: {e}")
            self.stats_text.insert(tk.END, f"Î£Ï†Î¬Î»Î¼Î± ÏƒÏ„Î·Î½ Î±Î½Î¬ÎºÏ„Î·ÏƒÎ· ÏƒÏ„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÏÎ½: {str(e)}")
    
    # Context menu methods for vehicles
    def _show_vehicle_context_menu(self, event):
        """Show context menu for vehicle right-click"""
        item = self.tree_vehicles.identify_row(event.y)
        if item:
            self.tree_vehicles.selection_set(item)
            self.vehicle_context_menu.post(event.x_root, event.y_root)

    def _get_selected_vehicle_plate(self):
        """Get the plate of the currently selected vehicle"""
        selection = self.tree_vehicles.selection()
        if not selection:
            messagebox.showwarning("âš ï¸ Î ÏÎ¿ÏƒÎ¿Ï‡Î®", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î­Î½Î± ÏŒÏ‡Î·Î¼Î±!")
            return None
        
        item = selection[0]
        values = self.tree_vehicles.item(item, 'values')
        return values[0] if values else None

    def _show_vehicle_history(self):
        """Show movement history for selected vehicle"""
        plate = self._get_selected_vehicle_plate()
        if not plate:
            return

        history_window = tk.Toplevel(self.root)
        history_window.title(f"Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½ - {plate}")
        history_window.geometry("900x600")
        history_window.configure(bg=THEMES[self.current_theme]["bg"])

        title_frame = tk.Frame(history_window, bg=THEMES[self.current_theme]["bg"])
        title_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(
            title_frame,
            text=f"ğŸ“‹ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½: {plate}",
            font=FONT_TITLE,
            fg=THEMES[self.current_theme]["accent"],
            bg=THEMES[self.current_theme]["bg"],
        ).pack()

        tree_frame = tk.Frame(history_window, bg=THEMES[self.current_theme]["bg"])
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        tree = ttk.Treeview(
            tree_frame,
            columns=("date", "driver", "movement_num", "start_km", "end_km", "total_km", "route", "purpose"),
            show="headings",
        )

        for col, text in [
            ("date", "Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±"),
            ("driver", "ÎŸÎ´Î·Î³ÏŒÏ‚"),
            ("movement_num", "Î‘Ï. ÎšÎ¯Î½Î·ÏƒÎ·Ï‚"),
            ("start_km", "Î§Î»Î¼ Î‘Î½Î±Ï‡."),
            ("end_km", "Î§Î»Î¼ Î•Ï€Î¹ÏƒÏ„Ï."),
            ("total_km", "Î£ÏÎ½Î¿Î»Î¿ Î§Î»Î¼"),
            ("route", "Î”Î¹Î±Î´ÏÎ¿Î¼Î®"),
            ("purpose", "Î£ÎºÎ¿Ï€ÏŒÏ‚"),
        ]:
            tree.heading(col, text=text)
            tree.column(col, width=100, anchor="center")

        tree.pack(fill="both", expand=True)

        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")

        try:
            self.db.cursor.execute(
                """
                SELECT m.date, d.name || ' ' || d.surname, COALESCE(m.movement_number, 0),
                       m.start_km, m.end_km,
                       (COALESCE(m.end_km, 0) - COALESCE(m.start_km, 0)) as total_km,
                       m.route, m.purpose
                FROM movements m
                JOIN drivers d ON m.driver_id = d.id
                JOIN vehicles v ON m.vehicle_id = v.id
                WHERE v.plate = ?
                ORDER BY m.date DESC, m.id DESC
                """,
                (plate,),
            )
            for row in self.db.cursor.fetchall():
                formatted_row = (
                    row[0],
                    row[1],
                    f"{row[2]:04d}" if row[2] > 0 else "---",
                    f"{row[3]:.0f}" if row[3] else "",
                    f"{row[4]:.0f}" if row[4] else "Î”ÎµÎ½ ÎµÏ€Î­ÏƒÏ„ÏÎµÏˆÎµ",
                    f"{row[5]:.1f}" if row[5] > 0 else "",
                    row[6] or "",
                    row[7] or "",
                )
                tree.insert("", "end", values=formatted_row)
        except Exception as e:
            logging.error(f"Error loading vehicle history: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÏƒÏ„Î·Î½ Î±Î½Î¬ÎºÏ„Î·ÏƒÎ· Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½: {str(e)}")

        stats_frame = tk.Frame(history_window, bg=THEMES[self.current_theme]["bg"])
        stats_frame.pack(fill="x", padx=10, pady=5)

        try:
            self.db.cursor.execute(
                """
                SELECT COUNT(*) as total_movements,
                       SUM(COALESCE(end_km, 0) - COALESCE(start_km, 0)) as total_km,
                       AVG(COALESCE(end_km, 0) - COALESCE(start_km, 0)) as avg_km
                FROM movements m
                JOIN vehicles v ON m.vehicle_id = v.id
                WHERE v.plate = ? AND m.end_km IS NOT NULL
                """,
                (plate,),
            )
            result = self.db.cursor.fetchone()
            total_movements = result[0] or 0
            total_km = result[1] or 0
            avg_km = result[2] or 0

            stats_text = (
                f"ğŸ“Š Î£ÏÎ½Î¿Î»Î¿ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½: {total_movements} | Î£ÏÎ½Î¿Î»Î¿ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±: {total_km:.1f} Ï‡Î»Î¼ | "
                f"ÎœÎ­ÏƒÎ¿Ï‚ ÎŒÏÎ¿Ï‚: {avg_km:.1f} Ï‡Î»Î¼/ÎºÎ¯Î½Î·ÏƒÎ·"
            )
            tk.Label(
                stats_frame,
                text=stats_text,
                font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"],
                fg=THEMES[self.current_theme]["fg"],
            ).pack()
        except Exception as e:
            logging.error(f"Error calculating vehicle statistics: {e}")
            tk.Label(
                stats_frame,
                text="Î£Ï†Î¬Î»Î¼Î± ÏƒÏ„Î¿Î½ Ï…Ï€Î¿Î»Î¿Î³Î¹ÏƒÎ¼ÏŒ ÏƒÏ„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÏÎ½",
                fg=THEMES[self.current_theme]["danger"],
                bg=THEMES[self.current_theme]["bg"],
            ).pack()

    def _show_vehicle_fuel_history(self):
        """Show fuel history for selected vehicle"""
        plate = self._get_selected_vehicle_plate()
        if not plate:
            return
        
        # Create fuel history window
        fuel_window = tk.Toplevel(self.root)
        fuel_window.title(f"Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½ - {plate}")
        fuel_window.geometry("800x500")
        fuel_window.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Title
        title_frame = tk.Frame(fuel_window, bg=THEMES[self.current_theme]["bg"])
        title_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(title_frame, text=f"â›½ Î™ÏƒÏ„Î¿ÏÎ¹ÎºÏŒ ÎšÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½: {plate}", 
                font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                bg=THEMES[self.current_theme]["bg"]).pack()
        
        # Create treeview for fuel records
        tree_frame = tk.Frame(fuel_window, bg=THEMES[self.current_theme]["bg"])
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        fuel_tree = ttk.Treeview(tree_frame, columns=("date", "driver", "liters", "mileage", "cost"), show="headings")
        
        fuel_tree.heading("date", text="Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±")
        fuel_tree.heading("driver", text="ÎŸÎ´Î·Î³ÏŒÏ‚")
        fuel_tree.heading("liters", text="Î›Î¯Ï„ÏÎ±")
        fuel_tree.heading("mileage", text="Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±")
        fuel_tree.heading("cost", text="ÎšÏŒÏƒÏ„Î¿Ï‚")
        
        for col in fuel_tree["columns"]:
            fuel_tree.column(col, width=150, anchor="center")
        
        fuel_tree.pack(fill="both", expand=True)
        
        # Add scrollbar
        fuel_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=fuel_tree.yview)
        fuel_tree.configure(yscrollcommand=fuel_scrollbar.set)
        fuel_scrollbar.pack(side="right", fill="y")
        
        # Load fuel data
        try:
            self.db.cursor.execute("""
                SELECT f.date, d.name || ' ' || d.surname, f.liters, f.mileage, f.cost
                FROM fuel f
                JOIN drivers d ON f.driver_id = d.id
                JOIN vehicles v ON f.vehicle_id = v.id
                WHERE v.plate = ?
                ORDER BY f.date DESC
            """, (plate,))
            
            total_fuel = 0
            total_cost = 0
            for row in self.db.cursor.fetchall():
                display_row = (
                    row[0], row[1], f"{row[2]:.1f}", 
                    f"{row[3]:.0f}" if row[3] else "",
                    format_currency(row[4]) if row[4] else "0.00 â‚¬"
                )
                fuel_tree.insert("", "end", values=display_row)
                total_fuel += row[2]
                total_cost += row[4] if row[4] else 0
                
            # Statistics
            stats_frame = tk.Frame(fuel_window, bg=THEMES[self.current_theme]["bg"])
            stats_frame.pack(fill="x", padx=10, pady=5)
            stats_text = f"ğŸ“Š Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ ÎšÎ±ÏÏƒÎ¹Î¼Î±: {total_fuel:.1f} Î»Î¯Ï„ÏÎ± | Î£Ï…Î½Î¿Î»Î¹ÎºÏŒ ÎšÏŒÏƒÏ„Î¿Ï‚: {format_currency(total_cost)}"
            tk.Label(stats_frame, text=stats_text, font=FONT_NORMAL,
                    bg=THEMES[self.current_theme]["bg"], 
                    fg=THEMES[self.current_theme]["fg"]).pack()
                    
        except Exception as e:
            logging.error(f"Error loading vehicle fuel history: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÏƒÏ„Î·Î½ Î±Î½Î¬ÎºÏ„Î·ÏƒÎ· Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½: {str(e)}")

    def _show_vehicle_statistics(self):
        """Show comprehensive statistics for selected vehicle"""
        plate = self._get_selected_vehicle_plate()
        if not plate:
            return
        
        # Create statistics window
        stats_window = tk.Toplevel(self.root)
        stats_window.title(f"Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÎ¬ ÎŸÏ‡Î®Î¼Î±Ï„Î¿Ï‚ - {plate}")
        stats_window.geometry("600x500")
        stats_window.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Title
        title_frame = tk.Frame(stats_window, bg=THEMES[self.current_theme]["bg"])
        title_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(title_frame, text=f"ğŸ“ˆ Î£Ï„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÎ¬: {plate}", 
                font=FONT_TITLE, fg=THEMES[self.current_theme]["accent"],
                bg=THEMES[self.current_theme]["bg"]).pack()
        
        # Text widget for statistics
        text_frame = tk.Frame(stats_window, bg=THEMES[self.current_theme]["bg"])
        text_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        stats_text = tk.Text(text_frame, font=("Courier", 10), wrap=tk.WORD,
                            bg=THEMES[self.current_theme]["entry_bg"], 
                            fg=THEMES[self.current_theme]["fg"],
                            relief="flat", borderwidth=1)
        stats_text.pack(fill="both", expand=True)
        
        # Add scrollbar
        text_scrollbar = ttk.Scrollbar(text_frame, orient="vertical", command=stats_text.yview)
        stats_text.configure(yscrollcommand=text_scrollbar.set)
        text_scrollbar.pack(side="right", fill="y")
        
        try:
            # Get vehicle info
            self.db.cursor.execute("SELECT brand, vtype, purpose FROM vehicles WHERE plate = ?", (plate,))
            vehicle_info = self.db.cursor.fetchone()
            
            stats_content = f"Î£Î¤Î‘Î¤Î™Î£Î¤Î™ÎšÎ‘ ÎŸÎ§Î—ÎœÎ‘Î¤ÎŸÎ£: {plate}\n"
            stats_content += "=" * 50 + "\n\n"
            
            if vehicle_info:
                stats_content += f"ÎœÎ¬ÏÎºÎ±: {vehicle_info[0] or 'Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ Î¿ÏÎ¹ÏƒÏ„ÎµÎ¯'}\n"
                stats_content += f"Î¤ÏÏ€Î¿Ï‚: {vehicle_info[1] or 'Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ Î¿ÏÎ¹ÏƒÏ„ÎµÎ¯'}\n"
                stats_content += f"Î£ÎºÎ¿Ï€ÏŒÏ‚: {vehicle_info[2] or 'Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ Î¿ÏÎ¹ÏƒÏ„ÎµÎ¯'}\n\n"
            
            # Movement statistics
            self.db.cursor.execute("""
                SELECT COUNT(*) as total_movements,
                       SUM(COALESCE(end_km, 0) - COALESCE(start_km, 0)) as total_km,
                       AVG(COALESCE(end_km, 0) - COALESCE(start_km, 0)) as avg_km,
                       MIN(start_km) as min_km,
                       MAX(COALESCE(end_km, start_km)) as max_km
                FROM movements m
                JOIN vehicles v ON m.vehicle_id = v.id
                WHERE v.plate = ? AND m.end_km IS NOT NULL
            """, (plate,))
            movement_stats = self.db.cursor.fetchone()
            if movement_stats:
                total_movements = movement_stats[0] or 0
                total_km = movement_stats[1] or 0
                avg_km = movement_stats[2] or 0
                min_km = movement_stats[3] or 0
                max_km = movement_stats[4] or 0
                stats_content += f"Î£Ï…Î½Î¿Î»Î¹ÎºÎ­Ï‚ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ (Î¿Î»Î¿ÎºÎ»Î·ÏÏ‰Î¼Î­Î½ÎµÏ‚): {total_movements}\n"
                stats_content += f"Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±: {total_km:.1f} Ï‡Î»Î¼\n"
                stats_content += f"ÎœÎ­ÏƒÎ¿Ï‚ ÎŒÏÎ¿Ï‚ Î§Î»Î¼/ÎšÎ¯Î½Î·ÏƒÎ·: {avg_km:.1f} Ï‡Î»Î¼\n"
                stats_content += f"Î‘Ï€ÏŒ Î§Î»Î¼: {min_km:.0f} Î­Ï‰Ï‚ {max_km:.0f}\n\n"
            # Active movements (not returned)
            self.db.cursor.execute("""
                SELECT COUNT(*) FROM movements m
                JOIN vehicles v ON m.vehicle_id = v.id
                WHERE v.plate = ? AND m.end_km IS NULL
            """, (plate,))
            active_movements = self.db.cursor.fetchone()[0] or 0
            stats_content += f"Î•Î½ÎµÏÎ³Î­Ï‚ ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚: {active_movements}\n\n"

            stats_text.insert(tk.END, stats_content)
        except Exception as e:
            logging.error(f"Error calculating vehicle statistics: {e}")
            stats_text.insert(tk.END, f"Î£Ï†Î¬Î»Î¼Î± ÏƒÏ„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÏÎ½: {e}")

        # End of statistics window creation

    def _view_vehicle_photo_from_tree(self):
        """View photo of selected vehicle from tree"""
        plate = self._get_selected_vehicle_plate()
        if not plate:
            return
        
        try:
            # Get vehicle photo path from database
            self.db.cursor.execute("SELECT photo_path FROM vehicles WHERE plate = ?", (plate,))
            result = self.db.cursor.fetchone()
            
            if not result or not result[0]:
                messagebox.showinfo("ğŸ“· Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±", f"Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡ÎµÎ¹ Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î± Î³Î¹Î± Ï„Î¿ ÏŒÏ‡Î·Î¼Î± {plate}")
                return
            
            photo_path = result[0]
            if not os.path.exists(photo_path):
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î— Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î± Î´ÎµÎ½ Î²ÏÎ­Î¸Î·ÎºÎµ ÏƒÏ„Î¿ Î´Î¯ÏƒÎºÎ¿: {photo_path}")
                return
            
            # Create photo viewer window
            if PIL_AVAILABLE:
                photo_window = tk.Toplevel(self.root)
                photo_window.title(f"Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î± ÎŸÏ‡Î®Î¼Î±Ï„Î¿Ï‚ - {plate}")
                photo_window.configure(bg=THEMES[self.current_theme]["bg"])
                
                # Load and resize image
                from PIL import Image, ImageTk
                pil_image = Image.open(photo_path)
                pil_image.thumbnail((600, 400), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(pil_image)
                
                label = tk.Label(photo_window, image=photo, bg=THEMES[self.current_theme]["bg"])
                label.image = photo  # Keep a reference
                label.pack(padx=10, pady=10)
                
                # Add vehicle info
                info_label = tk.Label(
                    photo_window, 
                    text=f"ÎŒÏ‡Î·Î¼Î±: {plate}",
                    font=FONT_NORMAL,
                    bg=THEMES[self.current_theme]["bg"],
                    fg=THEMES[self.current_theme]["fg"]
                )
                info_label.pack(pady=5)
            else:
                messagebox.showinfo("ğŸ“· Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±", f"Î¦Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î± Î¿Ï‡Î®Î¼Î±Ï„Î¿Ï‚ {plate}: {photo_path}")
                
        except Exception as e:
            logging.error(f"Error viewing vehicle photo: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Ï€ÏÎ¿Î²Î¿Î»Î® Ï†Ï‰Ï„Î¿Î³ÏÎ±Ï†Î¯Î±Ï‚: {str(e)}")

    def _add_movement(self):
        """Add new movement with improved validation"""
        try:
            # Get form data
            date = self.mov_date_var.get()
            driver = self.mov_driver_combo.get()
            vehicle = self.mov_vehicle_combo.get()
            start_km = self.mov_start_km_var.get()
            route = self.mov_route_entry.get()
            purpose = self.mov_purpose_combobox.get() if hasattr(self, 'mov_purpose_combobox') else ""
            
            # Validate required fields
            required_fields = {
                "Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±": date,
                "ÎŸÎ´Î·Î³ÏŒÏ‚": driver,
                "ÎŒÏ‡Î·Î¼Î±": vehicle,
                "Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ± Î‘Î½Î±Ï‡ÏÏÎ·ÏƒÎ·Ï‚": start_km,
                "Î£ÎºÎ¿Ï€ÏŒÏ‚": purpose
            }
            
            is_valid, error_msg = self.validate_required_fields(required_fields)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î± Î•Ï€Î¹ÎºÏÏÏ‰ÏƒÎ·Ï‚", error_msg)
                return
            
            # Validate date format
            if not validate_date(date):
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "ÎœÎ· Î­Î³ÎºÏ…ÏÎ· Î·Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±! Î§ÏÎ·ÏƒÎ¹Î¼Î¿Ï€Î¿Î¹Î®ÏƒÏ„Îµ Ï„Î· Î¼Î¿ÏÏ†Î® YYYY-MM-DD")
                return
            
            # Validate kilometers
            is_valid, km_value = DataValidator.is_valid_km(start_km)
            if not is_valid:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", km_value)
                return
            
            # Get IDs
            driver_id = self._get_driver_id(driver)
            vehicle_id = self._get_vehicle_id(vehicle)
            
            if not driver_id:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "ÎœÎ· Î­Î³ÎºÏ…ÏÎ¿Ï‚ Î¿Î´Î·Î³ÏŒÏ‚!")
                return
            
            if not vehicle_id:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "ÎœÎ· Î­Î³ÎºÏ…ÏÎ¿ ÏŒÏ‡Î·Î¼Î±!")
                return
            
            # Check if vehicle is already active
            if self._is_vehicle_active(vehicle_id):
                if not ConfirmDialog(
                    self.root,
                    "âš ï¸ Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·",
                    f"Î¤Î¿ ÏŒÏ‡Î·Î¼Î± {vehicle} Î­Ï‡ÎµÎ¹ Î®Î´Î· ÎµÎ½ÎµÏÎ³Î® ÎºÎ¯Î½Î·ÏƒÎ·!\nÎ˜Î­Î»ÎµÏ„Îµ Î½Î± ÏƒÏ…Î½ÎµÏ‡Î¯ÏƒÎµÏ„Îµ;"
                ).show():
                    return
            
            # Get movement number
            movement_number = self.db.get_next_movement_number()
            
            # Insert movement
            self.db.cursor.execute("""
                INSERT INTO movements (movement_number, vehicle_id, driver_id, date, start_km, route, purpose)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (movement_number, vehicle_id, driver_id, date, km_value, route, purpose))
            
            self.db.conn.commit()
            
            # Clear form
            self._clear_movement_form()
            
            # Refresh data immediately
            self._load_movements()
            
            # Update status
            self.status_bar.set_status(f"ÎšÎ¯Î½Î·ÏƒÎ· #{movement_number:04d} Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚")
            
            log_user_action("Movement created", f"Movement #{movement_number:04d} for vehicle {vehicle}")
            
            # Ask user if they want to print the document
            if ConfirmDialog(
                self.root,
                "ï¸ Î•ÎºÏ„ÏÏ€Ï‰ÏƒÎ· Î•Î³Î³ÏÎ¬Ï†Î¿Ï…",
                f"Î— ÎºÎ¯Î½Î·ÏƒÎ· #{movement_number:04d} ÎºÎ±Ï„Î±Ï‡Ï‰ÏÎ®Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!\n\n"
                f"Î˜Î­Î»ÎµÏ„Îµ Î½Î± ÎµÎºÏ„Ï…Ï€ÏÏƒÎµÏ„Îµ Ï„Î¿ Î­Î³Î³ÏÎ±Ï†Î¿ ÎºÎ¯Î½Î·ÏƒÎ·Ï‚;"
            ).show():
                self._generate_and_print_movement_document(movement_number, date, driver, vehicle, km_value, route, purpose)
            
        except Exception as e:
            logging.error(f"Error adding movement: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎºÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ· ÎºÎ¯Î½Î·ÏƒÎ·Ï‚:\n{str(e)}")

    def _generate_and_print_movement_document(self, movement_number, date, driver, vehicle, start_km, route, purpose):
        """Generate and print movement document"""
        try:
            # Create document content
            doc_content = self._create_movement_document_content(movement_number, date, driver, vehicle, start_km, route, purpose)
            
            # Create organized folder structure for movements
            base_dir = os.getcwd()
            movements_dir = os.path.join(base_dir, "ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚")
            
            # Extract year and month from date (format: YYYY-MM-DD)
            date_parts = date.split('-')
            year = date_parts[0]
            month = date_parts[1]
            
            # Create year and month directories
            year_dir = os.path.join(movements_dir, year)
            month_dir = os.path.join(year_dir, month)
            
            # Ensure directories exist
            os.makedirs(month_dir, exist_ok=True)
            
            # Save to organized file structure
            filename = f"movement_{movement_number:04d}_{date}.txt"
            filepath = os.path.join(month_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(doc_content)
            
            # Show success message with file location
            messagebox.showinfo("âœ… ÎˆÎ³Î³ÏÎ±Ï†Î¿ Î”Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ", 
                               f"Î¤Î¿ Î­Î³Î³ÏÎ±Ï†Î¿ ÎºÎ¯Î½Î·ÏƒÎ·Ï‚ Î±Ï€Î¿Î¸Î·ÎºÎµÏÎ¸Î·ÎºÎµ:\n{filepath}\n\n"
                               f"Î¤Î¿ Î±ÏÏ‡ÎµÎ¯Î¿ Î¸Î± Î±Î½Î¿Î¯Î¾ÎµÎ¹ Î³Î¹Î± ÎµÎºÏ„ÏÏ€Ï‰ÏƒÎ·.")
            
            # Open file for printing
            try:
                os.startfile(filepath)  # Windows
            except AttributeError:
                try:
                    import subprocess
                    subprocess.run(['xdg-open', filepath])  # Linux
                except:
                    try:
                        subprocess.run(['open', filepath])  # macOS
                    except:
                        pass
            
            log_user_action("Movement document generated", f"File: {filename}")
            
        except Exception as e:
            logging.error(f"Error generating movement document: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î¯Î± ÎµÎ³Î³ÏÎ¬Ï†Î¿Ï…:\n{str(e)}")

    def _create_movement_document_content(self, movement_number, date, driver, vehicle, start_km, route, purpose):
        """Create the content for movement document"""
        # Format date
        try:
            date_obj = datetime.strptime(date, "%Y-%m-%d")
            formatted_date = date_obj.strftime("%d/%m/%Y")
        except:
            formatted_date = date
        
        # Create document content
        content = f"""
â•”â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•—
â•‘                    Î•Î“Î“Î¡Î‘Î¦ÎŸ ÎšÎ™ÎÎ—Î£Î—Î£ ÎŸÎ§Î—ÎœÎ‘Î¤ÎŸÎ£                                â•‘
â•šâ•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

 Î‘Î¡Î™Î˜ÎœÎŸÎ£ ÎšÎ™ÎÎ—Î£Î—Î£: {movement_number:04d}
ğŸ“… Î—ÎœÎ•Î¡ÎŸÎœÎ—ÎÎ™Î‘: {formatted_date}
ğŸ‘¤ ÎŸÎ”Î—Î“ÎŸÎ£: {driver}
ğŸš— ÎŸÎ§Î—ÎœÎ‘: {vehicle}
ï¸ Î§Î™Î›Î™ÎŸÎœÎ•Î¤Î¡Î‘ Î‘ÎÎ‘Î§Î©Î¡Î—Î£Î—Î£: {start_km:,} Ï‡Î»Î¼
ğŸ—ºï¸ Î”Î™Î‘Î”Î¡ÎŸÎœÎ—: {route or 'Î”ÎµÎ½ Î­Ï‡ÎµÎ¹ ÎºÎ±Î¸Î¿ÏÎ¹ÏƒÏ„ÎµÎ¯'}
 Î£ÎšÎŸÎ ÎŸÎ£: {purpose or 'Î”ÎµÎ½ Î¿ÏÎ¯ÏƒÏ„Î·ÎºÎµ'}

â•”â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•—
â•‘                              Î¥Î ÎŸÎ“Î¡Î‘Î¦Î•Î£                                     â•‘
â•šâ•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

ğŸ‘¤ Î¥Ï€Î¿Î³ÏÎ±Ï†Î® ÎŸÎ´Î·Î³Î¿Ï: _________________    ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±: _________________

 Î¥Ï€Î¿Î³ÏÎ±Ï†Î® Î¥Ï€ÎµÏ…Î¸ÏÎ½Î¿Ï…: _________________    ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±: _________________

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

ğŸ“‹ Î£Î—ÎœÎ•Î™Î©Î£Î•Î™Î£:
â€¢ Î¤Î¿ Ï€Î±ÏÏŒÎ½ Î­Î³Î³ÏÎ±Ï†Î¿ Ï€ÏÎ­Ï€ÎµÎ¹ Î½Î± ÏƒÏ…Î¼Ï€Î»Î·ÏÏ‰Î¸ÎµÎ¯ ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÏ€Î¹ÏƒÏ„ÏÎ¿Ï†Î® Ï„Î¿Ï… Î¿Ï‡Î®Î¼Î±Ï„Î¿Ï‚
â€¢ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ± ÎµÏ€Î¹ÏƒÏ„ÏÎ¿Ï†Î®Ï‚: _________________ Ï‡Î»Î¼
â€¢ Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ Ï‡Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±: _________________ Ï‡Î»Î¼
â€¢ ÎšÎ±ÏÏƒÎ¹Î¼Î± Ï€Î¿Ï… ÎºÎ±Ï„Î±Î½Î±Î»ÏÎ¸Î·ÎºÎµ: _________________ L

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

ğŸ“… Î”Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
ğŸ–¥ï¸ Î•Ï†Î±ÏÎ¼Î¿Î³Î®: Fleet Management System v2.0
"""
        
        return content

    def _clear_movement_form(self):
        """Clear movement form"""
        self.mov_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.mov_driver_combo.set("")
        self.mov_vehicle_combo.set("")
        self.mov_start_km_var.set("")
        self.mov_route_entry.delete(0, tk.END)
        if hasattr(self, 'mov_purpose_combobox'):
            self.mov_purpose_combobox.set("")
            # Î•Ï€Î±Î½Î±Ï†ÏŒÏÏ„Ï‰ÏƒÎ· ÏŒÎ»Ï‰Î½ Ï„Ï‰Î½ ÏƒÎºÎ¿Ï€ÏÎ½
            self.mov_purpose_combobox.set_values(self.db.get_purpose_names(active_only=True))
    
    def _auto_fill_last_km(self, *args):
        """Auto-fill last recorded kilometers for selected vehicle"""
        vehicle = self.mov_vehicle_combo.get()
        if not vehicle:
            return
        
        vehicle_id = self._get_vehicle_id(vehicle)
        if not vehicle_id:
            return
        
        try:
            self.db.cursor.execute("""
                SELECT MAX(COALESCE(end_km, start_km)) 
                FROM movements 
                WHERE vehicle_id = ? AND (end_km IS NOT NULL OR start_km IS NOT NULL)
            """, (vehicle_id,))
            
            result = self.db.cursor.fetchone()
            if result and result[0]:
                self.mov_start_km_var.set(str(result[0]))
                self.status_bar.set_status(f"Î‘Ï…Ï„ÏŒÎ¼Î±Ï„Î· ÏƒÏ…Î¼Ï€Î»Î®ÏÏ‰ÏƒÎ· Ï‡Î¹Î»Î¹Î¿Î¼Î­Ï„ÏÏ‰Î½ Î³Î¹Î± {vehicle}")
        except Exception as e:
            logging.error(f"Error auto-filling km: {e}")
    
    def _browse_movement_documents(self):
        """Browse and view movement documents organized by year/month"""
        base_dir = os.getcwd()
        movements_dir = os.path.join(base_dir, "ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚")
        
        # Check if movements directory exists
        if not os.path.exists(movements_dir):
            messagebox.showinfo("ğŸ“ Î¦Î¬ÎºÎµÎ»Î¿Ï‚ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½", 
                               "Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡Î¿Ï…Î½ Î±ÎºÏŒÎ¼Î· Î±Ï€Î¿Î¸Î·ÎºÎµÏ…Î¼Î­Î½ÎµÏ‚ ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚.\n"
                               "ÎŸ Ï†Î¬ÎºÎµÎ»Î¿Ï‚ 'ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚' Î¸Î± Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î·Î¸ÎµÎ¯ ÏŒÏ„Î±Î½ Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î®ÏƒÎµÏ„Îµ Ï„Î·Î½ Ï€ÏÏÏ„Î· ÎºÎ¯Î½Î·ÏƒÎ·.")
            return
        
        # Create browse window
        browse_window = tk.Toplevel(self.root)
        browse_window.title("ğŸ“ Î ÎµÏÎ¹Î®Î³Î·ÏƒÎ· ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½")
        browse_window.geometry("800x600")
        browse_window.configure(bg=THEMES[self.current_theme]["bg"])
        
        # Title
        title_frame = tk.Frame(browse_window, bg=THEMES[self.current_theme]["bg"])
        title_frame.pack(fill="x", padx=10, pady=5)
        tk.Label(
            title_frame,
            text="ğŸ“ Î ÎµÏÎ¹Î®Î³Î·ÏƒÎ· Î‘Ï€Î¿Î¸Î·ÎºÎµÏ…Î¼Î­Î½Ï‰Î½ ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½",
            font=FONT_TITLE,
            fg=THEMES[self.current_theme]["accent"],
            bg=THEMES[self.current_theme]["bg"],
        ).pack()
        
        # Create treeview for folder structure
        tree_frame = tk.Frame(browse_window, bg=THEMES[self.current_theme]["bg"])
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        tree = ttk.Treeview(tree_frame, columns=("type", "count"), show="tree headings")
        tree.heading("#0", text="Î¦Î¬ÎºÎµÎ»Î¿Ï‚/Î‘ÏÏ‡ÎµÎ¯Î¿")
        tree.heading("type", text="Î¤ÏÏ€Î¿Ï‚")
        tree.heading("count", text="Î£Ï„Î¿Î¹Ï‡ÎµÎ¯Î±")
        
        tree.column("#0", width=400)
        tree.column("type", width=100, anchor="center")
        tree.column("count", width=80, anchor="center")
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Load folder structure
        try:
            # Root folder
            root_id = tree.insert("", "end", text="ğŸ“ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", values=("Î¦Î¬ÎºÎµÎ»Î¿Ï‚", ""), open=True)
            
            # Get years
            years = sorted([d for d in os.listdir(movements_dir) 
                           if os.path.isdir(os.path.join(movements_dir, d))], reverse=True)
            
            total_files = 0
            for year in years:
                year_path = os.path.join(movements_dir, year)
                months = sorted([d for d in os.listdir(year_path) 
                               if os.path.isdir(os.path.join(year_path, d))])
                
                year_count = 0
                year_id = tree.insert(root_id, "end", text=f"ğŸ“… {year}", values=("ÎˆÏ„Î¿Ï‚", ""), open=True)
                
                for month in months:
                    month_path = os.path.join(year_path, month)
                    files = [f for f in os.listdir(month_path) if f.endswith('.txt')]
                    month_count = len(files)
                    year_count += month_count
                    total_files += month_count
                    
                    # Greek month names
                    month_names = {
                        "01": "Î™Î±Î½Î¿Ï…Î¬ÏÎ¹Î¿Ï‚", "02": "Î¦ÎµÎ²ÏÎ¿Ï…Î¬ÏÎ¹Î¿Ï‚", "03": "ÎœÎ¬ÏÏ„Î¹Î¿Ï‚",
                        "04": "Î‘Ï€ÏÎ¯Î»Î¹Î¿Ï‚", "05": "ÎœÎ¬Î¹Î¿Ï‚", "06": "Î™Î¿ÏÎ½Î¹Î¿Ï‚",
                        "07": "Î™Î¿ÏÎ»Î¹Î¿Ï‚", "08": "Î‘ÏÎ³Î¿Ï…ÏƒÏ„Î¿Ï‚", "09": "Î£ÎµÏ€Ï„Î­Î¼Î²ÏÎ¹Î¿Ï‚",
                        "10": "ÎŸÎºÏ„ÏÎ²ÏÎ¹Î¿Ï‚", "11": "ÎÎ¿Î­Î¼Î²ÏÎ¹Î¿Ï‚", "12": "Î”ÎµÎºÎ­Î¼Î²ÏÎ¹Î¿Ï‚"
                    }
                    month_name = month_names.get(month, month)
                    
                    month_id = tree.insert(year_id, "end", 
                                         text=f"ğŸ“‚ {month_name} ({month})", 
                                         values=("ÎœÎ®Î½Î±Ï‚", f"{month_count} ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚"))
                    
                    # Add files
                    for file in sorted(files):
                        file_path = os.path.join(month_path, file)
                        tree.insert(month_id, "end", 
                                   text=f"ğŸ“„ {file}", 
                                   values=("ÎšÎ¯Î½Î·ÏƒÎ·", ""),
                                   tags=(file_path,))
                
                # Update year count
                tree.set(year_id, "count", f"{year_count} ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚")
            
            # Update root count
            tree.set(root_id, "count", f"{total_files} ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚")
            
        except Exception as e:
            logging.error(f"Error loading movement folders: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· Ï†Î±ÎºÎ­Î»Ï‰Î½: {str(e)}")
        
        # Double-click to open file
        def on_double_click(event):
            item = tree.selection()[0]
            tags = tree.item(item, 'tags')
            if tags and os.path.isfile(tags[0]):
                try:
                    os.startfile(tags[0])  # Windows
                except:
                    try:
                        import subprocess
                        subprocess.run(['xdg-open', tags[0]])  # Linux
                    except:
                        try:
                            subprocess.run(['open', tags[0]])  # macOS
                        except:
                            messagebox.showinfo("ğŸ“„ Î‘ÏÏ‡ÎµÎ¯Î¿", f"Î‘ÏÏ‡ÎµÎ¯Î¿: {tags[0]}")
        
        tree.bind("<Double-1>", on_double_click)
        
        # Buttons frame
        buttons_frame = tk.Frame(browse_window, bg=THEMES[self.current_theme]["bg"])
        buttons_frame.pack(fill="x", padx=10, pady=5)
        
        ModernButton(
            buttons_frame,
            text="ğŸ“‚ Î†Î½Î¿Î¹Î³Î¼Î± Î¦Î±ÎºÎ­Î»Î¿Ï… ÎšÎ¹Î½Î®ÏƒÎµÏ‰Î½",
            style="info",
            command=lambda: os.startfile(movements_dir) if os.path.exists(movements_dir) else None
        ).pack(side="left", padx=5)
        
        ModernButton(
            buttons_frame,
            text="âŒ ÎšÎ»ÎµÎ¯ÏƒÎ¹Î¼Î¿",
            style="secondary",
            command=browse_window.destroy
        ).pack(side="right", padx=5)

    def _load_movements(self):
        """Load movements with improved filtering and performance"""
        try:
            # Clear trees
            for item in self.tree_active.get_children():
                self.tree_active.delete(item)
            for item in self.tree_completed.get_children():
                self.tree_completed.delete(item)
            
            # Get search terms
            active_search = self.active_search_var.get().lower() if hasattr(self, 'active_search_var') else ""
            completed_search = self.completed_search_var.get().lower() if hasattr(self, 'completed_search_var') else ""
            
            # Load active movements (not returned)
            self.db.cursor.execute("""
                SELECT m.id, COALESCE(m.movement_number, 0), m.date, 
                       d.name || ' ' || d.surname as driver, v.plate, 
                       m.start_km, m.route, m.purpose
                FROM movements m
                JOIN drivers d ON m.driver_id = d.id
                JOIN vehicles v ON m.vehicle_id = v.id
                WHERE m.end_km IS NULL
                ORDER BY m.date DESC, m.id DESC
            """)
            
            for row in self.db.cursor.fetchall():
                # Apply search filter
                if active_search:
                    searchable_text = f"{row[2]} {row[3]} {row[4]} {row[6]} {row[7]}".lower()
                    if active_search not in searchable_text:
                        continue
                
                mov_num_display = f"{row[1]:04d}" if row[1] > 0 else "---"
                display_values = (mov_num_display,) + row[2:]
                self.tree_active.insert("", "end", values=display_values, tags=(row[0],))
            
            # Load completed movements (today only)
            today = datetime.now().strftime("%Y-%m-%d")
            self.db.cursor.execute("""
                SELECT m.id, COALESCE(m.movement_number, 0), m.date, 
                       d.name || ' ' || d.surname as driver, v.plate, 
                       m.start_km, m.end_km, m.route, m.purpose,
                       (COALESCE(m.end_km, 0) - COALESCE(m.start_km, 0)) as total_km
                FROM movements m
                JOIN drivers d ON m.driver_id = d.id
                JOIN vehicles v ON m.vehicle_id = v.id
                WHERE m.end_km IS NOT NULL AND m.date = ?
                ORDER BY m.id DESC
            """, (today,))
            
            for row in self.db.cursor.fetchall():
                # Apply search filter
                if completed_search:
                    searchable_text = f"{row[2]} {row[3]} {row[4]} {row[7]} {row[8]}".lower()
                    if completed_search not in searchable_text:
                        continue
                
                mov_num_display = f"{row[1]:04d}" if row[1] > 0 else "---"
                display_values = (mov_num_display,) + row[2:]
                self.tree_completed.insert("", "end", values=display_values, tags=(row[0],))
            
            self.status_bar.set_status("ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ ÎµÎ½Î·Î¼ÎµÏÏÎ¸Î·ÎºÎ±Î½")
            
        except Exception as e:
            logging.error(f"Error loading movements: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î· Ï†ÏŒÏÏ„Ï‰ÏƒÎ· ÎºÎ¹Î½Î®ÏƒÎµÏ‰Î½: {str(e)}")
    
    # Utility methods
    def _get_driver_id(self, driver_name):
        """Get driver ID from name"""
        try:
            parts = driver_name.split(' ', 1)
            if len(parts) == 2:
                name, surname = parts
                self.db.cursor.execute(
                    "SELECT id FROM drivers WHERE name = ? AND surname = ?",
                    (name, surname)
                )
                result = self.db.cursor.fetchone()
                return result[0] if result else None
        except Exception as e:
            logging.error(f"Error getting driver ID: {e}")
        return None
    
    def _get_vehicle_id(self, plate):
        """Get vehicle ID from plate"""
        try:
            self.db.cursor.execute("SELECT id FROM vehicles WHERE plate = ?", (plate,))
            result = self.db.cursor.fetchone()
            return result[0] if result else None
        except Exception as e:
            logging.error(f"Error getting vehicle ID: {e}")
        return None
    
    def _is_vehicle_active(self, vehicle_id):
        """Check if vehicle has active movement"""
        try:
            self.db.cursor.execute(
                "SELECT COUNT(*) FROM movements WHERE vehicle_id = ? AND end_km IS NULL",
                (vehicle_id,)
            )
            return self.db.cursor.fetchone()[0] > 0
        except Exception as e:
            logging.error(f"Error checking vehicle status: {e}")
        return False
    
    def _refresh_all_data(self):
        """Refresh all data in the application"""
        try:
            progress = ProgressDialog(self.root, "Î‘Î½Î±Î½Î­Ï‰ÏƒÎ· Î”ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½", "Î¦ÏŒÏÏ„Ï‰ÏƒÎ· Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½...")
            
            progress.update_message("Î¦ÏŒÏÏ„Ï‰ÏƒÎ· ÎºÎ¹Î½Î®ÏƒÎµÏ‰Î½...")
            self._load_movements()
            
            progress.update_message("Î¦ÏŒÏÏ„Ï‰ÏƒÎ· Î¿Ï‡Î·Î¼Î¬Ï„Ï‰Î½...")
            self._refresh_movement_combos()
            
            progress.update_message("Î•Î½Î·Î¼Î­ÏÏ‰ÏƒÎ· Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚...")
            self._update_tank_level()
            
            progress.close()
            self.status_bar.set_status("ÎŒÎ»Î± Ï„Î± Î´ÎµÎ´Î¿Î¼Î­Î½Î± Î±Î½Î±Î½ÎµÏÎ¸Î·ÎºÎ±Î½")
            
        except Exception as e:
            if 'progress' in locals():
                progress.close()
            logging.error(f"Error refreshing data: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ Î±Î½Î±Î½Î­Ï‰ÏƒÎ·: {str(e)}")
    
    def _refresh_movement_combos(self):
        """Refresh movement combo boxes"""
        try:
            # Get drivers
            self.db.cursor.execute("SELECT id, name, surname FROM drivers ORDER BY name, surname")
            drivers = []
            self.driver_ids = {}
            
            for row in self.db.cursor.fetchall():
                driver_text = f"{row[1]} {row[2]}"
                drivers.append(driver_text)
                self.driver_ids[driver_text] = row[0]
            
            # Get vehicles
            self.db.cursor.execute("SELECT id, plate FROM vehicles ORDER BY plate")
            vehicles = []
            self.vehicle_ids = {}
            
            for row in self.db.cursor.fetchall():
                vehicles.append(row[1])
                self.vehicle_ids[row[1]] = row[0]
            
            # Update combo boxes
            self.mov_driver_combo.set_values(drivers)
            self.mov_vehicle_combo.set_values(vehicles)
            
        except Exception as e:
            logging.error(f"Error refreshing combos: {e}")
    
    def _update_tank_level(self):
        """Update tank level display"""
        try:
            self.db.cursor.execute(
                "SELECT SUM(CASE WHEN type = 'fill' THEN liters ELSE -liters END) FROM tank"
            )
            level = self.db.cursor.fetchone()[0] or 0
            
            # Calculate percentage
            percentage = (level / TANK_CAPACITY) * 100 if TANK_CAPACITY > 0 else 0
            
            # Determine status
            if level < TANK_MIN_LEVEL:
                color = THEMES[self.current_theme]["danger"]
                icon = "ğŸ”´"
                status = "Î§Î‘ÎœÎ—Î›ÎŸ"
            else:
                color = THEMES[self.current_theme]["success"]
                icon = "ğŸŸ¢"
                status = "ÎšÎ‘Î›ÎŸ"
            
            # Update status bar indicator if it exists
            if hasattr(self, 'tank_indicator'):
                self.tank_indicator.config(
                    text=f"{icon} Î”ÎµÎ¾Î±Î¼ÎµÎ½Î®: {level:.1f}L ({status})",
                    fg=color
                )
            
            # Update tank level label in fuel tab if it exists
            if hasattr(self, 'tank_level_label'):
                tank_text = f"ğŸ“Š Î¤ÏÎ­Ï‡Î¿Î½ Î•Ï€Î¯Ï€ÎµÎ´Î¿: {level:.1f}L / {TANK_CAPACITY:,.0f}L ({percentage:.1f}%)"
                self.tank_level_label.config(
                    text=tank_text,
                    fg=color
                )
            
            # Update progress bar if it exists
            if hasattr(self, 'tank_progress'):
                self.tank_progress['value'] = percentage
                
                # Style progress bar based on level
                style = ttk.Style()
                if level < TANK_MIN_LEVEL:
                    style.configure("Tank.Horizontal.TProgressbar",
                                  background=THEMES[self.current_theme]["danger"])
                elif percentage < 30:
                    style.configure("Tank.Horizontal.TProgressbar",
                                  background=THEMES[self.current_theme]["warning"])
                else:
                    style.configure("Tank.Horizontal.TProgressbar",
                                  background=THEMES[self.current_theme]["success"])
                
                self.tank_progress.configure(style="Tank.Horizontal.TProgressbar")
            
        except Exception as e:
            logging.error(f"Error updating tank level: {e}")
            if hasattr(self, 'tank_indicator'):
                self.tank_indicator.config(text="â›½ Î£Ï†Î¬Î»Î¼Î± Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚", fg="red")
            if hasattr(self, 'tank_level_label'):
                self.tank_level_label.config(text="âš ï¸ Î£Ï†Î¬Î»Î¼Î± Ï†ÏŒÏÏ„Ï‰ÏƒÎ·Ï‚ Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½", fg="red")
            if hasattr(self, 'tank_progress'):
                self.tank_progress['value'] = 0
    
    # Theme and UI methods
    def _change_theme(self, theme_name):
        """Change application theme"""
        try:
            if theme_name not in THEMES:
                self.status_bar.set_status(f"Î†Î³Î½Ï‰ÏƒÏ„Î¿ Î¸Î­Î¼Î±: {theme_name}", "error")
                return
                
            self.current_theme = theme_name
            theme = THEMES[theme_name]
            
            # Save theme preference
            from config import save_user_setting
            save_user_setting("theme", theme_name)
            
            # Update root window
            self.root.configure(bg=theme["bg"])
            
            # Update all widgets recursively
            self._update_widget_theme(self.root, theme)
            
            # Update menu bar to reflect new theme
            self._create_menu_bar()
            
            theme_names = {
                "light": "Î¦Ï‰Ï„ÎµÎ¹Î½ÏŒ",
                "dark": "Î£ÎºÎ¿Ï„ÎµÎ¹Î½ÏŒ", 
                "blue": "ÎœÏ€Î»Îµ",
                "green": "Î ÏÎ¬ÏƒÎ¹Î½Î¿",
                "purple": "ÎœÏ‰Î²"
            }
            
            self.status_bar.set_status(f"Î˜Î­Î¼Î± Î¬Î»Î»Î±Î¾Îµ ÏƒÎµ: {theme_names.get(theme_name, theme_name)}", "success")
            log_user_action("Theme changed", theme_name)
            
        except Exception as e:
            logging.error(f"Error changing theme: {e}")
            self.status_bar.set_status("Î£Ï†Î¬Î»Î¼Î± Î±Î»Î»Î±Î³Î®Ï‚ Î¸Î­Î¼Î±Ï„Î¿Ï‚", "error")
    
    def _update_widget_theme(self, widget, theme):
        """Recursively update widget themes"""
        try:
            widget_class = widget.winfo_class()
            
            # Handle different widget types with modern styling
            if widget_class in ["Frame", "Toplevel"]:
                # Check if it's a modern component
                if hasattr(widget, 'theme'):
                    widget.theme = self.current_theme
                if hasattr(widget, 'is_card') and widget.is_card:
                    widget.configure(bg=theme["card_bg"])
                elif hasattr(widget, '_apply_modern_styling'):
                    widget.configure(bg=theme["frame_bg"])
                    widget._apply_modern_styling()
                else:
                    widget.configure(bg=theme["bg"])
                    
            elif widget_class == "Label":
                # Handle different label types
                widget.configure(bg=theme["bg"], fg=theme["fg"])
                
            elif widget_class == "Button":
                # Handle modern buttons
                if hasattr(widget, 'theme'):
                    widget.theme = self.current_theme
                if hasattr(widget, 'style_name'):
                    # Update modern button styling
                    style = widget.style_name
                    if style in BUTTON_STYLES:
                        button_style = BUTTON_STYLES[style]
                        widget.configure(
                            bg=button_style["bg"], 
                            fg=button_style["fg"],
                            activebackground=button_style["activebackground"]
                        )
                        widget.original_bg = button_style["bg"]
                        widget.hover_bg = button_style["activebackground"]
                else:
                    widget.configure(bg=theme["button_bg_solid"], fg=theme["button_fg"])
                    
            elif widget_class == "Entry":
                # Handle modern entries
                if hasattr(widget, 'theme'):
                    widget.theme = self.current_theme
                    widget.configure(
                        bg=theme["entry_bg"], 
                        fg=theme["fg"],
                        highlightbackground=theme["entry_border"],
                        highlightcolor=theme["entry_focus"]
                    )
                else:
                    widget.configure(bg=theme["entry_bg"], fg=theme["fg"])
            
            # Update children
            for child in widget.winfo_children():
                self._update_widget_theme(child, theme)
                
        except Exception as e:
            logging.error(f"Error updating widget theme: {e}")
    
    # System methods
    def _backup_database(self):
        """Create database backup"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"fleet_backup_{timestamp}.db"
            
            backup_path = filedialog.asksaveasfilename(
                defaultextension=".db",
                filetypes=[("Database files", "*.db"), ("All files", "*.*")],
                initialname=backup_filename
            )
            
            if backup_path:
                if self.db.backup_database(backup_path):
                    messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", f"Backup Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚:\n{backup_path}")
                    self.status_bar.set_status("Backup Î¿Î»Î¿ÎºÎ»Î·ÏÏÎ¸Î·ÎºÎµ")
                else:
                    messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î‘Ï€Î¿Ï„Ï…Ï‡Î¯Î± Î´Î·Î¼Î¹Î¿Ï…ÏÎ³Î¯Î±Ï‚ backup!")
        except Exception as e:
            logging.error(f"Backup error: {e}")
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± backup: {str(e)}")
    
    def _show_system_stats(self):
        """Show comprehensive system statistics"""
        # Implementation for system statistics
        pass
    
    def _cleanup_old_data(self):
        """Cleanup old data with user confirmation"""
        # Implementation for data cleanup
        pass
    
    def _show_help(self):
        """Show help dialog"""
        help_text = """
ğŸš— ÎŸÎ”Î—Î“Î™Î•Î£ Î§Î¡Î—Î£Î—Î£ Î£Î¥Î£Î¤Î—ÎœÎ‘Î¤ÎŸÎ£ Î”Î™Î‘Î§Î•Î™Î¡Î™Î£Î—Î£ Î£Î¤ÎŸÎ›ÎŸÎ¥

ğŸ“‹ Î’Î‘Î£Î™ÎšÎ•Î£ Î›Î•Î™Î¤ÎŸÎ¥Î¡Î“Î™Î•Î£:
â€¢ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚: Î”Î·Î¼Î¹Î¿Ï…ÏÎ³Î¯Î± ÎºÎ±Î¹ Ï€Î±ÏÎ±ÎºÎ¿Î»Î¿ÏÎ¸Î·ÏƒÎ· ÎºÎ¹Î½Î®ÏƒÎµÏ‰Î½ Î¿Ï‡Î·Î¼Î¬Ï„Ï‰Î½
â€¢ ÎŸÏ‡Î®Î¼Î±Ï„Î±: Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· ÏƒÏ„ÏŒÎ»Î¿Ï… Î¿Ï‡Î·Î¼Î¬Ï„Ï‰Î½
â€¢ ÎŸÎ´Î·Î³Î¿Î¯: Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· Ï€ÏÎ¿ÏƒÏ‰Ï€Î¹ÎºÎ¿Ï Î¿Î´Î·Î³ÏÎ½
â€¢ ÎšÎ±ÏÏƒÎ¹Î¼Î±: Î Î±ÏÎ±ÎºÎ¿Î»Î¿ÏÎ¸Î·ÏƒÎ· ÎºÎ±Ï„Î±Î½Î¬Î»Ï‰ÏƒÎ·Ï‚ ÎºÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½
â€¢ Î‘Î½Î±Ï†Î¿ÏÎ­Ï‚: Î•Î¾Î±Î³Ï‰Î³Î® Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½ ÎºÎ±Î¹ ÏƒÏ„Î±Ï„Î¹ÏƒÏ„Î¹ÎºÏÎ½

âŒ¨ï¸ Î£Î¥ÎÎ¤ÎŸÎœÎ•Î¥Î£Î•Î™Î£ Î Î›Î—ÎšÎ¤Î¡ÎŸÎ›ÎŸÎ“Î™ÎŸÎ¥:
â€¢ F5: Î‘Î½Î±Î½Î­Ï‰ÏƒÎ· Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½
â€¢ Ctrl+S: Backup Î²Î¬ÏƒÎ·Ï‚ Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½
â€¢ Ctrl+Q: ÎˆÎ¾Î¿Î´Î¿Ï‚ Î±Ï€ÏŒ Ï„Î·Î½ ÎµÏ†Î±ÏÎ¼Î¿Î³Î®

ğŸ’¡ Î£Î¥ÎœÎ’ÎŸÎ¥Î›Î•Î£:
â€¢ Î§ÏÎ·ÏƒÎ¹Î¼Î¿Ï€Î¿Î¹Î®ÏƒÏ„Îµ Ï„Î· Î»ÎµÎ¹Ï„Î¿Ï…ÏÎ³Î¯Î± Î±Î½Î±Î¶Î®Ï„Î·ÏƒÎ·Ï‚ Î³Î¹Î± Î³ÏÎ®Î³Î¿ÏÎ· ÎµÏÏÎµÏƒÎ·
â€¢ ÎšÎ¬Î½Ï„Îµ Ï„Î±ÎºÏ„Î¹ÎºÎ¬ backup Ï„Î·Ï‚ Î²Î¬ÏƒÎ·Ï‚ Î´ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½
â€¢ Î•Î»Î­Î³Ï‡ÎµÏ„Îµ Ï„Î¿ ÎµÏ€Î¯Ï€ÎµÎ´Î¿ Ï„Î·Ï‚ Î´ÎµÎ¾Î±Î¼ÎµÎ½Î®Ï‚ ÎºÎ±Ï…ÏƒÎ¯Î¼Ï‰Î½
        """
        
        help_window = tk.Toplevel(self.root)
        help_window.title("ğŸ“– ÎŸÎ´Î·Î³Î¯ÎµÏ‚ Î§ÏÎ®ÏƒÎ·Ï‚")
        help_window.geometry("600x500")
        
        text_widget = tk.Text(help_window, font=FONT_NORMAL, wrap=tk.WORD, padx=20, pady=20)
        text_widget.pack(fill="both", expand=True)
        text_widget.insert(tk.END, help_text)
        text_widget.config(state=tk.DISABLED)
    
    def _show_about(self):
        """Show about dialog"""
        about_text = f"""
ğŸš— Î£ÏÏƒÏ„Î·Î¼Î± Î”Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ·Ï‚ Î£Ï„ÏŒÎ»Î¿Ï…
ÎˆÎºÎ´Î¿ÏƒÎ· 2.0 - Î’ÎµÎ»Ï„Î¹Ï‰Î¼Î­Î½Î·

Î”Î·Î¼Î¹Î¿Ï…ÏÎ³Î®Î¸Î·ÎºÎµ Î³Î¹Î± Ï„Î· Î´Î¹Î±Ï‡ÎµÎ¯ÏÎ¹ÏƒÎ· ÏƒÏ„ÏŒÎ»Î¿Ï… Î¿Ï‡Î·Î¼Î¬Ï„Ï‰Î½
Î¼Îµ ÏƒÏÎ³Ï‡ÏÎ¿Î½Î· Î´Î¹ÎµÏ€Î±Ï†Î® ÎºÎ±Î¹ Î²ÎµÎ»Ï„Î¹Ï‰Î¼Î­Î½ÎµÏ‚ Î»ÎµÎ¹Ï„Î¿Ï…ÏÎ³Î¯ÎµÏ‚.

ğŸ“… Î—Î¼ÎµÏÎ¿Î¼Î·Î½Î¯Î±: {datetime.now().strftime('%d/%m/%Y')}
ğŸ Python Version: {sys.version.split()[0]}
ğŸ’¾ Î’Î¬ÏƒÎ· Î”ÎµÎ´Î¿Î¼Î­Î½Ï‰Î½: SQLite3

Â© 2024 Fleet Management System
        """
        
        messagebox.showinfo("â„¹ï¸ Î£Ï‡ÎµÏ„Î¹ÎºÎ¬ Î¼Îµ Ï„Î·Î½ Î•Ï†Î±ÏÎ¼Î¿Î³Î®", about_text)
    
    def _edit_movement_return(self, event):
        """Edit movement return with improved UX"""
        selection = self.tree_active.selection()
        if not selection:
            messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ ÎºÎ¯Î½Î·ÏƒÎ·!")
            return

        item = selection[0]
        values = self.tree_active.item(item, "values")
        movement_num = values[0]
        start_km = values[4]

        # Î Î±ÏÎ¬Î¸Ï…ÏÎ¿ Î³Î¹Î± ÎºÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ· Ï‡Î»Î¼ ÎµÏ€Î¹ÏƒÏ„ÏÎ¿Ï†Î®Ï‚
        def submit():
            end_km = entry_end_km.get().strip()
            if not end_km.isdigit() or int(end_km) < int(start_km):
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", "Î¤Î± Ï‡Î»Î¼ ÎµÏ€Î¹ÏƒÏ„ÏÎ¿Ï†Î®Ï‚ Ï€ÏÎ­Ï€ÎµÎ¹ Î½Î± ÎµÎ¯Î½Î±Î¹ Î±ÏÎ¹Î¸Î¼ÏŒÏ‚ ÎºÎ±Î¹ >= Ï‡Î»Î¼ Î±Î½Î±Ï‡ÏÏÎ·ÏƒÎ·Ï‚!")
                return
            try:
                self.db.cursor.execute(
                    "UPDATE movements SET end_km=? WHERE movement_number=?",
                    (int(end_km), movement_num)
                )
                self.db.conn.commit()
                top.destroy()
                self._load_movements()
                self.status_bar.set_status(f"Î— ÎºÎ¯Î½Î·ÏƒÎ· {movement_num} Î­ÎºÎ»ÎµÎ¹ÏƒÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚.")
                messagebox.showinfo("âœ… Î•Ï€Î¹Ï„Ï…Ï‡Î¯Î±", "Î— ÎºÎ¯Î½Î·ÏƒÎ· Î­ÎºÎ»ÎµÎ¹ÏƒÎµ ÎµÏ€Î¹Ï„Ï…Ï‡ÏÏ‚!")
            except Exception as e:
                messagebox.showerror("âŒ Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎºÎ±Ï„Î¬ Ï„Î·Î½ ÎµÎ½Î·Î¼Î­ÏÏ‰ÏƒÎ·: {str(e)}")

        top = tk.Toplevel(self.root)
        top.title("ÎšÎ»ÎµÎ¯ÏƒÎ¹Î¼Î¿ ÎšÎ¯Î½Î·ÏƒÎ·Ï‚")
        top.geometry("300x180")
        top.configure(bg=THEMES[self.current_theme]["bg"])
        top.grab_set()
        
        # Center the window
        top.transient(self.root)
        top.focus_set()
        
        # Title label
        title_label = tk.Label(top, 
                              text=f"ğŸ“ ÎšÎ»ÎµÎ¯ÏƒÎ¹Î¼Î¿ ÎšÎ¯Î½Î·ÏƒÎ·Ï‚ #{movement_num}", 
                              font=FONT_SUBTITLE,
                              bg=THEMES[self.current_theme]["bg"],
                              fg=THEMES[self.current_theme]["accent"])
        title_label.pack(padx=10, pady=(10, 5))
        
        # Start km info
        tk.Label(top, 
                text=f"ğŸ›£ï¸ Î§Î»Î¼ Î‘Î½Î±Ï‡ÏÏÎ·ÏƒÎ·Ï‚: {start_km}", 
                font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"],
                fg=THEMES[self.current_theme]["fg"]).pack(padx=10, pady=(2, 2))
        
        # End km input
        tk.Label(top, 
                text="ğŸ Î§Î»Î¼ Î•Ï€Î¹ÏƒÏ„ÏÎ¿Ï†Î®Ï‚:", 
                font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["bg"],
                fg=THEMES[self.current_theme]["fg"]).pack(padx=10, pady=(2, 2))
        
        entry_end_km = tk.Entry(top, 
                               font=FONT_NORMAL,
                               bg=THEMES[self.current_theme]["entry_bg"],
                               fg=THEMES[self.current_theme]["fg"],
                               relief="flat",
                               borderwidth=1,
                               highlightthickness=1,
                               highlightcolor=THEMES[self.current_theme]["accent"])
        entry_end_km.pack(padx=10, pady=(2, 10), fill="x")
        entry_end_km.focus_set()
        
        # Add tooltip
        if hasattr(self, 'tooltip_manager'):
            self.tooltip_manager.add_tooltip(entry_end_km, "Î Î»Î·ÎºÏ„ÏÎ¿Î»Î¿Î³Î®ÏƒÏ„Îµ Ï„Î± Ï‡Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ± ÎµÏ€Î¹ÏƒÏ„ÏÎ¿Ï†Î®Ï‚ ÎºÎ±Î¹ Ï€Î±Ï„Î®ÏƒÏ„Îµ Enter")
        
        # Bind Enter key to submit
        entry_end_km.bind('<Return>', lambda event: submit())
        entry_end_km.bind('<KP_Enter>', lambda event: submit())  # Numeric keypad Enter
        
        # Buttons frame
        buttons_frame = tk.Frame(top, bg=THEMES[self.current_theme]["bg"])
        buttons_frame.pack(fill="x", padx=10, pady=(0, 10))
        
        btn_submit = ModernButton(buttons_frame, 
                                 text="âœ… ÎšÎ±Ï„Î±Ï‡ÏÏÎ·ÏƒÎ·", 
                                 style="success", 
                                 command=submit)
        btn_submit.pack(side="left", padx=(0, 5))
        
        btn_cancel = ModernButton(buttons_frame, 
                                 text="âŒ Î‘ÎºÏÏÏ‰ÏƒÎ·", 
                                 style="secondary", 
                                 command=top.destroy)
        btn_cancel.pack(side="right", padx=(5, 0))
    
    def _edit_completed_movement(self, event):
        """Edit completed movement"""
        # Implementation for editing completed movements
        pass
    
    # Driver Analytics Methods
    def _load_analytics_drivers(self):
        """Load drivers for analytics selection"""
        try:
            drivers = self.db.get_all_drivers()
            driver_names = [f"{driver[1]} (ID: {driver[0]})" for driver in drivers]
            self.analytics_driver_combo.set_values(driver_names)
            if driver_names:
                self.analytics_driver_combo.set(driver_names[0])
        except Exception as e:
            logging.error(f"Error loading analytics drivers: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± Ï†ÏŒÏÏ„Ï‰ÏƒÎ·Ï‚ Î¿Î´Î·Î³ÏÎ½: {e}")
    
    def _show_driver_analytics(self):
        """Show analytics for selected driver"""
        try:
            # Get selected driver ID
            selected_driver = self.analytics_driver_combo.get()
            if not selected_driver:
                messagebox.showwarning("Î ÏÎ¿ÎµÎ¹Î´Î¿Ï€Î¿Î¯Î·ÏƒÎ·", "Î Î±ÏÎ±ÎºÎ±Î»Ï ÎµÏ€Î¹Î»Î­Î¾Ï„Îµ Î¿Î´Î·Î³ÏŒ")
                return
            
            driver_id = int(selected_driver.split("ID: ")[1].split(")")[0])
            
            # Get date range
            start_date = self.analytics_start_date.get() if self.analytics_start_date.get() else None
            end_date = self.analytics_end_date.get() if self.analytics_end_date.get() else None
            
            # Get analytics data
            analytics = self.db.get_driver_analytics(driver_id, start_date, end_date)
            if not analytics:
                messagebox.showinfo("Î Î»Î·ÏÎ¿Ï†Î¿ÏÎ¯Î±", "Î”ÎµÎ½ Î²ÏÎ­Î¸Î·ÎºÎ±Î½ Î´ÎµÎ´Î¿Î¼Î­Î½Î± Î³Î¹Î± Ï„Î¿Î½ ÎµÏ€Î¹Î»ÎµÎ³Î¼Î­Î½Î¿ Î¿Î´Î·Î³ÏŒ")
                return
            
            # Display analytics
            self._display_driver_analytics(analytics)
            
        except Exception as e:
            logging.error(f"Error showing driver analytics: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎµÎ¼Ï†Î¬Î½Î¹ÏƒÎ·Ï‚ Î±Î½Î±Î»Ï…Ï„Î¹ÎºÏÎ½: {e}")
    
    def _display_driver_analytics(self, analytics):
        """Display driver analytics in the UI"""
        # Clear previous results
        for widget in self.summary_frame.winfo_children():
            widget.destroy()
        for widget in self.details_frame.winfo_children():
            widget.destroy()
        
        # Summary Tab
        summary_canvas = tk.Canvas(self.summary_frame, bg=THEMES[self.current_theme]["frame_bg"])
        summary_scrollbar = ttk.Scrollbar(self.summary_frame, orient="vertical", command=summary_canvas.yview)
        summary_canvas.configure(yscrollcommand=summary_scrollbar.set)
        
        summary_content = tk.Frame(summary_canvas, bg=THEMES[self.current_theme]["frame_bg"])
        summary_canvas.create_window((0, 0), window=summary_content, anchor="nw")
        
        # Driver info
        info_frame = tk.LabelFrame(summary_content, text=f"ğŸ‘¤ {analytics['driver_name']}", 
                                  font=FONT_SUBTITLE, bg=THEMES[self.current_theme]["frame_bg"],
                                  fg=THEMES[self.current_theme]["accent"])
        info_frame.pack(fill="x", padx=10, pady=10)
        
        # Key metrics in a grid
        metrics_frame = tk.Frame(info_frame, bg=THEMES[self.current_theme]["frame_bg"])
        metrics_frame.pack(fill="x", padx=10, pady=10)
        
        # Row 1
        self._create_metric_label(metrics_frame, "ğŸš— Î£Ï…Î½Î¿Î»Î¹ÎºÎ­Ï‚ ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", 
                                 str(analytics['total_movements']), 0, 0)
        self._create_metric_label(metrics_frame, "ğŸ“ Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", 
                                 f"{analytics['total_kilometers']:,.0f} km", 0, 1)
        
        # Row 2
        self._create_metric_label(metrics_frame, "â›½ Î£Ï…Î½Î¿Î»Î¹ÎºÎ¬ ÎšÎ±ÏÏƒÎ¹Î¼Î±", 
                                 f"{analytics['total_fuel_liters']:.1f} L", 1, 0)
        self._create_metric_label(metrics_frame, "ğŸ’° Î£Ï…Î½Î¿Î»Î¹ÎºÏŒ ÎšÏŒÏƒÏ„Î¿Ï‚", 
                                 f"{analytics['total_fuel_cost']:.2f} â‚¬", 1, 1)
        
        # Row 3
        self._create_metric_label(metrics_frame, "ğŸ“Š ÎœÎ­ÏƒÎ· ÎšÎ±Ï„Î±Î½Î¬Î»Ï‰ÏƒÎ·", 
                                 f"{analytics['avg_consumption_km_per_liter']:.2f} km/L", 2, 0)
        self._create_metric_label(metrics_frame, "ğŸ’¸ ÎšÏŒÏƒÏ„Î¿Ï‚/km", 
                                 f"{analytics['avg_cost_per_km']:.3f} â‚¬/km", 2, 1)
        
        # Most used vehicles
        if analytics['most_used_vehicles']:
            vehicles_frame = tk.LabelFrame(summary_content, text="ğŸš™ Î Î¹Î¿ Î£Ï…Ï‡Î½Î¬ Î§ÏÎ·ÏƒÎ¹Î¼Î¿Ï€Î¿Î¹Î¿ÏÎ¼ÎµÎ½Î± ÎŸÏ‡Î®Î¼Î±Ï„Î±", 
                                          font=FONT_SUBTITLE, bg=THEMES[self.current_theme]["frame_bg"],
                                          fg=THEMES[self.current_theme]["accent"])
            vehicles_frame.pack(fill="x", padx=10, pady=10)
            
            for i, vehicle in enumerate(analytics['most_used_vehicles'][:3]):
                vehicle_info = f"ğŸš— {vehicle['plate']} - {vehicle['usage_count']} ÎºÎ¹Î½Î®ÏƒÎµÎ¹Ï‚ - {vehicle['total_km']:,.0f} km"
                tk.Label(vehicles_frame, text=vehicle_info, font=FONT_NORMAL,
                        bg=THEMES[self.current_theme]["frame_bg"],
                        fg=THEMES[self.current_theme]["fg"]).pack(anchor="w", padx=10, pady=2)
        
        # Monthly breakdown
        if analytics['monthly_breakdown']:
            monthly_frame = tk.LabelFrame(summary_content, text="ğŸ“… ÎœÎ·Î½Î¹Î±Î¯Î± Î£Ï„Î¿Î¹Ï‡ÎµÎ¯Î±", 
                                         font=FONT_SUBTITLE, bg=THEMES[self.current_theme]["frame_bg"],
                                         fg=THEMES[self.current_theme]["accent"])
            monthly_frame.pack(fill="x", padx=10, pady=10)
            
            # Headers
            headers_frame = tk.Frame(monthly_frame, bg=THEMES[self.current_theme]["frame_bg"])
            headers_frame.pack(fill="x", padx=10, pady=5)
            
            tk.Label(headers_frame, text="ÎœÎ®Î½Î±Ï‚", font=FONT_NORMAL, width=10,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0)
            tk.Label(headers_frame, text="ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", font=FONT_NORMAL, width=10,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=1)
            tk.Label(headers_frame, text="Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", font=FONT_NORMAL, width=12,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=2)
            tk.Label(headers_frame, text="ÎšÎ±ÏÏƒÎ¹Î¼Î± (L)", font=FONT_NORMAL, width=12,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=3)
            
            # Data rows
            for i, month_data in enumerate(analytics['monthly_breakdown'][:6]):
                row_frame = tk.Frame(monthly_frame, bg=THEMES[self.current_theme]["frame_bg"])
                row_frame.pack(fill="x", padx=10, pady=1)
                
                tk.Label(row_frame, text=month_data['month'], font=FONT_SMALL, width=10,
                        bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0)
                tk.Label(row_frame, text=str(month_data['movements']), font=FONT_SMALL, width=10,
                        bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=1)
                tk.Label(row_frame, text=f"{month_data['kilometers']:,.0f}", font=FONT_SMALL, width=12,
                        bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=2)
                tk.Label(row_frame, text=f"{month_data['fuel_liters']:.1f}", font=FONT_SMALL, width=12,
                        bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=3)
        
        # Update scroll region
        summary_content.update_idletasks()
        summary_canvas.configure(scrollregion=summary_canvas.bbox("all"))
        
        # Pack canvas and scrollbar
        summary_canvas.pack(side="left", fill="both", expand=True)
        summary_scrollbar.pack(side="right", fill="y")
        
        # Details Tab - Purpose breakdown
        self._display_purpose_breakdown(analytics['purpose_breakdown'])
    
    def _create_metric_label(self, parent, title, value, row, col):
        """Create a metric display label"""
        metric_frame = tk.Frame(parent, bg=THEMES[self.current_theme]["frame_bg"], relief="ridge", bd=1)
        metric_frame.grid(row=row, column=col, padx=5, pady=5, sticky="ew")
        parent.grid_columnconfigure(col, weight=1)
        
        tk.Label(metric_frame, text=title, font=FONT_SMALL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["text_muted"]).pack(pady=(5, 0))
        tk.Label(metric_frame, text=value, font=FONT_NORMAL,
                bg=THEMES[self.current_theme]["frame_bg"], 
                fg=THEMES[self.current_theme]["accent"]).pack(pady=(0, 5))
    
    def _display_purpose_breakdown(self, purpose_data):
        """Display purpose breakdown in details tab"""
        # Clear details frame
        for widget in self.details_frame.winfo_children():
            widget.destroy()
        
        if not purpose_data:
            tk.Label(self.details_frame, text="Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡Î¿Ï…Î½ Î´ÎµÎ´Î¿Î¼Î­Î½Î± ÏƒÎºÎ¿Ï€ÏÎ½", 
                    font=FONT_NORMAL, bg=THEMES[self.current_theme]["frame_bg"],
                    fg=THEMES[self.current_theme]["text_muted"]).pack(pady=20)
            return
        
        # Purpose breakdown
        purpose_frame = tk.LabelFrame(self.details_frame, text="ğŸ¯ Î‘Î½Î¬Î»Ï…ÏƒÎ· Î±Î½Î¬ Î£ÎºÎ¿Ï€ÏŒ", 
                                     font=FONT_SUBTITLE, bg=THEMES[self.current_theme]["frame_bg"],
                                     fg=THEMES[self.current_theme]["accent"])
        purpose_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Headers
        headers_frame = tk.Frame(purpose_frame, bg=THEMES[self.current_theme]["frame_bg"])
        headers_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(headers_frame, text="Î£ÎºÎ¿Ï€ÏŒÏ‚", font=FONT_NORMAL, width=20,
                bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0)
        tk.Label(headers_frame, text="ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", font=FONT_NORMAL, width=10,
                bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=1)
        tk.Label(headers_frame, text="Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", font=FONT_NORMAL, width=12,
                bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=2)
        tk.Label(headers_frame, text="Î Î¿ÏƒÎ¿ÏƒÏ„ÏŒ", font=FONT_NORMAL, width=10,
                bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=3)
        
        # Calculate total for percentages
        total_movements = sum(p['count'] for p in purpose_data)
        
        # Data rows
        for i, purpose in enumerate(purpose_data):
            row_frame = tk.Frame(purpose_frame, bg=THEMES[self.current_theme]["frame_bg"])
            row_frame.pack(fill="x", padx=10, pady=1)
            
            percentage = (purpose['count'] / total_movements * 100) if total_movements > 0 else 0
            
            tk.Label(row_frame, text=purpose['purpose'], font=FONT_SMALL, width=20,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=0, sticky="w")
            tk.Label(row_frame, text=str(purpose['count']), font=FONT_SMALL, width=10,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=1)
            tk.Label(row_frame, text=f"{purpose['total_km']:,.0f}", font=FONT_SMALL, width=12,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=2)
            tk.Label(row_frame, text=f"{percentage:.1f}%", font=FONT_SMALL, width=10,
                    bg=THEMES[self.current_theme]["frame_bg"], fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=3)
    
    def _show_all_drivers_summary(self):
        """Show summary for all drivers"""
        try:
            # Get date range
            start_date = self.analytics_start_date.get() if self.analytics_start_date.get() else None
            end_date = self.analytics_end_date.get() if self.analytics_end_date.get() else None
            
            # Get summary data
            summary = self.db.get_all_drivers_summary(start_date, end_date)
            if not summary:
                messagebox.showinfo("Î Î»Î·ÏÎ¿Ï†Î¿ÏÎ¯Î±", "Î”ÎµÎ½ Î²ÏÎ­Î¸Î·ÎºÎ±Î½ Î´ÎµÎ´Î¿Î¼Î­Î½Î±")
                return
            
            # Display in comparison tab
            self._display_drivers_comparison(summary)
            
        except Exception as e:
            logging.error(f"Error showing drivers summary: {e}")
            messagebox.showerror("Î£Ï†Î¬Î»Î¼Î±", f"Î£Ï†Î¬Î»Î¼Î± ÎµÎ¼Ï†Î¬Î½Î¹ÏƒÎ·Ï‚ ÏƒÏÎ½Î¿ÏˆÎ·Ï‚: {e}")
    
    def _display_drivers_comparison(self, summary):
        """Display comparison of all drivers"""
        # Clear comparison frame
        for widget in self.comparison_frame.winfo_children():
            widget.destroy()
        
        if not summary:
            tk.Label(self.comparison_frame, text="Î”ÎµÎ½ Ï…Ï€Î¬ÏÏ‡Î¿Ï…Î½ Î´ÎµÎ´Î¿Î¼Î­Î½Î±", 
                    font=FONT_NORMAL, bg=THEMES[self.current_theme]["frame_bg"],
                    fg=THEMES[self.current_theme]["text_muted"]).pack(pady=20)
            return
        
        # Create scrollable frame
        canvas = tk.Canvas(self.comparison_frame, bg=THEMES[self.current_theme]["frame_bg"])
        scrollbar = ttk.Scrollbar(self.comparison_frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        content_frame = tk.Frame(canvas, bg=THEMES[self.current_theme]["frame_bg"])
        canvas.create_window((0, 0), window=content_frame, anchor="nw")
        
        # Title
        title_label = tk.Label(content_frame, text="ğŸ‘¥ Î£ÏÎ³ÎºÏÎ¹ÏƒÎ· ÎŒÎ»Ï‰Î½ Ï„Ï‰Î½ ÎŸÎ´Î·Î³ÏÎ½", 
                              font=FONT_SUBTITLE, bg=THEMES[self.current_theme]["frame_bg"],
                              fg=THEMES[self.current_theme]["accent"])
        title_label.pack(pady=10)
        
        # Headers
        headers_frame = tk.Frame(content_frame, bg=THEMES[self.current_theme]["frame_bg"])
        headers_frame.pack(fill="x", padx=10, pady=5)
        
        headers = ["ÎŒÎ½Î¿Î¼Î±", "ÎšÎ¹Î½Î®ÏƒÎµÎ¹Ï‚", "Î§Î¹Î»Î¹ÏŒÎ¼ÎµÏ„ÏÎ±", "ÎšÎ±ÏÏƒÎ¹Î¼Î± (L)", "ÎšÏŒÏƒÏ„Î¿Ï‚ (â‚¬)", "ÎšÎ±Ï„Î±Î½Î¬Î»Ï‰ÏƒÎ·"]
        widths = [20, 10, 12, 12, 12, 12]
        
        for i, (header, width) in enumerate(zip(headers, widths)):
            tk.Label(headers_frame, text=header, font=FONT_NORMAL, width=width,
                    bg=THEMES[self.current_theme]["frame_bg"], 
                    fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=i, padx=2)
        
        # Data rows
        for i, driver in enumerate(summary):
            row_frame = tk.Frame(content_frame, bg=THEMES[self.current_theme]["frame_bg"])
            row_frame.pack(fill="x", padx=10, pady=1)
            
            values = [
                driver['driver_name'],
                str(driver['total_movements']),
                f"{driver['total_kilometers']:,.0f}",
                f"{driver['total_fuel_liters']:.1f}",
                f"{driver['total_fuel_cost']:.2f}",
                f"{driver['avg_consumption']:.2f}"
            ]
            
            for j, (value, width) in enumerate(zip(values, widths)):
                tk.Label(row_frame, text=value, font=FONT_SMALL, width=width,
                        bg=THEMES[self.current_theme]["frame_bg"], 
                        fg=THEMES[self.current_theme]["fg"]).grid(row=0, column=j, padx=2, sticky="w")
        
        # Update scroll region
        content_frame.update_idletasks()
        canvas.configure(scrollregion=canvas.bbox("all"))
        
        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def _on_close(self):
        """Handle application close with cleanup"""
        try:
            # Ask for confirmation
            if ConfirmDialog(
                self.root,
                "ğŸšª ÎˆÎ¾Î¿Î´Î¿Ï‚ Î±Ï€ÏŒ Ï„Î·Î½ Î•Ï†Î±ÏÎ¼Î¿Î³Î®",
                "Î•Î¯ÏƒÏ„Îµ ÏƒÎ¯Î³Î¿Ï…ÏÎ¿Ï‚ ÏŒÏ„Î¹ Î¸Î­Î»ÎµÏ„Îµ Î½Î± ÎºÎ»ÎµÎ¯ÏƒÎµÏ„Îµ Ï„Î·Î½ ÎµÏ†Î±ÏÎ¼Î¿Î³Î®;"
            ).show():
                
                log_user_action("Application closing")
                
                # Close database connection
                if self.db:
                    self.db.close()
                
                # Destroy window
                self.root.destroy()
                
        except Exception as e:
            logging.error(f"Error during application close: {e}")
            self.root.destroy()

    def _update_font_sizes(self, screen_width, screen_height):
        """Update font sizes based on screen dimensions"""
        from config import get_adaptive_font_sizes
        
        # Get adaptive font sizes
        font_sizes = get_adaptive_font_sizes(screen_width, screen_height)
        
        # Update global font variables
        global FONT_BIG, FONT_NORMAL, FONT_SMALL, FONT_TITLE, FONT_SUBTITLE
        FONT_BIG = font_sizes['FONT_BIG']
        FONT_NORMAL = font_sizes['FONT_NORMAL']
        FONT_SMALL = font_sizes['FONT_SMALL']
        FONT_TITLE = font_sizes['FONT_TITLE']
        FONT_SUBTITLE = font_sizes['FONT_SUBTITLE']
        
        # Update button styles with new font sizes
        for style in BUTTON_STYLES.values():
            style['font'] = FONT_NORMAL
        
        logging.info(f"Font sizes updated for {screen_width}x{screen_height} screen")

    def _get_adaptive_column_widths(self, base_widths):
        """Get adaptive column widths based on window size"""
        if not hasattr(self, 'window_width'):
            return base_widths
        
        # Scale factor based on window width
        if self.window_width < 1200:
            scale_factor = 0.8
        elif self.window_width < 1600:
            scale_factor = 0.9
        else:
            scale_factor = 1.0
        
        return {col: int(width * scale_factor) for col, width in base_widths.items()}

def main():
    """Main application entry point"""
    try:
        # Create root window
        root = tk.Tk()
        
        # Initialize application
        app = FleetManagerImproved(root)
        
        # Start main loop
        root.mainloop()
        
    except Exception as e:
        logging.critical(f"Critical error in main: {e}")
        if 'root' in locals():
            messagebox.showerror("âŒ ÎšÏÎ¯ÏƒÎ¹Î¼Î¿ Î£Ï†Î¬Î»Î¼Î±", 
                               f"ÎšÏÎ¯ÏƒÎ¹Î¼Î¿ ÏƒÏ†Î¬Î»Î¼Î± ÎµÏ†Î±ÏÎ¼Î¿Î³Î®Ï‚:\n{str(e)}")
        print(f"Critical error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()